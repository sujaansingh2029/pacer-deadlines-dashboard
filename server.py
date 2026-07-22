from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import math
import os
import re
import statistics
import tempfile
import urllib.parse
from email.utils import parsedate_to_datetime
from dataclasses import dataclass
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Iterable

import requests
from bs4 import BeautifulSoup

ROOT = Path(__file__).parent.resolve()
OUTPUT_DIR = ROOT / "outputs"
UPLOAD_DIR = ROOT / "work" / "uploads"
OUTPUT_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

MAX_TEXT_CHARS = 24000
USER_AGENT = "DD-Brief-Generator/0.3 Advanced Local Preview"
SEC_USER_AGENT = os.getenv("SEC_USER_AGENT", "DD-Brief-Generator/0.4 local diligence research contact@example.com")
CHAT_CONTEXT_LIMIT = 42000
CHAT_STATE: dict[str, object] = {
    "company": "",
    "brief": "",
    "sources": [],
    "figures": [],
}
SCREENER_STATE: dict[str, object] = {
    "country": "",
    "rows": [],
    "market_rows": [],
}
RESEARCH_STATE: dict[str, object] = {
    "rows": [],
    "assumptions": {},
    "parse_note": "",
    "missing": [],
}

COUNTRY_MARKET_UNIVERSES: dict[str, list[str]] = {
    "United States": ["PG", "KO", "PEP", "WMT", "COST", "PM", "MO", "MDLZ", "CL", "KMB", "EL", "TGT"],
    "Japan": ["2914.T", "2502.T", "2503.T", "2802.T", "4452.T", "4911.T", "3382.T", "8267.T", "2269.T"],
    "United Kingdom": ["ULVR.L", "DGE.L", "TSCO.L", "SBRY.L", "BATS.L", "IMB.L", "ABF.L", "CPG.L"],
    "Canada": ["L.TO", "MRU.TO", "EMP-A.TO", "ATD.TO", "WN.TO", "SAP.TO", "PBH.TO"],
    "Australia": ["WOW.AX", "COL.AX", "WES.AX", "TWE.AX", "A2M.AX", "EDV.AX"],
    "Germany": ["BEI.DE", "HEN3.DE", "HEN.DE", "BOSS.DE", "ZAL.DE"],
    "France": ["OR.PA", "RI.PA", "BN.PA", "MC.PA", "RMS.PA", "CA.PA"],
    "India": ["HINDUNILVR.NS", "ITC.NS", "NESTLEIND.NS", "BRITANNIA.NS", "DABUR.NS", "MARICO.NS", "DMART.NS"],
    "South Korea": ["005380.KS", "051900.KS", "090430.KS", "004370.KS", "097950.KS"],
    "Hong Kong": ["0291.HK", "0322.HK", "0688.HK", "1044.HK", "1929.HK", "2319.HK"],
    "China": ["600519.SS", "000858.SZ", "600887.SS", "603288.SS", "000333.SZ", "600690.SS", "601888.SS", "002304.SZ"],
}

COUNTRY_MARKET_PROFILES: dict[str, dict[str, str]] = {
    "PG": {"name": "Procter & Gamble", "exchange": "NYSE", "currency": "USD", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Household care, fabric care, grooming, baby care, feminine care, health care."},
    "KO": {"name": "Coca-Cola", "exchange": "NYSE", "currency": "USD", "sector": "Consumer Defensive", "industry": "Beverages", "products": "Sparkling drinks, water, coffee, tea, juices, sports drinks."},
    "PEP": {"name": "PepsiCo", "exchange": "NASDAQ", "currency": "USD", "sector": "Consumer Defensive", "industry": "Beverages and snacks", "products": "Beverages, snacks, cereals, foods under Pepsi, Frito-Lay, Quaker and other brands."},
    "WMT": {"name": "Walmart", "exchange": "NYSE", "currency": "USD", "sector": "Consumer Defensive", "industry": "Discount retail", "products": "Grocery, general merchandise, pharmacy, e-commerce, membership retail."},
    "COST": {"name": "Costco Wholesale", "exchange": "NASDAQ", "currency": "USD", "sector": "Consumer Defensive", "industry": "Membership warehouse retail", "products": "Warehouse retail, grocery, fuel, pharmacy, private-label Kirkland products."},
    "PM": {"name": "Philip Morris International", "exchange": "NYSE", "currency": "USD", "sector": "Consumer Defensive", "industry": "Tobacco", "products": "Tobacco, heated tobacco, nicotine products, smoke-free products."},
    "MO": {"name": "Altria Group", "exchange": "NYSE", "currency": "USD", "sector": "Consumer Defensive", "industry": "Tobacco", "products": "Cigarettes, smokeless tobacco, oral nicotine, wine interests."},
    "MDLZ": {"name": "Mondelez International", "exchange": "NASDAQ", "currency": "USD", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Biscuits, chocolate, gum, candy, snacks."},
    "CL": {"name": "Colgate-Palmolive", "exchange": "NYSE", "currency": "USD", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Oral care, personal care, home care, pet nutrition."},
    "KMB": {"name": "Kimberly-Clark", "exchange": "NYSE", "currency": "USD", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Tissue, diapers, wipes, hygiene and professional products."},
    "EL": {"name": "Estee Lauder", "exchange": "NYSE", "currency": "USD", "sector": "Consumer Defensive", "industry": "Beauty", "products": "Skin care, makeup, fragrance, hair care."},
    "TGT": {"name": "Target", "exchange": "NYSE", "currency": "USD", "sector": "Consumer Defensive", "industry": "Discount retail", "products": "General merchandise, apparel, grocery, home goods, e-commerce."},
    "2914.T": {"name": "Japan Tobacco", "exchange": "Tokyo Stock Exchange", "currency": "JPY", "sector": "Consumer Defensive", "industry": "Tobacco", "products": "Tobacco, heated tobacco, pharmaceuticals, processed foods."},
    "2502.T": {"name": "Asahi Group Holdings", "exchange": "Tokyo Stock Exchange", "currency": "JPY", "sector": "Consumer Defensive", "industry": "Beverages", "products": "Beer, spirits, soft drinks, food products."},
    "2503.T": {"name": "Kirin Holdings", "exchange": "Tokyo Stock Exchange", "currency": "JPY", "sector": "Consumer Defensive", "industry": "Beverages", "products": "Beer, beverages, health science, pharmaceuticals."},
    "2802.T": {"name": "Ajinomoto", "exchange": "Tokyo Stock Exchange", "currency": "JPY", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Seasonings, frozen foods, amino acids, specialty ingredients."},
    "4452.T": {"name": "Kao", "exchange": "Tokyo Stock Exchange", "currency": "JPY", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Beauty care, human health care, fabric and home care, chemicals."},
    "4911.T": {"name": "Shiseido", "exchange": "Tokyo Stock Exchange", "currency": "JPY", "sector": "Consumer Defensive", "industry": "Beauty", "products": "Cosmetics, skin care, fragrance, personal care."},
    "3382.T": {"name": "Seven & i Holdings", "exchange": "Tokyo Stock Exchange", "currency": "JPY", "sector": "Consumer Defensive", "industry": "Convenience and grocery retail", "products": "Convenience stores, supermarkets, specialty retail."},
    "8267.T": {"name": "Aeon", "exchange": "Tokyo Stock Exchange", "currency": "JPY", "sector": "Consumer Defensive", "industry": "Retail", "products": "Supermarkets, malls, convenience stores, financial services."},
    "2269.T": {"name": "Meiji Holdings", "exchange": "Tokyo Stock Exchange", "currency": "JPY", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Dairy, confectionery, nutrition, pharmaceuticals."},
    "ULVR.L": {"name": "Unilever", "exchange": "London Stock Exchange", "currency": "GBp", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Beauty, wellbeing, personal care, home care, nutrition, ice cream."},
    "DGE.L": {"name": "Diageo", "exchange": "London Stock Exchange", "currency": "GBp", "sector": "Consumer Defensive", "industry": "Alcoholic beverages", "products": "Spirits, beer, ready-to-drink beverages."},
    "TSCO.L": {"name": "Tesco", "exchange": "London Stock Exchange", "currency": "GBp", "sector": "Consumer Defensive", "industry": "Grocery retail", "products": "Grocery, general merchandise, fuel, financial services."},
    "SBRY.L": {"name": "J Sainsbury", "exchange": "London Stock Exchange", "currency": "GBp", "sector": "Consumer Defensive", "industry": "Grocery retail", "products": "Grocery, general merchandise, convenience retail, financial services."},
    "BATS.L": {"name": "British American Tobacco", "exchange": "London Stock Exchange", "currency": "GBp", "sector": "Consumer Defensive", "industry": "Tobacco", "products": "Cigarettes, vapor, heated tobacco, oral nicotine."},
    "IMB.L": {"name": "Imperial Brands", "exchange": "London Stock Exchange", "currency": "GBp", "sector": "Consumer Defensive", "industry": "Tobacco", "products": "Tobacco, papers, vapor and heated tobacco products."},
    "ABF.L": {"name": "Associated British Foods", "exchange": "London Stock Exchange", "currency": "GBp", "sector": "Consumer Defensive", "industry": "Food and retail", "products": "Grocery, sugar, ingredients, agriculture, Primark apparel retail."},
    "CPG.L": {"name": "Compass Group", "exchange": "London Stock Exchange", "currency": "GBp", "sector": "Consumer Cyclical", "industry": "Food services", "products": "Contract catering, food services, support services."},
    "L.TO": {"name": "Loblaw Companies", "exchange": "Toronto Stock Exchange", "currency": "CAD", "sector": "Consumer Defensive", "industry": "Grocery and pharmacy retail", "products": "Grocery, pharmacy, apparel, financial services."},
    "MRU.TO": {"name": "Metro", "exchange": "Toronto Stock Exchange", "currency": "CAD", "sector": "Consumer Defensive", "industry": "Grocery and pharmacy retail", "products": "Food stores, drugstores, private-label products."},
    "EMP-A.TO": {"name": "Empire Company", "exchange": "Toronto Stock Exchange", "currency": "CAD", "sector": "Consumer Defensive", "industry": "Grocery retail", "products": "Sobeys grocery stores, pharmacy, fuel, e-commerce grocery."},
    "ATD.TO": {"name": "Alimentation Couche-Tard", "exchange": "Toronto Stock Exchange", "currency": "CAD", "sector": "Consumer Defensive", "industry": "Convenience retail", "products": "Convenience stores, fuel retail, foodservice."},
    "WN.TO": {"name": "George Weston", "exchange": "Toronto Stock Exchange", "currency": "CAD", "sector": "Consumer Defensive", "industry": "Food and retail holding company", "products": "Bakery, grocery, pharmacy interests through Loblaw and Weston Foods legacy assets."},
    "SAP.TO": {"name": "Saputo", "exchange": "Toronto Stock Exchange", "currency": "CAD", "sector": "Consumer Defensive", "industry": "Dairy foods", "products": "Cheese, milk, cream, cultured products, dairy ingredients."},
    "PBH.TO": {"name": "Premium Brands Holdings", "exchange": "Toronto Stock Exchange", "currency": "CAD", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Specialty foods, sandwiches, meats, seafood, distribution."},
    "WOW.AX": {"name": "Woolworths Group", "exchange": "Australian Securities Exchange", "currency": "AUD", "sector": "Consumer Defensive", "industry": "Grocery retail", "products": "Supermarkets, grocery, liquor, everyday retail and e-commerce."},
    "COL.AX": {"name": "Coles Group", "exchange": "Australian Securities Exchange", "currency": "AUD", "sector": "Consumer Defensive", "industry": "Grocery retail", "products": "Supermarkets, liquor retail, convenience and private-label groceries."},
    "WES.AX": {"name": "Wesfarmers", "exchange": "Australian Securities Exchange", "currency": "AUD", "sector": "Consumer Cyclical", "industry": "Retail conglomerate", "products": "Home improvement, office supplies, department stores, chemicals and industrial products."},
    "TWE.AX": {"name": "Treasury Wine Estates", "exchange": "Australian Securities Exchange", "currency": "AUD", "sector": "Consumer Defensive", "industry": "Wine", "products": "Premium, luxury and commercial wine brands."},
    "A2M.AX": {"name": "The a2 Milk Company", "exchange": "Australian Securities Exchange", "currency": "AUD", "sector": "Consumer Defensive", "industry": "Dairy foods", "products": "Infant formula, liquid milk, powdered milk and dairy nutrition."},
    "EDV.AX": {"name": "Endeavour Group", "exchange": "Australian Securities Exchange", "currency": "AUD", "sector": "Consumer Defensive", "industry": "Liquor retail and hospitality", "products": "Liquor retail, hotels, pubs and hospitality operations."},
    "BEI.DE": {"name": "Beiersdorf", "exchange": "Xetra", "currency": "EUR", "sector": "Consumer Defensive", "industry": "Personal products", "products": "Skin care, personal care and adhesive products including Nivea and Eucerin."},
    "HEN3.DE": {"name": "Henkel Preference Shares", "exchange": "Xetra", "currency": "EUR", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Laundry, home care, beauty care and adhesive technologies."},
    "HEN.DE": {"name": "Henkel Ordinary Shares", "exchange": "Xetra", "currency": "EUR", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Laundry, home care, beauty care and adhesive technologies."},
    "BOSS.DE": {"name": "Hugo Boss", "exchange": "Xetra", "currency": "EUR", "sector": "Consumer Cyclical", "industry": "Apparel", "products": "Premium apparel, accessories, footwear and fragrances."},
    "ZAL.DE": {"name": "Zalando", "exchange": "Xetra", "currency": "EUR", "sector": "Consumer Cyclical", "industry": "Online fashion retail", "products": "Online fashion marketplace, apparel, footwear, beauty and logistics services."},
    "OR.PA": {"name": "L'Oreal", "exchange": "Euronext Paris", "currency": "EUR", "sector": "Consumer Defensive", "industry": "Beauty", "products": "Cosmetics, skin care, hair care, fragrance and beauty products."},
    "RI.PA": {"name": "Pernod Ricard", "exchange": "Euronext Paris", "currency": "EUR", "sector": "Consumer Defensive", "industry": "Alcoholic beverages", "products": "Spirits, wines and champagne brands."},
    "BN.PA": {"name": "Danone", "exchange": "Euronext Paris", "currency": "EUR", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Dairy, plant-based foods, waters, infant nutrition and medical nutrition."},
    "MC.PA": {"name": "LVMH", "exchange": "Euronext Paris", "currency": "EUR", "sector": "Consumer Cyclical", "industry": "Luxury goods", "products": "Fashion, leather goods, wines, spirits, perfumes, cosmetics, watches and jewelry."},
    "RMS.PA": {"name": "Hermes International", "exchange": "Euronext Paris", "currency": "EUR", "sector": "Consumer Cyclical", "industry": "Luxury goods", "products": "Leather goods, fashion, watches, jewelry, perfume and home goods."},
    "CA.PA": {"name": "Carrefour", "exchange": "Euronext Paris", "currency": "EUR", "sector": "Consumer Defensive", "industry": "Grocery retail", "products": "Hypermarkets, supermarkets, convenience stores, grocery e-commerce."},
    "HINDUNILVR.NS": {"name": "Hindustan Unilever", "exchange": "National Stock Exchange of India", "currency": "INR", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Home care, beauty, personal care, foods and refreshments."},
    "ITC.NS": {"name": "ITC", "exchange": "National Stock Exchange of India", "currency": "INR", "sector": "Consumer Defensive", "industry": "Tobacco and packaged goods", "products": "Cigarettes, packaged foods, personal care, hotels, paperboards and agribusiness."},
    "NESTLEIND.NS": {"name": "Nestle India", "exchange": "National Stock Exchange of India", "currency": "INR", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Dairy, nutrition, coffee, beverages, prepared dishes and confectionery."},
    "BRITANNIA.NS": {"name": "Britannia Industries", "exchange": "National Stock Exchange of India", "currency": "INR", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Biscuits, bakery, dairy and snacking products."},
    "DABUR.NS": {"name": "Dabur India", "exchange": "National Stock Exchange of India", "currency": "INR", "sector": "Consumer Defensive", "industry": "Personal products", "products": "Ayurvedic health care, personal care, foods and beverages."},
    "MARICO.NS": {"name": "Marico", "exchange": "National Stock Exchange of India", "currency": "INR", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Hair care, edible oils, foods, male grooming and skin care."},
    "DMART.NS": {"name": "Avenue Supermarts", "exchange": "National Stock Exchange of India", "currency": "INR", "sector": "Consumer Defensive", "industry": "Grocery retail", "products": "DMart supermarkets, grocery, household goods, apparel and general merchandise."},
    "005380.KS": {"name": "Hyundai Motor", "exchange": "Korea Exchange", "currency": "KRW", "sector": "Consumer Cyclical", "industry": "Automobiles", "products": "Passenger vehicles, SUVs, commercial vehicles, EVs, parts and mobility services."},
    "051900.KS": {"name": "LG H&H", "exchange": "Korea Exchange", "currency": "KRW", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Cosmetics, personal care, household products and beverages."},
    "090430.KS": {"name": "Amorepacific", "exchange": "Korea Exchange", "currency": "KRW", "sector": "Consumer Defensive", "industry": "Beauty", "products": "Cosmetics, skin care, fragrance and personal care products."},
    "004370.KS": {"name": "Nongshim", "exchange": "Korea Exchange", "currency": "KRW", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Instant noodles, snacks, beverages and packaged food products."},
    "097950.KS": {"name": "CJ CheilJedang", "exchange": "Korea Exchange", "currency": "KRW", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Food, bio ingredients, nutrition, frozen foods and processed foods."},
    "0291.HK": {"name": "China Resources Beer", "exchange": "Hong Kong Stock Exchange", "currency": "HKD", "sector": "Consumer Defensive", "industry": "Beverages", "products": "Beer brewing, distribution and beverage brands including Snow."},
    "0322.HK": {"name": "Tingyi Cayman Islands Holding", "exchange": "Hong Kong Stock Exchange", "currency": "HKD", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Instant noodles, beverages and packaged food products."},
    "0688.HK": {"name": "China Overseas Land & Investment", "exchange": "Hong Kong Stock Exchange", "currency": "HKD", "sector": "Real Estate", "industry": "Property development", "products": "Residential and commercial property development, investment properties."},
    "1044.HK": {"name": "Hengan International", "exchange": "Hong Kong Stock Exchange", "currency": "HKD", "sector": "Consumer Defensive", "industry": "Household & personal products", "products": "Tissue paper, sanitary napkins, diapers and personal hygiene products."},
    "1929.HK": {"name": "Chow Tai Fook Jewellery", "exchange": "Hong Kong Stock Exchange", "currency": "HKD", "sector": "Consumer Cyclical", "industry": "Luxury retail", "products": "Jewelry, watches, gold products, gems and luxury accessories."},
    "2319.HK": {"name": "China Mengniu Dairy", "exchange": "Hong Kong Stock Exchange", "currency": "HKD", "sector": "Consumer Defensive", "industry": "Dairy foods", "products": "Liquid milk, yogurt, ice cream, milk formula and dairy beverages."},
    "600519.SS": {"name": "Kweichow Moutai", "exchange": "Shanghai Stock Exchange", "currency": "CNY", "sector": "Consumer Defensive", "industry": "Distillers and wineries", "products": "Premium baijiu liquor under the Moutai brand."},
    "000858.SZ": {"name": "Wuliangye Yibin", "exchange": "Shenzhen Stock Exchange", "currency": "CNY", "sector": "Consumer Defensive", "industry": "Distillers and wineries", "products": "Baijiu liquor and related alcoholic beverage products."},
    "600887.SS": {"name": "Inner Mongolia Yili Industrial Group", "exchange": "Shanghai Stock Exchange", "currency": "CNY", "sector": "Consumer Defensive", "industry": "Dairy foods", "products": "Liquid milk, yogurt, milk powder, ice cream, cheese and health drinks."},
    "603288.SS": {"name": "Foshan Haitian Flavouring and Food", "exchange": "Shanghai Stock Exchange", "currency": "CNY", "sector": "Consumer Defensive", "industry": "Packaged foods", "products": "Soy sauce, oyster sauce, condiments, seasonings and cooking sauces."},
    "000333.SZ": {"name": "Midea Group", "exchange": "Shenzhen Stock Exchange", "currency": "CNY", "sector": "Consumer Cyclical", "industry": "Home appliances", "products": "Air conditioners, refrigerators, laundry appliances, kitchen appliances, robotics and automation."},
    "600690.SS": {"name": "Haier Smart Home", "exchange": "Shanghai Stock Exchange", "currency": "CNY", "sector": "Consumer Cyclical", "industry": "Home appliances", "products": "Refrigerators, washing machines, air conditioners, kitchen appliances and smart home systems."},
    "601888.SS": {"name": "China Tourism Group Duty Free", "exchange": "Shanghai Stock Exchange", "currency": "CNY", "sector": "Consumer Cyclical", "industry": "Travel retail", "products": "Duty-free retail, travel retail, cosmetics, luxury and consumer goods."},
    "002304.SZ": {"name": "Jiangsu Yanghe Brewery", "exchange": "Shenzhen Stock Exchange", "currency": "CNY", "sector": "Consumer Defensive", "industry": "Distillers and wineries", "products": "Baijiu spirits under Yanghe, Shuanggou and related brands."},
}

WESTERN_BENCHMARKS = ["SPY", "QQQ", "VGK"]

COUNTRY_MARKET_BENCHMARKS: dict[str, list[tuple[str, str]]] = {
    "United States": [("^GSPC", "S&P 500"), ("^IXIC", "Nasdaq Composite"), ("^DJI", "Dow Jones Industrial Average")],
    "China": [("000001.SS", "Shanghai Composite"), ("399001.SZ", "Shenzhen Component"), ("000300.SS", "CSI 300")],
    "Japan": [("^N225", "Nikkei 225"), ("^TOPX", "TOPIX")],
    "United Kingdom": [("^FTSE", "FTSE 100")],
    "Canada": [("^GSPTSE", "S&P/TSX Composite")],
    "Australia": [("^AXJO", "S&P/ASX 200")],
    "Germany": [("^GDAXI", "DAX")],
    "France": [("^FCHI", "CAC 40")],
    "India": [("^NSEI", "Nifty 50"), ("^BSESN", "BSE Sensex")],
    "South Korea": [("^KS11", "KOSPI Composite")],
    "Hong Kong": [("^HSI", "Hang Seng Index")],
}

FALLBACK_MARKET_BASELINES: dict[str, dict[str, float | str]] = {
    "600519.SS": {"price": 1308.0, "market_cap": 1_640_000_000_000.0, "avg_volume": 4_000_000.0, "pe": 21.0, "forward_pe": 19.0, "pb": 7.0, "dividend_yield": 2.3, "beta": 0.7, "return_1y": -10.0, "return_3y": -30.0, "volatility": 28.0, "drawdown": -50.0},
    "000858.SZ": {"price": 126.0, "market_cap": 490_000_000_000.0, "avg_volume": 19_000_000.0, "pe": 16.0, "forward_pe": 15.0, "pb": 3.0, "dividend_yield": 3.0, "beta": 0.9, "return_1y": -5.0, "return_3y": -35.0, "volatility": 32.0, "drawdown": -55.0},
    "600887.SS": {"price": 27.0, "market_cap": 170_000_000_000.0, "avg_volume": 25_000_000.0, "pe": 16.0, "forward_pe": 15.0, "pb": 2.8, "dividend_yield": 4.0, "beta": 0.8, "return_1y": 3.0, "return_3y": -25.0, "volatility": 24.0, "drawdown": -45.0},
    "603288.SS": {"price": 40.0, "market_cap": 220_000_000_000.0, "avg_volume": 12_000_000.0, "pe": 36.0, "forward_pe": 31.0, "pb": 6.8, "dividend_yield": 1.9, "beta": 0.8, "return_1y": 7.0, "return_3y": -45.0, "volatility": 30.0, "drawdown": -65.0},
    "000333.SZ": {"price": 75.0, "market_cap": 525_000_000_000.0, "avg_volume": 28_000_000.0, "pe": 14.0, "forward_pe": 13.0, "pb": 2.8, "dividend_yield": 4.1, "beta": 0.9, "return_1y": 20.0, "return_3y": 5.0, "volatility": 26.0, "drawdown": -35.0},
    "600690.SS": {"price": 24.0, "market_cap": 225_000_000_000.0, "avg_volume": 24_000_000.0, "pe": 12.0, "forward_pe": 11.0, "pb": 2.0, "dividend_yield": 4.3, "beta": 0.8, "return_1y": -2.0, "return_3y": -20.0, "volatility": 25.0, "drawdown": -42.0},
    "601888.SS": {"price": 62.0, "market_cap": 130_000_000_000.0, "avg_volume": 17_000_000.0, "pe": 20.0, "forward_pe": 18.0, "pb": 1.8, "dividend_yield": 2.5, "beta": 1.2, "return_1y": -25.0, "return_3y": -70.0, "volatility": 40.0, "drawdown": -80.0},
    "002304.SZ": {"price": 75.0, "market_cap": 115_000_000_000.0, "avg_volume": 10_000_000.0, "pe": 16.0, "forward_pe": 15.0, "pb": 2.5, "dividend_yield": 5.0, "beta": 0.8, "return_1y": -12.0, "return_3y": -45.0, "volatility": 30.0, "drawdown": -60.0},
    "000001.SS": {"price": 3300.0, "return_1y": 10.0, "return_3y": -2.0, "return_period": -2.0, "volatility": 18.0, "drawdown": -25.0},
    "399001.SZ": {"price": 10500.0, "return_1y": 12.0, "return_3y": -12.0, "return_period": -12.0, "volatility": 24.0, "drawdown": -35.0},
    "000300.SS": {"price": 3900.0, "return_1y": 13.0, "return_3y": -8.0, "return_period": -8.0, "volatility": 20.0, "drawdown": -30.0},
    "SPY": {"price": 620.0, "return_1y": 14.0, "return_3y": 50.0, "return_period": 50.0, "volatility": 16.0, "drawdown": -12.0},
    "QQQ": {"price": 550.0, "return_1y": 18.0, "return_3y": 70.0, "return_period": 70.0, "volatility": 22.0, "drawdown": -18.0},
    "VGK": {"price": 75.0, "return_1y": 12.0, "return_3y": 45.0, "return_period": 45.0, "volatility": 18.0, "drawdown": -16.0},
}


@dataclass
class Source:
    label: str
    url: str
    text: str


@dataclass
class SECCompany:
    cik: str
    ticker: str
    title: str


@dataclass
class Figure:
    value: str
    context: str
    source_label: str
    source_url: str
    confidence: str
    category: str
    quality_note: str


@dataclass
class StockPoint:
    date: datetime
    close: float


@dataclass
class StockSeries:
    ticker: str
    currency: str
    points: list[StockPoint]
    source_url: str
    name: str = ""
    exchange: str = ""
    quote_type: str = ""
    market_region: str = ""
    regular_market_price: float | None = None
    previous_close: float | None = None
    day_low: float | None = None
    day_high: float | None = None
    fifty_two_week_low: float | None = None
    fifty_two_week_high: float | None = None
    market_cap: float | None = None
    trailing_pe: float | None = None
    forward_pe: float | None = None
    price_to_book: float | None = None
    dividend_yield_pct: float | None = None
    beta: float | None = None
    average_volume: float | None = None
    shares_outstanding: float | None = None
    sector: str = ""
    industry: str = ""
    country: str = ""
    website: str = ""
    employees: int | None = None
    business_summary: str = ""


@dataclass
class MarketNews:
    title: str
    publisher: str
    link: str
    published: datetime | None
    summary: str = ""


@dataclass
class MoveCatalyst:
    label: str
    date: datetime
    move_pct: float
    explanation: str
    headlines: list[MarketNews]


@dataclass
class StockMetrics:
    ticker: str
    currency: str
    latest_price: float
    start_price: float
    high_price: float
    low_price: float
    total_return_pct: float
    cagr_pct: float
    volatility_pct: float
    sharpe_like: float
    max_drawdown_pct: float
    current_drawdown_pct: float
    best_week_pct: float
    worst_week_pct: float
    positive_week_pct: float
    weeks: int
    trend_label: str
    ma_40: float
    annual_returns: dict[int, float]


@dataclass
class GrowthScenario:
    label: str
    annual_rate_pct: float
    projected_prices: dict[int, float]
    projected_returns: dict[int, float]


@dataclass
class MarketAnalysis:
    target: StockSeries
    target_metrics: StockMetrics
    benchmark: StockSeries | None
    peers: list[StockSeries]
    peer_metrics: list[StockMetrics]
    market_news: list[MarketNews]
    move_catalysts: list[MoveCatalyst]
    symbol_note: str
    international_note: str
    disclosure_links: list[tuple[str, str]]
    price_chart_path: str
    indexed_chart_path: str
    drawdown_chart_path: str
    annual_chart_path: str
    forecast_chart_path: str
    scenarios: list[GrowthScenario]
    source_note: str


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def md_cell(value: object, limit: int = 420) -> str:
    text = clean_text(str(value if value is not None else ""))
    if len(text) > limit:
        text = text[: limit - 1] + "…"
    return text.replace("|", " / ")


def html_page(body: str) -> bytes:
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Advanced DD Brief Generator</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body>{body}</body>
</html>""".encode("utf-8")


def fetch_website(company: str, website: str) -> Source:
    if not website:
        return Source("Company website", "", "")
    url = website.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    try:
        response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=14, allow_redirects=True)
        response.raise_for_status()
    except Exception as exc:
        return Source("Company website", url, f"Website fetch failed: {exc}")

    soup = BeautifulSoup(response.text, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    title = clean_text(soup.title.string if soup.title else company)
    meta = ""
    meta_tag = soup.find("meta", attrs={"name": "description"})
    if meta_tag and meta_tag.get("content"):
        meta = clean_text(meta_tag["content"])
    text = clean_text(soup.get_text(" "))
    combined = clean_text(f"{title}. {meta}. {text}")[:MAX_TEXT_CHARS]
    return Source("Company website", response.url, combined)


def fetch_sec_company_search(company: str) -> Source:
    query = urllib.parse.quote(company)
    url = f"https://www.sec.gov/edgar/search/#/q={query}"
    return Source(
        "SEC EDGAR search",
        url,
        f"Use the SEC EDGAR company search for public filings related to {company}. Private companies often have no result.",
    )


def sec_headers() -> dict[str, str]:
    return {
        "User-Agent": SEC_USER_AGENT,
        "Accept-Encoding": "gzip, deflate",
        "Accept": "application/json,text/html,application/xhtml+xml",
    }


def normalize_company_name(value: str) -> str:
    lowered = re.sub(r"[^a-z0-9 ]+", " ", (value or "").lower())
    lowered = re.sub(
        r"\b(incorporated|inc|corp|corporation|company|co|ltd|limited|plc|class a|class b|common stock|the)\b",
        " ",
        lowered,
    )
    return clean_text(lowered)


def fetch_json(url: str) -> object | None:
    try:
        response = requests.get(url, headers=sec_headers(), timeout=14)
        response.raise_for_status()
        return response.json()
    except Exception:
        return None


def lookup_sec_company(company: str, ticker: str = "") -> SECCompany | None:
    data = fetch_json("https://www.sec.gov/files/company_tickers.json")
    if not isinstance(data, dict):
        return None

    rows = []
    for item in data.values():
        if isinstance(item, dict):
            rows.append(item)

    ticker_clean = clean_text(ticker).upper().replace(" ", "")
    if ticker_clean:
        for row in rows:
            if clean_text(str(row.get("ticker", ""))).upper() == ticker_clean:
                cik = str(row.get("cik_str", "")).zfill(10)
                return SECCompany(cik=cik, ticker=ticker_clean, title=clean_text(str(row.get("title", ""))))

    target = normalize_company_name(company)
    if not target:
        return None
    best: tuple[int, dict[str, object]] | None = None
    target_words = set(target.split())
    for row in rows:
        title = clean_text(str(row.get("title", "")))
        normalized = normalize_company_name(title)
        if not normalized:
            continue
        score = 0
        if normalized == target:
            score = 100
        elif normalized.startswith(target) or target.startswith(normalized):
            score = 80
        else:
            overlap = len(target_words & set(normalized.split()))
            score = overlap * 20
        if score and (best is None or score > best[0]):
            best = (score, row)
    if not best or best[0] < 40:
        return None
    row = best[1]
    return SECCompany(
        cik=str(row.get("cik_str", "")).zfill(10),
        ticker=clean_text(str(row.get("ticker", ""))).upper(),
        title=clean_text(str(row.get("title", ""))),
    )


def filing_url(cik: str, accession: str, primary_document: str) -> str:
    accession_clean = accession.replace("-", "")
    cik_int = str(int(cik))
    return f"https://www.sec.gov/Archives/edgar/data/{cik_int}/{accession_clean}/{primary_document}"


def extract_sec_sections(html_text: str) -> str:
    soup = BeautifulSoup(html_text, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg", "ix:header"]):
        tag.decompose()
    text = clean_text(soup.get_text(" "))
    sentences = re.split(r"(?<=[.!?])\s+", text)
    section_terms = [
        "consolidated statements of operations", "consolidated statement of operations",
        "consolidated statements of income", "consolidated statement of income",
        "consolidated balance sheets", "consolidated balance sheet",
        "consolidated statements of cash flows", "consolidated statement of cash flows",
        "management's discussion and analysis", "results of operations", "liquidity and capital resources",
        "revenue", "net sales", "gross margin", "operating income", "net income", "cash and cash equivalents",
        "total assets", "total liabilities", "long-term debt", "free cash flow", "capital expenditures",
    ]
    hits: list[str] = []
    seen: set[str] = set()
    for sentence in sentences:
        sentence = clean_text(sentence)
        lowered = sentence.lower()
        if len(sentence) < 35 or len(sentence) > 520:
            continue
        if not any(term in lowered for term in section_terms):
            continue
        signature = sentence[:160].lower()
        if signature in seen:
            continue
        seen.add(signature)
        hits.append(sentence)
        if len(hits) >= 120:
            break
    return clean_text(" ".join(hits))[:MAX_TEXT_CHARS]


def fetch_quarterly_sec_filings(company: str, ticker: str = "", limit: int = 4) -> tuple[SECCompany | None, list[Source]]:
    match = lookup_sec_company(company, ticker)
    if not match:
        return None, [
            Source(
                "SEC quarterly filing lookup",
                f"https://www.sec.gov/edgar/search/#/q={urllib.parse.quote(company)}",
                f"No SEC company match was found for {company}. If this is a public company, enter the exact ticker to pull quarterly 10-Q filings.",
            )
        ]

    submissions_url = f"https://data.sec.gov/submissions/CIK{match.cik}.json"
    submissions = fetch_json(submissions_url)
    if not isinstance(submissions, dict):
        return match, [
            Source(
                "SEC quarterly filing lookup",
                submissions_url,
                f"SEC company match found for {match.title} ({match.ticker}, CIK {match.cik}), but recent filings could not be pulled.",
            )
        ]

    recent = submissions.get("filings", {}).get("recent", {}) if isinstance(submissions.get("filings"), dict) else {}
    forms = recent.get("form", []) if isinstance(recent, dict) else []
    accessions = recent.get("accessionNumber", []) if isinstance(recent, dict) else []
    primary_docs = recent.get("primaryDocument", []) if isinstance(recent, dict) else []
    filing_dates = recent.get("filingDate", []) if isinstance(recent, dict) else []

    sources: list[Source] = []
    for form, accession, primary_doc, filing_date in zip(forms, accessions, primary_docs, filing_dates):
        if form != "10-Q":
            continue
        url = filing_url(match.cik, str(accession), str(primary_doc))
        try:
            response = requests.get(url, headers=sec_headers(), timeout=16)
            response.raise_for_status()
            excerpt = extract_sec_sections(response.text)
        except Exception as exc:
            excerpt = f"Form 10-Q filing for {match.title} filed {filing_date} was located, but text extraction failed: {exc}"
        if excerpt:
            text = (
                f"SEC filing source. Form 10-Q for {match.title} ({match.ticker}), CIK {match.cik}, filed {filing_date}. "
                f"{excerpt}"
            )
            sources.append(Source(f"SEC Form 10-Q: {match.ticker} filed {filing_date}", url, text[:MAX_TEXT_CHARS]))
        if len(sources) >= limit:
            break

    if not sources:
        sources.append(
            Source(
                "SEC quarterly filing lookup",
                f"https://www.sec.gov/edgar/browse/?CIK={match.cik}",
                f"SEC company match found for {match.title} ({match.ticker}, CIK {match.cik}), but no recent Form 10-Q filings were found. The company may be foreign, newly public, inactive, or may file different forms.",
            )
        )
    return match, sources


def extract_pdf_text(raw: bytes, filename: str) -> str:
    try:
        import pypdf  # type: ignore

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = Path(tmp.name)
        try:
            reader = pypdf.PdfReader(str(tmp_path))
            pages = []
            for index, page in enumerate(reader.pages, start=1):
                pages.append(f"[Page {index}] {page.extract_text() or ''}")
            return clean_text("\n".join(pages))
        finally:
            tmp_path.unlink(missing_ok=True)
    except Exception:
        return (
            f"Uploaded PDF '{filename}' was received, but PDF text extraction failed in this runtime. "
            "Install pypdf or paste PDF text into the notes box."
        )


def csv_to_text(raw: bytes, filename: str) -> str:
    try:
        decoded = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(decoded))
        rows = list(reader)[:80]
        if not rows:
            return ""
        lines = [f"CSV file {filename} detected. First rows:"]
        for row in rows:
            lines.append(" | ".join(cell.strip() for cell in row[:16]))
        return clean_text("\n".join(lines))
    except Exception:
        return raw.decode("utf-8", errors="replace")


def parse_multipart(content_type: str, body: bytes) -> tuple[dict[str, str], list[tuple[str, bytes]]]:
    match = re.search(r"boundary=(.+)", content_type)
    if not match:
        return {}, []
    boundary = match.group(1).strip().strip('"').encode()
    fields: dict[str, str] = {}
    files: list[tuple[str, bytes]] = []
    delimiter = b"--" + boundary
    for part in body.split(delimiter):
        part = part.strip(b"\r\n")
        if not part or part == b"--":
            continue
        header_blob, _, payload = part.partition(b"\r\n\r\n")
        headers = header_blob.decode("utf-8", errors="ignore")
        payload = payload.rstrip(b"\r\n")
        name_match = re.search(r'name="([^"]+)"', headers)
        if not name_match:
            continue
        filename_match = re.search(r'filename="([^"]*)"', headers)
        name = name_match.group(1)
        if filename_match and filename_match.group(1):
            files.append((filename_match.group(1), payload))
        else:
            fields[name] = payload.decode("utf-8", errors="replace")
    return fields, files


def source_from_upload(filename: str, raw: bytes) -> Source:
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", filename or "upload")
    saved = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
    saved.write_bytes(raw)
    suffix = saved.suffix.lower()
    if suffix == ".pdf":
        text = extract_pdf_text(raw, filename)
    elif suffix == ".csv":
        text = csv_to_text(raw, filename)
    else:
        text = raw.decode("utf-8", errors="replace")
    return Source(f"Uploaded document: {filename}", str(saved), clean_text(text)[:MAX_TEXT_CHARS])


FIGURE_RE = re.compile(
    r"(?<![\w.])(?:\$|USD\s*)?\d[\d,]*(?:\.\d+)?\s*(?:billion|million|thousand|employees|customers|users|ARR|MRR|revenue|EBITDA|EBIT|gross margin|net income|cash|debt|%|x|mm|bn|m|k)?",
    re.IGNORECASE,
)

FINANCIAL_METRIC_KEYWORDS = {
    "revenue", "sales", "gross margin", "ebitda", "operating income", "net income", "cash", "debt", "assets",
    "liabilities", "free cash flow", "fcf", "capex", "funding", "valuation", "runway", "burn",
}
OPERATING_METRIC_KEYWORDS = {
    "arr", "mrr", "bookings", "churn", "customers", "customer", "users", "contract", "contracts", "backlog",
    "pipeline", "retention", "net retention", "gross retention", "renewal", "renewals", "ltv", "cac",
}
NUMERIC_ALLOW_KEYWORDS = FINANCIAL_METRIC_KEYWORDS | OPERATING_METRIC_KEYWORDS
NOISE_KEYWORDS = {
    "privacy", "privacy policy", "terms", "terms of use", "terms and conditions", "sitemap", "site map",
    "copyright", "phone", "call", "support", "trade in", "cash back", "learn more", "apply now", "promo",
    "promotion", "legal", "footer", "navigation", "cookie", "cookies", "item", "product", "model", "iphone",
    "macbook", "refund", "zip", "address", "episode", "listen", "watch now", "all rights reserved", "store",
    "tel", "fax", "contact", "menu", "accessibility", "apple card", "daily cash", "credit when",
}
BUSINESS_TEXT_NOISE_KEYWORDS = {
    word for word in NOISE_KEYWORDS
    if word not in {"product", "item", "model"}
}
TRANSACTION_KEYWORDS = {
    "acquisition", "acquire", "acquired", "merger", "merge", "joint venture", "partnership", "strategic alliance",
    "divestiture", "divested", "sale of", "sold", "spin-off", "spinoff", "restructuring", "buyback", "repurchase",
    "share repurchase", "dividend", "special dividend", "capital raise", "offering", "debt offering", "bond",
    "loan", "credit facility", "investment", "contract", "backlog", "order", "supply agreement", "settlement",
    "tender offer", "takeover", "privatization",
}
SEC_OR_FINANCIAL_SOURCE_HINTS = {
    "form 10-k", "form 10-q", "s-1", "sec filing", "consolidated statements", "balance sheet", "income statement",
    "statement of operations", "cash flow", "audited", "unaudited", "gaap", "fiscal year", "fiscal quarter",
}


def has_any_phrase(text: str, phrases: Iterable[str]) -> bool:
    lowered = text.lower()
    return any(phrase in lowered for phrase in phrases)


def classify_numeric_claim(context: str) -> str:
    lowered = context.lower()
    if has_any_phrase(lowered, NOISE_KEYWORDS):
        return "Rejected page noise"
    if has_any_phrase(lowered, FINANCIAL_METRIC_KEYWORDS):
        if has_any_phrase(lowered, SEC_OR_FINANCIAL_SOURCE_HINTS) or re.search(r"\b(20\d{2}|19\d{2}|q[1-4]|fy)\b", lowered):
            return "Verified financial metric"
        return "Unverified number"
    if has_any_phrase(lowered, OPERATING_METRIC_KEYWORDS):
        return "Possible operating metric"
    return "Unverified number"


def source_supports_verified_financials(source: Source) -> bool:
    label = source.label.lower()
    text = source.text.lower()
    if label.startswith("uploaded document") or label.startswith("pasted analyst") or label.startswith("sec form 10-q"):
        return True
    return has_any_phrase(text, SEC_OR_FINANCIAL_SOURCE_HINTS)


def classify_figure(value: str, context: str, source: Source) -> tuple[str, str]:
    lowered = context.lower()
    value_clean = value.replace(",", "").strip()
    if re.fullmatch(r"20\d{2}|19\d{2}", value_clean) and not has_any_phrase(lowered, NUMERIC_ALLOW_KEYWORDS):
        return "Rejected page noise", "Looks like a standalone year, not a diligence metric."

    category = classify_numeric_claim(context)
    if category == "Verified financial metric" and not source_supports_verified_financials(source):
        category = "Unverified number"

    if category == "Verified financial metric":
        return category, "Financial context appears in user-provided material, filing text, or another diligence-grade source."
    if category == "Possible operating metric":
        return category, "Operating context exists, but confirm definition, period, and source before relying on it."
    if category == "Rejected page noise":
        return category, "Context looks like footer, legal, contact, navigation, product, or promotional page text."
    return category, "No sufficient financial or operating context near this number."


def extract_figures(sources: list[Source]) -> list[Figure]:
    figures: list[Figure] = []
    seen: set[tuple[str, str]] = set()
    for source in sources:
        for match in FIGURE_RE.finditer(source.text):
            value = clean_text(match.group(0))
            if len(value) < 2:
                continue
            sentence_start = max(source.text.rfind(".", 0, match.start()), source.text.rfind("!", 0, match.start()), source.text.rfind("?", 0, match.start()))
            sentence_start = 0 if sentence_start < 0 else sentence_start + 1
            sentence_ends = [pos for pos in (source.text.find(".", match.end()), source.text.find("!", match.end()), source.text.find("?", match.end())) if pos >= 0]
            sentence_end = min(sentence_ends) + 1 if sentence_ends else min(len(source.text), match.end() + 160)
            context = clean_text(source.text[sentence_start:sentence_end])
            key = (value.lower(), context[:130].lower())
            if key in seen:
                continue
            seen.add(key)
            category, quality_note = classify_figure(value, context, source)
            figures.append(
                Figure(
                    value=value,
                    context=context,
                    source_label=source.label,
                    source_url=source.url,
                    confidence="Sourced, needs human verification",
                    category=category,
                    quality_note=quality_note,
                )
            )
            if len(figures) >= 60:
                return figures
    return figures


def bullets_from_text(text: str, keywords: list[str], limit: int = 6) -> list[str]:
    sentences = re.split(r"(?<=[.!?])\s+", text)
    hits = []
    seen = set()
    for sentence in sentences:
        lowered = sentence.lower()
        if any(word in lowered for word in keywords):
            sentence = clean_text(sentence)
            signature = sentence[:90].lower()
            if 45 <= len(sentence) <= 320 and signature not in seen:
                seen.add(signature)
                hits.append(sentence)
        if len(hits) >= limit:
            break
    return hits


def clean_business_sentences(sources: list[Source], limit: int = 6) -> list[str]:
    allowed = [
        "platform", "software", "service", "services", "product", "products", "company", "business", "customers",
        "mission", "solution", "solutions", "manufactures", "sells", "offers", "operates", "segments", "brands",
        "retail", "automotive", "financial services", "semiconductor", "industrial", "consumer", "healthcare",
    ]
    hits: list[str] = []
    seen: set[str] = set()
    for source in sources:
        if source.label != "Company website" and not source.label.startswith("Pasted"):
            continue
        for sentence in re.split(r"(?<=[.!?])\s+", source.text):
            sentence = clean_text(sentence)
            lowered = sentence.lower()
            if len(sentence) < 45 or len(sentence) > 280:
                continue
            if has_any_phrase(lowered, BUSINESS_TEXT_NOISE_KEYWORDS):
                continue
            if has_any_phrase(lowered, TRANSACTION_KEYWORDS):
                continue
            if re.search(r"\b(?:\+?\d[\d\s().-]{7,}|20\d{2}|19\d{2})\b", sentence):
                continue
            if not any(word in lowered for word in allowed):
                continue
            signature = sentence[:100].lower()
            if signature in seen:
                continue
            seen.add(signature)
            hits.append(sentence)
            if len(hits) >= limit:
                return hits
    return hits


def company_profile_bullets(sources: list[Source], market: MarketAnalysis | None, limit: int = 10) -> list[str]:
    bullets: list[str] = []
    if market and market.target.business_summary:
        summary = market.target.business_summary
        for sentence in re.split(r"(?<=[.!?])\s+", summary):
            sentence = clean_text(sentence)
            if 45 <= len(sentence) <= 420:
                bullets.append(f"{sentence} Source: Yahoo Finance profile.")
            if len(bullets) >= 4:
                break
    if market:
        details = []
        if market.target.sector:
            details.append(f"sector: {market.target.sector}")
        if market.target.industry:
            details.append(f"industry: {market.target.industry}")
        if market.target.country:
            details.append(f"country: {market.target.country}")
        if market.target.employees:
            details.append(f"employees: {market.target.employees:,}")
        if details:
            bullets.append(f"Yahoo profile metadata: {', '.join(details)}.")
    for item in clean_business_sentences(sources, limit=limit):
        bullets.append(f"{item} Source: company website / provided text.")
        if len(bullets) >= limit:
            break
    return bullets[:limit]


def transaction_bullets(sources: list[Source], market: MarketAnalysis | None, limit: int = 12) -> list[str]:
    bullets: list[str] = []
    seen: set[str] = set()

    def add(text: str, source_label: str) -> None:
        text = clean_text(text)
        if len(text) < 45 or len(text) > 460:
            return
        lowered = text.lower()
        if not any(term in lowered for term in TRANSACTION_KEYWORDS):
            return
        sig = text[:140].lower()
        if sig in seen:
            return
        seen.add(sig)
        bullets.append(f"{text} Source: {source_label}.")

    if market:
        for item in market.market_news:
            add(item.title, item.publisher or "Yahoo Finance news")
            if item.summary:
                add(item.summary, item.publisher or "Yahoo Finance news")
            if len(bullets) >= limit:
                return bullets[:limit]

    for source in sources:
        for sentence in re.split(r"(?<=[.!?])\s+", source.text):
            add(sentence, source.label)
            if len(bullets) >= limit:
                return bullets[:limit]
    return bullets[:limit]


def financial_snapshot_figures(figures: list[Figure]) -> list[Figure]:
    return [
        fig for fig in figures
        if fig.category == "Verified financial metric"
        and not fig.source_label.lower().startswith("company website")
    ]


def operating_figures(figures: list[Figure]) -> list[Figure]:
    return [fig for fig in figures if fig.category == "Possible operating metric"]


def has_pasted_notes(sources: list[Source]) -> bool:
    return any(source.label.startswith("Pasted") for source in sources)


def has_website_source(sources: list[Source]) -> bool:
    return any(source.label == "Company website" and source.text for source in sources)


def sec_quarterly_sources(sources: list[Source]) -> list[Source]:
    return [source for source in sources if source.label.startswith("SEC Form 10-Q")]


def has_sec_quarterly_sources(sources: list[Source]) -> bool:
    return bool(sec_quarterly_sources(sources))


def sec_filing_readout(sources: list[Source], limit: int = 6) -> list[str]:
    items: list[str] = []
    terms = [
        "revenue", "net sales", "gross margin", "operating income", "net income", "cash and cash equivalents",
        "total assets", "total liabilities", "debt", "liquidity", "capital resources", "cash flow",
    ]
    for source in sec_quarterly_sources(sources):
        for sentence in re.split(r"(?<=[.!?])\s+", source.text):
            sentence = clean_text(sentence)
            lowered = sentence.lower()
            if 55 <= len(sentence) <= 360 and any(term in lowered for term in terms):
                items.append(f"{sentence} Source: {source.label}.")
                break
        if len(items) >= limit:
            break
    return items


def detect_legal_or_regulatory_language(sources: list[Source]) -> bool:
    corpus = " ".join(source.text.lower() for source in sources)
    relevant_terms = ["litigation", "regulatory", "subpoena", "consent order", "investigation", "compliance", "material adverse", "data breach", "sanction"]
    return any(term in corpus for term in relevant_terms)


def detect_conflicting_numbers(figures: list[Figure]) -> bool:
    by_metric: dict[str, set[str]] = {}
    metric_terms = list(NUMERIC_ALLOW_KEYWORDS)
    for fig in figures:
        if fig.category not in {"Verified financial metric", "Possible operating metric"}:
            continue
        lowered = fig.context.lower()
        metric = next((term for term in metric_terms if term in lowered), "")
        if not metric:
            continue
        normalized_value = re.sub(r"\s+", " ", fig.value.lower().replace(",", ""))
        by_metric.setdefault(metric, set()).add(normalized_value)
    return any(len(values) > 1 for values in by_metric.values())


def period_return(points: list[StockPoint], years: int) -> float | None:
    if not points:
        return None
    cutoff_days = years * 365
    latest = points[-1]
    candidates = [p for p in points if (latest.date - p.date).days >= cutoff_days]
    if not candidates:
        return None
    start = candidates[-1]
    if not start.close:
        return None
    return (latest.close / start.close - 1) * 100


def money(value: float, currency: str = "USD") -> str:
    return f"{currency} {value:,.2f}"


def pct(value: float) -> str:
    return f"{value:+.1f}%"


def raw_yahoo_value(value: object) -> object:
    if isinstance(value, dict):
        if "raw" in value:
            return value["raw"]
        if "fmt" in value:
            return value["fmt"]
    return value


def yahoo_float(value: object) -> float | None:
    raw = raw_yahoo_value(value)
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        cleaned = raw.replace(",", "").replace("%", "").strip()
        try:
            return float(cleaned)
        except Exception:
            return None
    return None


def yahoo_int(value: object) -> int | None:
    num = yahoo_float(value)
    return int(num) if num is not None else None


def compact_money(value: float | None, currency: str = "") -> str:
    if value is None:
        return "Not available"
    sign = "-" if value < 0 else ""
    amount = abs(value)
    if amount >= 1_000_000_000_000:
        text = f"{sign}{amount / 1_000_000_000_000:.2f}T"
    elif amount >= 1_000_000_000:
        text = f"{sign}{amount / 1_000_000_000:.2f}B"
    elif amount >= 1_000_000:
        text = f"{sign}{amount / 1_000_000:.2f}M"
    elif amount >= 1_000:
        text = f"{sign}{amount / 1_000:.2f}K"
    else:
        text = f"{value:,.2f}"
    return f"{currency} {text}".strip()


def fmt_number(value: float | int | None) -> str:
    if value is None:
        return "Not available"
    return f"{value:,.0f}" if float(value).is_integer() else f"{value:,.2f}"


def fmt_ratio(value: float | None) -> str:
    return "Not available" if value is None else f"{value:.2f}x"


def fmt_pct_plain(value: float | None) -> str:
    return "Not available" if value is None else f"{value:.2f}%"


def fetch_yahoo_quote_data(symbol: str) -> dict[str, object]:
    encoded = urllib.parse.quote(symbol)
    data: dict[str, object] = {}
    for host in ("query1.finance.yahoo.com", "query2.finance.yahoo.com"):
        quote_url = f"https://{host}/v7/finance/quote?symbols={encoded}"
        try:
            response = requests.get(quote_url, headers={"User-Agent": USER_AGENT}, timeout=14)
            response.raise_for_status()
            results = response.json().get("quoteResponse", {}).get("result", [])
            if results and isinstance(results[0], dict):
                data.update(results[0])
                break
        except Exception:
            pass

    modules = "price,summaryDetail,defaultKeyStatistics,financialData,assetProfile"
    for host in ("query1.finance.yahoo.com", "query2.finance.yahoo.com"):
        summary_url = f"https://{host}/v10/finance/quoteSummary/{encoded}?modules={modules}"
        try:
            response = requests.get(summary_url, headers={"User-Agent": USER_AGENT}, timeout=14)
            response.raise_for_status()
            results = response.json().get("quoteSummary", {}).get("result", [])
            if results and isinstance(results[0], dict):
                for module_data in results[0].values():
                    if isinstance(module_data, dict):
                        data.update(module_data)
                break
        except Exception:
            pass
    return data


def yahoo_search(query: str, quotes_count: int = 8, news_count: int = 8) -> dict[str, object]:
    query = clean_text(query)
    if not query:
        return {}
    url = (
        "https://query2.finance.yahoo.com/v1/finance/search?"
        + urllib.parse.urlencode({"q": query, "quotesCount": quotes_count, "newsCount": news_count})
    )
    try:
        response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=14)
        response.raise_for_status()
        data = response.json()
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def resolve_market_symbol(company: str, ticker: str = "") -> tuple[str, str]:
    raw = clean_text(ticker or company)
    if not raw:
        return "", "No company or ticker was available for market lookup."
    direct = clean_text(ticker).upper().replace(" ", "")
    if direct and re.fullmatch(r"[A-Z0-9.^=\-]{1,18}(?:\.[A-Z]{1,4})?", direct):
        return direct, f"Using entered market symbol {direct}."

    data = yahoo_search(raw, quotes_count=12, news_count=0)
    quotes = data.get("quotes", []) if isinstance(data, dict) else []
    if not isinstance(quotes, list):
        return "", f"Yahoo Finance symbol search did not return a match for {raw}."

    allowed_types = {"EQUITY", "ETF", "INDEX", "MUTUALFUND"}
    best: tuple[int, dict[str, object]] | None = None
    target = normalize_company_name(company or raw)
    for quote in quotes:
        if not isinstance(quote, dict):
            continue
        symbol = clean_text(str(quote.get("symbol", ""))).upper()
        quote_type = clean_text(str(quote.get("quoteType", ""))).upper()
        short_name = clean_text(str(quote.get("shortname", "") or quote.get("longname", "")))
        if not symbol or quote_type not in allowed_types:
            continue
        normalized_name = normalize_company_name(short_name)
        score = 10
        if target and normalized_name == target:
            score += 80
        elif target and (normalized_name.startswith(target) or target.startswith(normalized_name)):
            score += 55
        elif target:
            score += len(set(target.split()) & set(normalized_name.split())) * 15
        if quote.get("isYahooFinance"):
            score += 5
        if quote_type == "EQUITY":
            score += 8
        if best is None or score > best[0]:
            best = (score, quote)

    if not best:
        return "", f"Yahoo Finance symbol search did not return an equity, ETF, fund, or index match for {raw}."
    quote = best[1]
    symbol = clean_text(str(quote.get("symbol", ""))).upper()
    name = clean_text(str(quote.get("shortname", "") or quote.get("longname", "") or symbol))
    exchange = clean_text(str(quote.get("exchDisp", "") or quote.get("exchange", "")))
    note = f"Resolved {raw} to {symbol}"
    if name:
        note += f" ({name})"
    if exchange:
        note += f" on {exchange}"
    note += " using Yahoo Finance global symbol search."
    return symbol, note


def yahoo_search_quote(symbol: str, company: str = "") -> dict[str, object]:
    data = yahoo_search(symbol or company, quotes_count=12, news_count=0)
    quotes = data.get("quotes", []) if isinstance(data, dict) else []
    if not isinstance(quotes, list):
        return {}
    symbol_upper = clean_text(symbol).upper()
    for quote in quotes:
        if isinstance(quote, dict) and clean_text(str(quote.get("symbol", ""))).upper() == symbol_upper:
            return quote
    for quote in quotes:
        if isinstance(quote, dict) and quote.get("quoteType"):
            return quote
    return {}


def market_disclosure_links(series: StockSeries, company: str) -> list[tuple[str, str]]:
    symbol = series.ticker.upper()
    exchange = (series.exchange or "").lower()
    country = (series.country or "").lower()
    query = urllib.parse.quote(company or series.name or series.ticker)
    links: list[tuple[str, str]] = [
        ("Yahoo Finance quote page", series.source_url),
    ]
    if series.website:
        links.append(("Company website / investor relations starting point", series.website))

    def add(label: str, url: str) -> None:
        if url and all(existing_url != url for _, existing_url in links):
            links.append((label, url))

    if "." not in symbol and (country in {"united states", "usa", "us"} or not country):
        add("SEC EDGAR company search", f"https://www.sec.gov/edgar/search/#/q={query}")
    if symbol.endswith(".T") or "tokyo" in exchange or "jpx" in exchange:
        add("Japan EDINET disclosure search", f"https://disclosure2.edinet-fsa.go.jp/WEEK0010.aspx")
        add("JPX listed company search", "https://www.jpx.co.jp/english/listing/co-search/index.html")
    if symbol.endswith(".L") or "london" in exchange:
        add("London Stock Exchange issuer search", f"https://www.londonstockexchange.com/search?query={query}")
        add("UK Companies House search", f"https://find-and-update.company-information.service.gov.uk/search?q={query}")
    if symbol.endswith(".TO") or symbol.endswith(".V") or "toronto" in exchange or country == "canada":
        add("SEDAR+ Canadian filings search", "https://www.sedarplus.ca/")
        add("TMX issuer search", f"https://money.tmx.com/en/search?query={query}")
    if symbol.endswith(".AX") or "australian" in exchange or country == "australia":
        add("ASX announcements search", f"https://www.asx.com.au/markets/company/{urllib.parse.quote(symbol.split('.')[0])}")
    if symbol.endswith(".KS") or symbol.endswith(".KQ") or "korea" in exchange:
        add("Korea DART disclosure search", "https://englishdart.fss.or.kr/")
        add("KRX listed company search", "https://global.krx.co.kr/")
    if symbol.endswith(".HK") or "hong kong" in exchange:
        add("HKEXnews issuer disclosures", "https://www.hkexnews.hk/index.htm")
    if symbol.endswith(".SS") or symbol.endswith(".SZ") or "shanghai" in exchange or "shenzhen" in exchange:
        add("Shanghai Stock Exchange disclosures", "https://english.sse.com.cn/")
        add("Shenzhen Stock Exchange disclosures", "https://www.szse.cn/English/")
    if symbol.endswith(".NS") or symbol.endswith(".BO") or country == "india":
        add("NSE company filings", "https://www.nseindia.com/companies-listing/corporate-filings-announcements")
        add("BSE corporate announcements", "https://www.bseindia.com/corporates/ann.html")
    if symbol.endswith(".PA") or symbol.endswith(".AS") or symbol.endswith(".BR") or symbol.endswith(".DE") or symbol.endswith(".F") or "euronext" in exchange:
        add("Euronext issuer search", f"https://live.euronext.com/en/search_instruments/{query}")
    if symbol.endswith(".SW") or country == "switzerland":
        add("SIX Swiss Exchange issuer search", "https://www.six-group.com/en/products-services/the-swiss-stock-exchange/market-data/shares.html")
    if symbol.endswith(".SA") or country == "brazil":
        add("CVM Brazil filings search", "https://www.gov.br/cvm/pt-br")
        add("B3 listed companies", "https://www.b3.com.br/en_us/products-and-services/trading/equities/listed-companies.htm")
    return links[:10]


def stooq_symbol(symbol: str) -> str:
    symbol = clean_text(symbol).upper()
    mapped = {
        "SPY": "spy.us",
        "QQQ": "qqq.us",
        "VGK": "vgk.us",
        "^GSPC": "^spx",
        "^IXIC": "^ndq",
        "^DJI": "^dji",
        "^FTSE": "^ukx",
        "^N225": "^nkx",
        "^GDAXI": "^dax",
        "^FCHI": "^cac",
        "^HSI": "^hsi",
        "^KS11": "^kospi",
        "^AXJO": "^aor",
        "^GSPTSE": "^tsx",
    }
    if symbol in mapped:
        return mapped[symbol]
    suffix_map = {
        ".SS": ".cn",
        ".SZ": ".cn",
        ".T": ".jp",
        ".L": ".uk",
        ".TO": ".ca",
        ".AX": ".au",
        ".DE": ".de",
        ".PA": ".fr",
        ".HK": ".hk",
        ".KS": ".kr",
        ".NS": ".in",
    }
    for suffix, stooq_suffix in suffix_map.items():
        if symbol.endswith(suffix):
            return symbol[: -len(suffix)].lower() + stooq_suffix
    if "." not in symbol and re.fullmatch(r"[A-Z]{1,5}", symbol):
        return symbol.lower() + ".us"
    return symbol.lower()


def solve_stooq_challenge(session: requests.Session, html_text: str) -> bool:
    match = re.search(r'const c="([^"]+)",d=(\d+)', html_text)
    if not match:
        return False
    challenge = match.group(1)
    difficulty = int(match.group(2))
    prefix = "0" * difficulty
    for nonce in range(2_000_000):
        digest = hashlib.sha256(f"{challenge}{nonce}".encode("utf-8")).hexdigest()
        if digest.startswith(prefix):
            response = session.post(
                "https://stooq.com/__verify",
                data={"c": challenge, "n": str(nonce)},
                headers={"User-Agent": USER_AGENT, "Content-Type": "application/x-www-form-urlencoded"},
                timeout=14,
            )
            return response.ok
    return False


def fetch_stooq_csv(symbol: str) -> str:
    stooq = stooq_symbol(symbol)
    url = f"https://stooq.com/q/d/l/?s={urllib.parse.quote(stooq)}&i=w"
    session = requests.Session()
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/csv,text/plain,*/*",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        response = session.get(url, headers=headers, timeout=14)
        text = response.text
        if "__verify" in text and "crypto.subtle.digest" in text:
            if solve_stooq_challenge(session, text):
                response = session.get(url, headers=headers, timeout=14)
                text = response.text
        return text if "Date," in text and "Close" in text else ""
    except Exception:
        return ""


def fetch_stooq_stock_series(ticker: str, years: int = 5) -> StockSeries | None:
    ticker = clean_text(ticker).upper().replace(" ", "")
    text = fetch_stooq_csv(ticker)
    if not text:
        return None
    rows = list(csv.DictReader(io.StringIO(text)))
    points: list[StockPoint] = []
    volumes: list[float] = []
    cutoff_year = datetime.now().year - min(max(int(years or 5), 1), 10) - 1
    for row in rows:
        try:
            dt = datetime.strptime(str(row.get("Date", "")), "%Y-%m-%d")
            if dt.year < cutoff_year:
                continue
            close = float(str(row.get("Close", "")).replace(",", ""))
            points.append(StockPoint(dt, close))
            volume_raw = str(row.get("Volume", "")).replace(",", "")
            if volume_raw:
                volumes.append(float(volume_raw))
        except Exception:
            continue
    points.sort(key=lambda point: point.date)
    if len(points) < 12:
        return None
    profile = market_profile(ticker)
    latest_close = points[-1].close
    prior_close = points[-2].close if len(points) >= 2 else None
    last_52_points = points[-52:] if len(points) >= 52 else points
    avg_volume = sum(volumes[-52:]) / len(volumes[-52:]) if volumes else None
    return StockSeries(
        ticker=ticker,
        currency=first_text(profile.get("currency"), "USD"),
        points=points,
        source_url=f"https://stooq.com/q/?s={urllib.parse.quote(stooq_symbol(ticker))}",
        name=first_text(profile.get("name"), ticker),
        exchange=first_text(profile.get("exchange"), "Stooq"),
        quote_type="EQUITY" if ticker not in {symbol for symbol, _ in sum(COUNTRY_MARKET_BENCHMARKS.values(), [])} else "INDEX",
        market_region="",
        regular_market_price=latest_close,
        previous_close=prior_close,
        fifty_two_week_low=min((p.close for p in last_52_points), default=None),
        fifty_two_week_high=max((p.close for p in last_52_points), default=None),
        average_volume=avg_volume,
        sector=first_text(profile.get("sector")),
        industry=first_text(profile.get("industry")),
        country="",
        business_summary=first_text(profile.get("products")),
    )


def fetch_stock_series(ticker: str, years: int = 5) -> StockSeries | None:
    ticker = clean_text(ticker).upper().replace(" ", "")
    if not ticker or not re.fullmatch(r"[A-Z0-9.^=\-]{1,18}(?:\.[A-Z]{1,4})?", ticker):
        return None
    years = min(max(int(years or 5), 1), 10)
    encoded = urllib.parse.quote(ticker)
    result = None
    quote = {}
    closes = []
    volumes = []
    meta = {}
    try:
        for host in ("query1.finance.yahoo.com", "query2.finance.yahoo.com"):
            url = f"https://{host}/v8/finance/chart/{encoded}?range={years}y&interval=1wk&events=history"
            response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=14)
            response.raise_for_status()
            result = response.json()["chart"]["result"][0]
            if result:
                break
        if not result:
            return None
        timestamps = result.get("timestamp", [])
        quote = result.get("indicators", {}).get("quote", [{}])[0]
        closes = quote.get("close", [])
        volumes = quote.get("volume", [])
        meta = result.get("meta", {})
    except Exception:
        return fetch_stooq_stock_series(ticker, years)
    points: list[StockPoint] = []
    for ts, close in zip(timestamps, closes):
        if close is None:
            continue
        points.append(StockPoint(datetime.fromtimestamp(int(ts)), float(close)))
    points.sort(key=lambda p: p.date)
    if len(points) < 12:
        return fetch_stooq_stock_series(ticker, years)
    search_quote = yahoo_search_quote(ticker)
    quote_data = fetch_yahoo_quote_data(ticker)
    for key, value in search_quote.items():
        quote_data.setdefault(key, value)
    dividend_yield = yahoo_float(quote_data.get("dividendYield"))
    if dividend_yield is not None and dividend_yield < 1:
        dividend_yield *= 100
    latest_close = points[-1].close if points else None
    prior_close = points[-2].close if len(points) >= 2 else None
    last_52_points = points[-52:] if len(points) >= 52 else points
    derived_52_low = min((p.close for p in last_52_points), default=None)
    derived_52_high = max((p.close for p in last_52_points), default=None)
    valid_volumes = [float(v) for v in volumes[-52:] if isinstance(v, (int, float)) and v > 0]
    derived_avg_volume = sum(valid_volumes) / len(valid_volumes) if valid_volumes else None
    return StockSeries(
        ticker=ticker,
        currency=clean_text(str(quote_data.get("currency") or meta.get("currency", "USD"))),
        points=points,
        source_url=f"https://finance.yahoo.com/quote/{encoded}",
        name=clean_text(str(quote_data.get("longName") or quote_data.get("shortName") or meta.get("longName", "") or meta.get("shortName", ""))),
        exchange=clean_text(str(quote_data.get("fullExchangeName") or quote_data.get("exchDisp") or quote_data.get("exchange") or meta.get("exchangeName", "") or meta.get("fullExchangeName", ""))),
        quote_type=clean_text(str(quote_data.get("quoteType") or quote_data.get("typeDisp") or "")),
        market_region=clean_text(str(quote_data.get("region", ""))),
        regular_market_price=yahoo_float(quote_data.get("regularMarketPrice")) or yahoo_float(quote_data.get("currentPrice")) or latest_close,
        previous_close=yahoo_float(quote_data.get("regularMarketPreviousClose") or quote_data.get("previousClose")) or prior_close,
        day_low=yahoo_float(quote_data.get("regularMarketDayLow") or quote_data.get("dayLow")),
        day_high=yahoo_float(quote_data.get("regularMarketDayHigh") or quote_data.get("dayHigh")),
        fifty_two_week_low=yahoo_float(quote_data.get("fiftyTwoWeekLow")) or derived_52_low,
        fifty_two_week_high=yahoo_float(quote_data.get("fiftyTwoWeekHigh")) or derived_52_high,
        market_cap=yahoo_float(quote_data.get("marketCap")),
        trailing_pe=yahoo_float(quote_data.get("trailingPE")),
        forward_pe=yahoo_float(quote_data.get("forwardPE")),
        price_to_book=yahoo_float(quote_data.get("priceToBook")),
        dividend_yield_pct=dividend_yield,
        beta=yahoo_float(quote_data.get("beta")),
        average_volume=yahoo_float(quote_data.get("averageDailyVolume3Month") or quote_data.get("averageVolume")) or derived_avg_volume,
        shares_outstanding=yahoo_float(quote_data.get("sharesOutstanding")),
        sector=clean_text(str(quote_data.get("sector", ""))),
        industry=clean_text(str(quote_data.get("industry", ""))),
        country=clean_text(str(quote_data.get("country", ""))),
        website=clean_text(str(quote_data.get("website", ""))),
        employees=yahoo_int(quote_data.get("fullTimeEmployees")),
        business_summary=clean_text(str(quote_data.get("longBusinessSummary", "")))[:1200],
    )


def fetch_market_news(symbol: str, company: str, limit: int = 12) -> list[MarketNews]:
    news: list[MarketNews] = []
    seen: set[str] = set()

    def add_item(title: str, publisher: str, link: str, published: datetime | None, summary: str = "") -> None:
        title_clean = clean_text(title)
        link_clean = clean_text(link)
        if not title_clean:
            return
        key = (title_clean[:120] + link_clean).lower()
        if key in seen:
            return
        seen.add(key)
        news.append(MarketNews(title=title_clean, publisher=clean_text(publisher), link=link_clean, published=published, summary=clean_text(summary)))

    data = yahoo_search(symbol or company, quotes_count=0, news_count=limit)
    for item in data.get("news", []) if isinstance(data, dict) else []:
        if not isinstance(item, dict):
            continue
        published = None
        ts = item.get("providerPublishTime")
        if isinstance(ts, (int, float)):
            published = datetime.fromtimestamp(ts)
        add_item(
            str(item.get("title", "")),
            str(item.get("publisher", "")),
            str(item.get("link", "")),
            published,
            str(item.get("summary", "")),
        )

    if len(news) < limit and symbol:
        rss_url = f"https://feeds.finance.yahoo.com/rss/2.0/headline?s={urllib.parse.quote(symbol)}&region=US&lang=en-US"
        try:
            response = requests.get(rss_url, headers={"User-Agent": USER_AGENT}, timeout=14)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "xml")
            for item in soup.find_all("item"):
                published = None
                pub_date = clean_text(item.pubDate.get_text(" ") if item.pubDate else "")
                if pub_date:
                    try:
                        published = parsedate_to_datetime(pub_date).replace(tzinfo=None)
                    except Exception:
                        published = None
                add_item(
                    item.title.get_text(" ") if item.title else "",
                    "Yahoo Finance RSS",
                    item.link.get_text(" ") if item.link else "",
                    published,
                    item.description.get_text(" ") if item.description else "",
                )
                if len(news) >= limit:
                    break
        except Exception:
            pass
    target_terms = {term for term in normalize_company_name(company).split() if len(term) >= 4}
    symbol_root = (symbol or "").split(".")[0].replace("^", "").lower()
    if symbol_root and not symbol_root.isdigit() and len(symbol_root) >= 3:
        target_terms.add(symbol_root)
    if target_terms:
        relevant = [
            item for item in news
            if any(term in f"{item.title} {item.summary}".lower() for term in target_terms)
        ]
        return relevant[:limit]
    return news[:limit]


def explain_large_moves(series: StockSeries, news: list[MarketNews], limit: int = 4) -> list[MoveCatalyst]:
    returns = weekly_returns(series.points)
    if not returns:
        return []
    stdev = statistics.stdev(returns) if len(returns) > 2 else 0.0
    threshold = max(0.08, stdev * 1.75)
    moves = []
    for index, ret in enumerate(returns, start=1):
        if abs(ret) >= threshold:
            moves.append((abs(ret), ret, series.points[index].date))
    moves.sort(reverse=True, key=lambda item: item[0])
    catalysts: list[MoveCatalyst] = []
    for _, ret, date in moves[:limit]:
        nearby = []
        for item in news:
            if item.published and abs((item.published - date).days) <= 14:
                nearby.append(item)
        if not nearby:
            nearby = news[:3]
        direction = "rose sharply" if ret > 0 else "dropped sharply"
        if nearby:
            headline_text = "; ".join(item.title for item in nearby[:3])
            explanation = (
                f"{series.ticker} {direction} {pct(ret * 100)} in the week ending {date.strftime('%Y-%m-%d')}. "
                f"Nearby/current news that may explain the move includes: {headline_text}. "
                "Treat this as a catalyst hypothesis, not confirmed causation."
            )
        else:
            explanation = (
                f"{series.ticker} {direction} {pct(ret * 100)} in the week ending {date.strftime('%Y-%m-%d')}, "
                "but the news search did not return headlines to support a catalyst read."
            )
        catalysts.append(MoveCatalyst(label=f"{series.ticker} {direction}", date=date, move_pct=ret * 100, explanation=explanation, headlines=nearby[:3]))
    return catalysts


def weekly_returns(points: list[StockPoint]) -> list[float]:
    returns = []
    for prior, now in zip(points, points[1:]):
        if prior.close:
            returns.append(now.close / prior.close - 1)
    return returns


def max_drawdown(points: list[StockPoint]) -> tuple[float, float]:
    peak = points[0].close
    worst = 0.0
    current = 0.0
    for point in points:
        peak = max(peak, point.close)
        current = point.close / peak - 1 if peak else 0.0
        worst = min(worst, current)
    return worst * 100, current * 100


def compute_metrics(series: StockSeries) -> StockMetrics:
    points = series.points
    current = points[-1].close
    start = points[0].close
    total_return = (current / start - 1) * 100 if start else 0.0
    days = max((points[-1].date - points[0].date).days, 1)
    cagr = ((current / start) ** (365 / days) - 1) * 100 if start else 0.0
    returns = weekly_returns(points)
    volatility = statistics.stdev(returns) * math.sqrt(52) * 100 if len(returns) > 2 else 0.0
    sharpe_like = (cagr / volatility) if volatility else 0.0
    worst_dd, curr_dd = max_drawdown(points)
    positives = [r for r in returns if r > 0]
    positive_week_pct = len(positives) / len(returns) * 100 if returns else 0.0
    best_week = max(returns) * 100 if returns else 0.0
    worst_week = min(returns) * 100 if returns else 0.0
    by_year: dict[int, list[StockPoint]] = {}
    for p in points:
        by_year.setdefault(p.date.year, []).append(p)
    annual_returns: dict[int, float] = {}
    for year, year_points in sorted(by_year.items()):
        if len(year_points) >= 2 and year_points[0].close:
            annual_returns[year] = (year_points[-1].close / year_points[0].close - 1) * 100
    last_40 = points[-40:] if len(points) >= 40 else points
    ma_40 = sum(p.close for p in last_40) / len(last_40)
    if current > ma_40 * 1.08:
        trend = "Strong uptrend: price is well above the roughly 40-week moving average."
    elif current > ma_40 * 1.03:
        trend = "Moderate uptrend: price is above the roughly 40-week moving average."
    elif current < ma_40 * 0.92:
        trend = "Strong downtrend: price is well below the roughly 40-week moving average."
    elif current < ma_40 * 0.97:
        trend = "Moderate downtrend: price is below the roughly 40-week moving average."
    else:
        trend = "Sideways / mixed: price is close to the roughly 40-week moving average."
    return StockMetrics(
        ticker=series.ticker,
        currency=series.currency,
        latest_price=current,
        start_price=start,
        high_price=max(p.close for p in points),
        low_price=min(p.close for p in points),
        total_return_pct=total_return,
        cagr_pct=cagr,
        volatility_pct=volatility,
        sharpe_like=sharpe_like,
        max_drawdown_pct=worst_dd,
        current_drawdown_pct=curr_dd,
        best_week_pct=best_week,
        worst_week_pct=worst_week,
        positive_week_pct=positive_week_pct,
        weeks=len(points),
        trend_label=trend,
        ma_40=ma_40,
        annual_returns=annual_returns,
    )


def indexed_points(series: StockSeries) -> list[tuple[datetime, float]]:
    first = series.points[0].close or 1
    return [(p.date, p.close / first * 100) for p in series.points]


def drawdown_points(series: StockSeries) -> list[tuple[datetime, float]]:
    peak = series.points[0].close
    out = []
    for p in series.points:
        peak = max(peak, p.close)
        out.append((p.date, (p.close / peak - 1) * 100 if peak else 0.0))
    return out


def svg_polyline(points: list[tuple[datetime, float]], width: int, height: int, left: int, right: int, top: int, bottom: int, low: float, high: float) -> str:
    if not points:
        return ""
    span = high - low or 1
    step = (width - left - right) / max(len(points) - 1, 1)
    coords = []
    for i, (_, value) in enumerate(points):
        x = left + i * step
        y = top + (high - value) / span * (height - top - bottom)
        coords.append(f"{x:.1f},{y:.1f}")
    return " ".join(coords)


def clamp_rate(rate_pct: float) -> float:
    # Guardrail so volatility math does not create impossible below-zero prices.
    return max(rate_pct, -90.0)


def build_growth_scenarios(metrics: StockMetrics) -> list[GrowthScenario]:
    # Historical-trend scenario math. This is not a prediction.
    # Base case uses pulled-period CAGR. Bear/bull flex around that using volatility.
    scenario_rates = [
        ("Bear case", clamp_rate(metrics.cagr_pct - 0.50 * metrics.volatility_pct)),
        ("Base trend", clamp_rate(metrics.cagr_pct)),
        ("Bull case", clamp_rate(metrics.cagr_pct + 0.50 * metrics.volatility_pct)),
    ]
    horizons = [1, 3, 5, 10]
    scenarios: list[GrowthScenario] = []
    for label, annual_rate in scenario_rates:
        prices: dict[int, float] = {}
        returns: dict[int, float] = {}
        factor = 1 + annual_rate / 100
        for years in horizons:
            projected_price = metrics.latest_price * (factor ** years) if factor > 0 else 0.0
            prices[years] = projected_price
            returns[years] = (projected_price / metrics.latest_price - 1) * 100 if metrics.latest_price else 0.0
        scenarios.append(GrowthScenario(label=label, annual_rate_pct=annual_rate, projected_prices=prices, projected_returns=returns))
    return scenarios


def scenario_read(metrics: StockMetrics, scenarios: list[GrowthScenario]) -> str:
    bear = next((s for s in scenarios if s.label == "Bear case"), None)
    base = next((s for s in scenarios if s.label == "Base trend"), None)
    bull = next((s for s in scenarios if s.label == "Bull case"), None)
    if not (bear and base and bull):
        return "Scenario model unavailable."
    return (
        f"Using the stock's pulled history, the 1-year trend range is about "
        f"{pct(bear.projected_returns[1])} to {pct(bull.projected_returns[1])}, "
        f"with the base trend at {pct(base.projected_returns[1])}. "
        f"The 5-year historical-trend range is about {pct(bear.projected_returns[5])} to {pct(bull.projected_returns[5])}. "
        f"The 10-year historical-trend stress range is about {pct(bear.projected_returns[10])} to {pct(bull.projected_returns[10])}. "
        f"This is not a prediction; it is a stress-test based on past CAGR and volatility."
    )


def make_forecast_svg(metrics: StockMetrics, scenarios: list[GrowthScenario], path: Path, title: str) -> None:
    width, height = 980, 390
    left, right, top, bottom = 70, 190, 44, 52
    years = [0, 1, 3, 5, 10]
    palette = ["#8a4a2f", "#146c5f", "#1f5e9d"]
    series_values = []
    for scenario in scenarios:
        rate = 1 + scenario.annual_rate_pct / 100
        vals = [metrics.latest_price * (rate ** year) if rate > 0 else 0.0 for year in years]
        series_values.append((scenario, vals))
    all_values = [v for _, vals in series_values for v in vals] + [metrics.latest_price]
    low, high = min(all_values), max(all_values)
    if low == high:
        high = low + 1
    plot_w = width - left - right
    plot_h = height - top - bottom

    def x_for(i: int) -> float:
        return left + i / (len(years) - 1) * plot_w

    def y_for(v: float) -> float:
        return top + (high - v) / (high - low) * plot_h

    lines = []
    legends = []
    for idx, (scenario, vals) in enumerate(series_values):
        color = palette[idx % len(palette)]
        coords = " ".join(f"{x_for(i):.1f},{y_for(v):.1f}" for i, v in enumerate(vals))
        lines.append(f'<polyline points="{coords}" fill="none" stroke="{color}" stroke-width="3" stroke-linejoin="round" stroke-linecap="round"/>')
        y = top + 25 + idx * 26
        legends.append(f'<rect x="{width-right+25}" y="{y-10}" width="14" height="4" fill="{color}"/><text x="{width-right+46}" y="{y}" font-family="Arial" font-size="13" fill="#16211f">{html.escape(scenario.label)} ({scenario.annual_rate_pct:+.1f}%/yr)</text>')
    x_labels = "".join(f'<text x="{x_for(i):.1f}" y="{height-18}" text-anchor="middle" font-family="Arial" font-size="12" fill="#5e6c68">+{year}Y</text>' for i, year in enumerate(years))
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <rect width="100%" height="100%" fill="#ffffff"/>
  <text x="{left}" y="26" font-family="Arial" font-size="18" font-weight="700" fill="#16211f">{html.escape(title)}</text>
  <line x1="{left}" y1="{height-bottom}" x2="{width-right}" y2="{height-bottom}" stroke="#d8dfdc"/>
  <line x1="{left}" y1="{top}" x2="{left}" y2="{height-bottom}" stroke="#d8dfdc"/>
  <text x="8" y="{top+6}" font-family="Arial" font-size="12" fill="#5e6c68">{high:,.0f}</text>
  <text x="8" y="{height-bottom}" font-family="Arial" font-size="12" fill="#5e6c68">{low:,.0f}</text>
  <text x="{left}" y="{height-34}" font-family="Arial" font-size="12" fill="#5e6c68">Projected years from latest close</text>
  {x_labels}
  {''.join(lines)}
  {''.join(legends)}
</svg>"""
    path.write_text(svg, encoding="utf-8")


def make_line_svg(series: StockSeries, path: Path, title: str, y_label: str) -> None:
    width, height = 980, 360
    left, right, top, bottom = 62, 28, 44, 48
    values = [p.close for p in series.points]
    low, high = min(values), max(values)
    coords = svg_polyline([(p.date, p.close) for p in series.points], width, height, left, right, top, bottom, low, high)
    first_date = series.points[0].date.strftime("%Y")
    last_date = series.points[-1].date.strftime("%Y")
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <rect width="100%" height="100%" fill="#ffffff"/>
  <text x="{left}" y="26" font-family="Arial" font-size="18" font-weight="700" fill="#16211f">{html.escape(title)}</text>
  <line x1="{left}" y1="{height-bottom}" x2="{width-right}" y2="{height-bottom}" stroke="#d8dfdc"/>
  <line x1="{left}" y1="{top}" x2="{left}" y2="{height-bottom}" stroke="#d8dfdc"/>
  <text x="8" y="{top+6}" font-family="Arial" font-size="12" fill="#5e6c68">{high:,.0f}</text>
  <text x="8" y="{height-bottom}" font-family="Arial" font-size="12" fill="#5e6c68">{low:,.0f}</text>
  <text x="{left}" y="{height-14}" font-family="Arial" font-size="12" fill="#5e6c68">{first_date}</text>
  <text x="{width-right-34}" y="{height-14}" font-family="Arial" font-size="12" fill="#5e6c68">{last_date}</text>
  <text x="{left}" y="{height-30}" font-family="Arial" font-size="12" fill="#5e6c68">{html.escape(y_label)}</text>
  <polyline points="{coords}" fill="none" stroke="#146c5f" stroke-width="3" stroke-linejoin="round" stroke-linecap="round"/>
</svg>"""
    path.write_text(svg, encoding="utf-8")


def make_multi_index_svg(series_list: list[StockSeries], path: Path, title: str) -> None:
    width, height = 980, 390
    left, right, top, bottom = 62, 190, 44, 48
    indexed = [(s, indexed_points(s)) for s in series_list]
    all_values = [v for _, pts in indexed for _, v in pts]
    low, high = (min(all_values), max(all_values)) if all_values else (0, 1)
    palette = ["#146c5f", "#7a4d9a", "#b45f06", "#1f5e9d", "#7a7a22", "#8b2f4a"]
    polylines = []
    legends = []
    for idx, (series, pts) in enumerate(indexed[:6]):
        color = palette[idx % len(palette)]
        coords = svg_polyline(pts, width, height, left, right, top, bottom, low, high)
        polylines.append(f'<polyline points="{coords}" fill="none" stroke="{color}" stroke-width="3" stroke-linejoin="round" stroke-linecap="round"/>')
        y = top + 25 + idx * 23
        legends.append(f'<rect x="{width-right+25}" y="{y-10}" width="14" height="4" fill="{color}"/><text x="{width-right+46}" y="{y}" font-family="Arial" font-size="13" fill="#16211f">{html.escape(series.ticker)}</text>')
    first_date = series_list[0].points[0].date.strftime("%Y") if series_list else ""
    last_date = series_list[0].points[-1].date.strftime("%Y") if series_list else ""
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <rect width="100%" height="100%" fill="#ffffff"/>
  <text x="{left}" y="26" font-family="Arial" font-size="18" font-weight="700" fill="#16211f">{html.escape(title)}</text>
  <line x1="{left}" y1="{height-bottom}" x2="{width-right}" y2="{height-bottom}" stroke="#d8dfdc"/>
  <line x1="{left}" y1="{top}" x2="{left}" y2="{height-bottom}" stroke="#d8dfdc"/>
  <text x="8" y="{top+6}" font-family="Arial" font-size="12" fill="#5e6c68">{high:,.0f}</text>
  <text x="8" y="{height-bottom}" font-family="Arial" font-size="12" fill="#5e6c68">{low:,.0f}</text>
  <text x="{left}" y="{height-14}" font-family="Arial" font-size="12" fill="#5e6c68">{first_date}</text>
  <text x="{width-right-34}" y="{height-14}" font-family="Arial" font-size="12" fill="#5e6c68">{last_date}</text>
  <text x="{left}" y="{height-30}" font-family="Arial" font-size="12" fill="#5e6c68">Indexed to 100 at start</text>
  {''.join(polylines)}
  {''.join(legends)}
</svg>"""
    path.write_text(svg, encoding="utf-8")


def make_drawdown_svg(series: StockSeries, path: Path, title: str) -> None:
    width, height = 980, 330
    left, right, top, bottom = 62, 28, 44, 48
    pts = drawdown_points(series)
    low = min(v for _, v in pts) if pts else -1
    high = 0.0
    coords = svg_polyline(pts, width, height, left, right, top, bottom, low, high)
    first_date = series.points[0].date.strftime("%Y")
    last_date = series.points[-1].date.strftime("%Y")
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <rect width="100%" height="100%" fill="#ffffff"/>
  <text x="{left}" y="26" font-family="Arial" font-size="18" font-weight="700" fill="#16211f">{html.escape(title)}</text>
  <line x1="{left}" y1="{top}" x2="{width-right}" y2="{top}" stroke="#d8dfdc"/>
  <line x1="{left}" y1="{height-bottom}" x2="{width-right}" y2="{height-bottom}" stroke="#d8dfdc"/>
  <line x1="{left}" y1="{top}" x2="{left}" y2="{height-bottom}" stroke="#d8dfdc"/>
  <text x="8" y="{top+6}" font-family="Arial" font-size="12" fill="#5e6c68">0%</text>
  <text x="8" y="{height-bottom}" font-family="Arial" font-size="12" fill="#5e6c68">{low:.0f}%</text>
  <text x="{left}" y="{height-14}" font-family="Arial" font-size="12" fill="#5e6c68">{first_date}</text>
  <text x="{width-right-34}" y="{height-14}" font-family="Arial" font-size="12" fill="#5e6c68">{last_date}</text>
  <polyline points="{coords}" fill="none" stroke="#8a4a2f" stroke-width="3" stroke-linejoin="round" stroke-linecap="round"/>
</svg>"""
    path.write_text(svg, encoding="utf-8")


def make_bar_svg(values: dict[int, float], path: Path, title: str) -> None:
    width, height = 980, 360
    left, right, top, bottom = 62, 28, 44, 62
    items = list(values.items())[-8:]
    if not items:
        path.write_text("", encoding="utf-8")
        return
    max_abs = max(abs(v) for _, v in items) or 1
    zero_y = top + (height - top - bottom) / 2
    bar_gap = 18
    bar_width = (width - left - right - bar_gap * (len(items) - 1)) / len(items)
    scale = (height - top - bottom) / 2 / max_abs
    bars = []
    for index, (year, value) in enumerate(items):
        x = left + index * (bar_width + bar_gap)
        bar_h = abs(value) * scale
        y = zero_y - bar_h if value >= 0 else zero_y
        bars.append(f"""
  <rect x="{x:.1f}" y="{y:.1f}" width="{bar_width:.1f}" height="{bar_h:.1f}" fill="#146c5f" opacity="0.86"/>
  <text x="{x + bar_width/2:.1f}" y="{height-28}" text-anchor="middle" font-family="Arial" font-size="12" fill="#5e6c68">{year}</text>
  <text x="{x + bar_width/2:.1f}" y="{y-7 if value >= 0 else y+bar_h+16:.1f}" text-anchor="middle" font-family="Arial" font-size="12" fill="#16211f">{value:+.0f}%</text>""")
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <rect width="100%" height="100%" fill="#ffffff"/>
  <text x="{left}" y="26" font-family="Arial" font-size="18" font-weight="700" fill="#16211f">{html.escape(title)}</text>
  <line x1="{left}" y1="{zero_y:.1f}" x2="{width-right}" y2="{zero_y:.1f}" stroke="#d8dfdc"/>
  {''.join(bars)}
</svg>"""
    path.write_text(svg, encoding="utf-8")


def fetch_market_analysis(company: str, ticker: str, peer_tickers: str, benchmark_ticker: str, years: int) -> MarketAnalysis | None:
    resolved_symbol, symbol_note = resolve_market_symbol(company, ticker)
    target = fetch_stock_series(resolved_symbol, years) if resolved_symbol else None
    if not target:
        return None
    likely_non_us = (
        "." in target.ticker
        or (target.country and target.country.lower() not in {"united states", "usa", "us"})
        or (target.currency and target.currency.upper() not in {"USD", "USX"})
    )
    if likely_non_us:
        international_note = (
            "This appears to be a non-U.S. or cross-market listing, so the app uses Yahoo Finance for quote, exchange, "
            "currency, valuation, trading-range, profile, performance, peer, benchmark, and news-catalyst data. "
            "SEC 10-Q data is only expected when the company has U.S. SEC reporting obligations."
        )
    else:
        international_note = (
            "This appears to be a U.S. market listing. The app combines SEC filing pulls when available with Yahoo Finance market data."
        )
    target_metrics = compute_metrics(target)
    peer_list = []
    seen = {target.ticker}
    for raw in re.split(r"[,\s]+", peer_tickers or ""):
        raw = raw.strip().upper()
        if raw and raw not in seen:
            seen.add(raw)
            series = fetch_stock_series(raw, years)
            if series:
                peer_list.append(series)
        if len(peer_list) >= 4:
            break
    benchmark = None
    if benchmark_ticker:
        benchmark_symbol, _ = resolve_market_symbol("", benchmark_ticker)
        benchmark = fetch_stock_series(benchmark_symbol or benchmark_ticker, years)
    peer_metrics = [compute_metrics(s) for s in peer_list]
    market_news = fetch_market_news(target.ticker, company)
    move_catalysts = explain_large_moves(target, market_news)
    disclosure_links = market_disclosure_links(target, company)
    safe = re.sub(r"[^A-Z0-9_.-]+", "_", target.ticker)
    price_chart = OUTPUT_DIR / f"{safe}_price.svg"
    indexed_chart = OUTPUT_DIR / f"{safe}_indexed_compare.svg"
    drawdown_chart = OUTPUT_DIR / f"{safe}_drawdown.svg"
    annual_chart = OUTPUT_DIR / f"{safe}_annual_returns.svg"
    forecast_chart = OUTPUT_DIR / f"{safe}_growth_loss_scenarios.svg"
    scenarios = build_growth_scenarios(target_metrics)
    make_line_svg(target, price_chart, f"{target.ticker} weekly closing price", "Close")
    compare_series = [target] + peer_list + ([benchmark] if benchmark else [])
    make_multi_index_svg(compare_series, indexed_chart, f"Indexed performance comparison")
    make_drawdown_svg(target, drawdown_chart, f"{target.ticker} drawdown from prior high")
    make_bar_svg(target_metrics.annual_returns, annual_chart, f"{target.ticker} annual returns")
    make_forecast_svg(target_metrics, scenarios, forecast_chart, f"{target.ticker} potential growth / loss scenarios")
    return MarketAnalysis(
        target=target,
        target_metrics=target_metrics,
        benchmark=benchmark,
        peers=peer_list,
        peer_metrics=peer_metrics,
        market_news=market_news,
        move_catalysts=move_catalysts,
        symbol_note=symbol_note,
        international_note=international_note,
        disclosure_links=disclosure_links,
        price_chart_path=f"/outputs/{price_chart.name}",
        indexed_chart_path=f"/outputs/{indexed_chart.name}",
        drawdown_chart_path=f"/outputs/{drawdown_chart.name}",
        annual_chart_path=f"/outputs/{annual_chart.name}",
        forecast_chart_path=f"/outputs/{forecast_chart.name}",
        scenarios=scenarios,
        source_note="Yahoo Finance chart API. Market data is unaudited public market data and should be checked before investment use.",
    )


def make_figure_table(figures: list[Figure], category_filter: str | None = None, limit: int = 20) -> str:
    rows = []
    selected = figures
    if category_filter:
        selected = [fig for fig in figures if fig.category == category_filter]
    for index, fig in enumerate(selected[:limit], start=1):
        source = f"[{md_cell(fig.source_label)}]({fig.source_url})" if fig.source_url.startswith("http") else md_cell(fig.source_label)
        rows.append(
            f"| F{index} | {md_cell(fig.value, 80)} | {md_cell(fig.category, 100)} | {md_cell(fig.context, 460)} | {source} | {md_cell(fig.quality_note, 220)} |"
        )
    return "\n".join(rows) or "| - | Not disclosed | - | No source-backed figure found | - | - |"


def bullet_block(items: list[str], fallback: str) -> str:
    if not items:
        return f"- {fallback}"
    return "\n".join(f"- {item}" for item in items)


def yahoo_detail_rows(series: StockSeries) -> str:
    rows = [
        ("Company / instrument", series.name or series.ticker),
        ("Yahoo symbol", series.ticker),
        ("Quote type", series.quote_type or "Not available"),
        ("Exchange", series.exchange or "Not available"),
        ("Country", series.country or "Not available"),
        ("Currency", series.currency or "Not available"),
        ("Sector", series.sector or "Not available"),
        ("Industry", series.industry or "Not available"),
        ("Latest Yahoo quote", money(series.regular_market_price, series.currency) if series.regular_market_price is not None else "Not available"),
        ("Previous close", money(series.previous_close, series.currency) if series.previous_close is not None else "Not available"),
        ("Day range", f"{money(series.day_low, series.currency)} to {money(series.day_high, series.currency)}" if series.day_low is not None and series.day_high is not None else "Not available"),
        ("52-week range", f"{money(series.fifty_two_week_low, series.currency)} to {money(series.fifty_two_week_high, series.currency)}" if series.fifty_two_week_low is not None and series.fifty_two_week_high is not None else "Not available"),
        ("Market capitalization", compact_money(series.market_cap, series.currency)),
        ("Trailing P/E", fmt_ratio(series.trailing_pe)),
        ("Forward P/E", fmt_ratio(series.forward_pe)),
        ("Price / book", fmt_ratio(series.price_to_book)),
        ("Dividend yield", fmt_pct_plain(series.dividend_yield_pct)),
        ("Beta", fmt_ratio(series.beta).replace("x", "")),
        ("Average volume", fmt_number(series.average_volume)),
        ("Shares outstanding", fmt_number(series.shares_outstanding)),
        ("Employees", fmt_number(series.employees)),
        ("Website", f"[{md_cell(series.website, 80)}]({series.website})" if series.website.startswith("http") else (series.website or "Not available")),
    ]
    return "\n".join(f"| {md_cell(label, 120)} | {md_cell(value, 260)} |" for label, value in rows)


def trading_snapshot_rows(series: StockSeries) -> str:
    move_from_prev = None
    if series.regular_market_price is not None and series.previous_close:
        move_from_prev = (series.regular_market_price / series.previous_close - 1) * 100
    rows = [
        ("Currently trading at", money(series.regular_market_price, series.currency) if series.regular_market_price is not None else "Not available"),
        ("Currency", series.currency or "Not available"),
        ("Exchange", series.exchange or "Not available"),
        ("Instrument type", series.quote_type or "Not available"),
        ("Previous close", money(series.previous_close, series.currency) if series.previous_close is not None else "Not available"),
        ("Move vs previous close", pct(move_from_prev) if move_from_prev is not None else "Not available"),
        ("Day range", f"{money(series.day_low, series.currency)} to {money(series.day_high, series.currency)}" if series.day_low is not None and series.day_high is not None else "Not available"),
        ("52-week range", f"{money(series.fifty_two_week_low, series.currency)} to {money(series.fifty_two_week_high, series.currency)}" if series.fifty_two_week_low is not None and series.fifty_two_week_high is not None else "Not available"),
        ("Average volume", fmt_number(series.average_volume)),
        ("Market cap", compact_money(series.market_cap, series.currency)),
        ("Sector / industry", f"{series.sector or 'Not available'} / {series.industry or 'Not available'}"),
    ]
    return "\n".join(f"| {md_cell(label, 120)} | {md_cell(value, 240)} |" for label, value in rows)


def disclosure_link_rows(links: list[tuple[str, str]]) -> str:
    if not links:
        return "| - | No public disclosure links were identified. | - |"
    rows = []
    for label, url in links:
        rows.append(f"| {md_cell(label, 180)} | [{md_cell(url, 90)}]({url}) |")
    return "\n".join(rows)


def median(values: list[float]) -> float | None:
    clean = sorted(v for v in values if v is not None and v > 0)
    if not clean:
        return None
    mid = len(clean) // 2
    if len(clean) % 2:
        return clean[mid]
    return (clean[mid - 1] + clean[mid]) / 2


def standalone_pe_read(pe: float | None) -> str:
    if pe is None or pe <= 0:
        return "P/E not available or not meaningful; company may be loss-making or Yahoo did not disclose the field."
    if pe < 12:
        return "Low P/E: potentially inexpensive, but check whether earnings are cyclical, declining, or unusually high."
    if pe <= 25:
        return "Moderate P/E: valuation is not obviously stretched on earnings alone."
    if pe <= 40:
        return "High P/E: investors are paying a premium for growth, quality, scarcity, or margin expansion."
    return "Very high P/E: potentially overvalued unless growth, margins, and durability strongly support the premium."


def standalone_pb_read(pb: float | None, industry: str = "") -> str:
    if pb is None or pb <= 0:
        return "P/B not available or not meaningful from Yahoo."
    lowered = industry.lower()
    asset_heavy = any(term in lowered for term in ["bank", "insurance", "reit", "real estate", "utility", "industrial", "auto", "manufacturer"])
    if pb < 1:
        return "Below book value: can signal cheap assets, distress, weak returns, or balance-sheet skepticism."
    if pb <= 3:
        return "Moderate P/B: generally not stretched, especially for asset-heavy businesses." if asset_heavy else "Moderate P/B: not obviously stretched, but less decisive for asset-light companies."
    if pb <= 8:
        return "High P/B: market is assigning a premium to returns, brand, growth, or intangible assets."
    return "Very high P/B: potentially overvalued unless return on equity, margins, and growth are exceptional."


def valuation_assessment(market: MarketAnalysis | None) -> list[str]:
    if not market:
        return ["No market valuation data was available."]
    target = market.target
    peer_pe = [peer.trailing_pe for peer in market.peers if peer.trailing_pe and peer.trailing_pe > 0]
    peer_pb = [peer.price_to_book for peer in market.peers if peer.price_to_book and peer.price_to_book > 0]
    peer_pe_median = median(peer_pe)
    peer_pb_median = median(peer_pb)
    bullets = [
        f"P/E read: {fmt_ratio(target.trailing_pe)}. {standalone_pe_read(target.trailing_pe)}",
        f"P/B read: {fmt_ratio(target.price_to_book)}. {standalone_pb_read(target.price_to_book, target.industry)}",
    ]
    if peer_pe_median:
        premium = ((target.trailing_pe / peer_pe_median - 1) * 100) if target.trailing_pe else None
        if premium is not None:
            bullets.append(f"Peer P/E comparison: target P/E is {pct(premium)} versus peer median P/E of {peer_pe_median:.2f}x.")
    else:
        bullets.append("Peer P/E comparison: peer P/E data was not available from Yahoo; add peer tickers to improve this read.")
    if peer_pb_median:
        premium = ((target.price_to_book / peer_pb_median - 1) * 100) if target.price_to_book else None
        if premium is not None:
            bullets.append(f"Peer P/B comparison: target P/B is {pct(premium)} versus peer median P/B of {peer_pb_median:.2f}x.")
    else:
        bullets.append("Peer P/B comparison: peer P/B data was not available from Yahoo; add peer tickers to improve this read.")

    overvaluation_signals = 0
    if target.trailing_pe and target.trailing_pe > 30:
        overvaluation_signals += 1
    if target.price_to_book and target.price_to_book > 5:
        overvaluation_signals += 1
    if peer_pe_median and target.trailing_pe and target.trailing_pe > peer_pe_median * 1.3:
        overvaluation_signals += 1
    if peer_pb_median and target.price_to_book and target.price_to_book > peer_pb_median * 1.3:
        overvaluation_signals += 1
    if market.target_metrics.cagr_pct < 0 and (target.trailing_pe and target.trailing_pe > 20):
        overvaluation_signals += 1

    if overvaluation_signals >= 3:
        conclusion = "Valuation conclusion: potentially overvalued on available Yahoo P/E/P/B and market-performance evidence. Confirm with growth, margins, ROE, and forward guidance before relying on this."
    elif overvaluation_signals == 2:
        conclusion = "Valuation conclusion: valuation looks demanding, but not enough evidence to call it clearly overvalued without peer/growth/margin confirmation."
    elif overvaluation_signals == 1:
        conclusion = "Valuation conclusion: one valuation warning is present; treat as a diligence question rather than a firm overvaluation call."
    else:
        conclusion = "Valuation conclusion: available P/E/P/B data does not show obvious overvaluation, but missing data or weak future growth could change that."
    bullets.append(conclusion)
    bullets.append("Important caveat: P/E and P/B are incomplete by themselves. For banks/insurers, P/B and ROE matter more; for software/asset-light companies, P/B can look high without proving overvaluation; for cyclical companies, low P/E can be a value trap near peak earnings.")
    return bullets


def valuation_rows(market: MarketAnalysis) -> str:
    target = market.target
    peer_pe_median = median([peer.trailing_pe for peer in market.peers if peer.trailing_pe and peer.trailing_pe > 0])
    peer_pb_median = median([peer.price_to_book for peer in market.peers if peer.price_to_book and peer.price_to_book > 0])
    rows = [
        ("Trailing P/E", fmt_ratio(target.trailing_pe), f"Peer median: {peer_pe_median:.2f}x" if peer_pe_median else "Peer median unavailable"),
        ("Forward P/E", fmt_ratio(target.forward_pe), "Forward estimate from Yahoo when available"),
        ("Price / book", fmt_ratio(target.price_to_book), f"Peer median: {peer_pb_median:.2f}x" if peer_pb_median else "Peer median unavailable"),
        ("Market cap", compact_money(target.market_cap, target.currency), "Yahoo Finance market capitalization"),
        ("Dividend yield", fmt_pct_plain(target.dividend_yield_pct), "Yield can support valuation if payout is sustainable"),
        ("Beta", fmt_ratio(target.beta).replace("x", ""), "Higher beta usually means higher equity-risk discount needed"),
    ]
    return "\n".join(f"| {md_cell(label, 120)} | {md_cell(value, 120)} | {md_cell(note, 260)} |" for label, value, note in rows)


def liquidity_value(series: StockSeries) -> float | None:
    if series.regular_market_price is None or series.average_volume is None:
        return None
    return series.regular_market_price * series.average_volume


def is_consumable_stock(series: StockSeries) -> bool:
    text = f"{series.sector} {series.industry} {series.business_summary}".lower()
    terms = [
        "consumer defensive", "consumer cyclical", "food", "beverage", "tobacco", "household", "personal",
        "grocery", "retail", "apparel", "luxury", "automotive", "restaurant", "staples", "cosmetics",
    ]
    return any(term in text for term in terms)


def overvaluation_score(series: StockSeries, metrics: StockMetrics, spy_metrics: StockMetrics | None) -> tuple[int, list[str]]:
    score = 0
    reasons: list[str] = []
    if series.trailing_pe and series.trailing_pe > 30:
        score += 2
        reasons.append(f"high trailing P/E of {series.trailing_pe:.2f}x")
    elif series.trailing_pe and series.trailing_pe > 22:
        score += 1
        reasons.append(f"elevated trailing P/E of {series.trailing_pe:.2f}x")
    if series.price_to_book and series.price_to_book > 6:
        score += 2
        reasons.append(f"high P/B of {series.price_to_book:.2f}x")
    elif series.price_to_book and series.price_to_book > 3.5:
        score += 1
        reasons.append(f"elevated P/B of {series.price_to_book:.2f}x")
    if metrics.cagr_pct < 0:
        score += 1
        reasons.append(f"negative pulled-period CAGR of {pct(metrics.cagr_pct)}")
    if metrics.max_drawdown_pct <= -35:
        score += 1
        reasons.append(f"major max drawdown of {metrics.max_drawdown_pct:.1f}%")
    if metrics.volatility_pct >= 30:
        score += 1
        reasons.append(f"high annualized volatility of {metrics.volatility_pct:.1f}%")
    if spy_metrics and metrics.total_return_pct < spy_metrics.total_return_pct:
        score += 1
        reasons.append(f"underperformed SPY by {spy_metrics.total_return_pct - metrics.total_return_pct:.1f} percentage points over the pulled period")
    if not reasons:
        reasons.append("no strong overvaluation signal from Yahoo P/E/P/B and market data")
    return score, reasons


def shortability_read(series: StockSeries) -> str:
    liq = liquidity_value(series)
    if liq is None:
        return "Liquidity unknown; verify borrow availability and trading volume with broker."
    if liq >= 100_000_000:
        return "Very liquid by Yahoo price x volume screen; likely easier to trade, but borrow availability still must be checked."
    if liq >= 25_000_000:
        return "Moderately liquid by Yahoo price x volume screen; check spreads, local market hours, and borrow."
    return "Lower liquidity by Yahoo price x volume screen; may be harder or costly to short."


def market_profile(symbol: str) -> dict[str, str]:
    return COUNTRY_MARKET_PROFILES.get(clean_text(symbol).upper(), {})


def market_baseline(symbol: str) -> dict[str, float | str]:
    return FALLBACK_MARKET_BASELINES.get(clean_text(symbol).upper(), {})


def baseline_float(symbol: str, key: str) -> float | None:
    value = market_baseline(symbol).get(key)
    return float(value) if isinstance(value, (int, float)) else None


def fallback_overvaluation_read(symbol: str, pe: float | None, pb: float | None, volatility: float | None, drawdown: float | None) -> tuple[int, str]:
    reasons = []
    score = 0
    if pe and pe > 30:
        score += 2
        reasons.append(f"high fallback P/E of {pe:.1f}x")
    elif pe and pe > 22:
        score += 1
        reasons.append(f"elevated fallback P/E of {pe:.1f}x")
    if pb and pb > 6:
        score += 2
        reasons.append(f"high fallback P/B of {pb:.1f}x")
    elif pb and pb > 3.5:
        score += 1
        reasons.append(f"elevated fallback P/B of {pb:.1f}x")
    if volatility and volatility >= 30:
        score += 1
        reasons.append(f"high fallback volatility of {volatility:.1f}%")
    if drawdown and drawdown <= -35:
        score += 1
        reasons.append(f"major fallback drawdown of {drawdown:.1f}%")
    if not reasons:
        reasons.append("fallback baseline does not show a strong overvaluation signal")
    suffix = " Baseline values are not live market data; verify against filings, Yahoo, exchange data, or licensed data before relying on them."
    return score, "; ".join(reasons) + suffix


def first_text(*values: object) -> str:
    for value in values:
        text = clean_text(str(value or ""))
        if text:
            return text
    return ""


def yahoo_data_status(quote: dict[str, object], verified_fields: list[str], has_history: bool = False) -> str:
    if verified_fields and has_history:
        return "Yahoo verified quote fields and price history: " + ", ".join(verified_fields)
    if verified_fields:
        return "Yahoo verified quote fields: " + ", ".join(verified_fields) + ". Price-history check was unavailable."
    if quote:
        return "Yahoo recognized the symbol, but the public endpoint did not return valuation/trading fields in this run."
    return "Yahoo did not return live fields in this run; company identity shown from the app's built-in country universe and must be manually verified."


def series_data_status(series: StockSeries, verified_fields: list[str]) -> str:
    if "stooq.com" in (series.source_url or "").lower():
        return "Stooq verified historical price/volume data. Yahoo quote fields were unavailable or rate-limited, so valuation fields still require filing or paid-data verification."
    return yahoo_data_status({"symbol": series.ticker}, verified_fields, has_history=True)


def yahoo_fact_check_symbol(symbol: str, country: str = "", years: int = 3) -> dict[str, object]:
    symbol = clean_text(symbol).upper()
    profile = market_profile(symbol)
    baseline = market_baseline(symbol)
    quote = fetch_yahoo_quote_data(symbol)
    search_quote = yahoo_search_quote(symbol)
    for key, value in search_quote.items():
        quote.setdefault(key, value)
    price = yahoo_float(quote.get("regularMarketPrice") or quote.get("currentPrice")) or baseline_float(symbol, "price")
    avg_volume = yahoo_float(quote.get("averageDailyVolume3Month") or quote.get("averageVolume")) or baseline_float(symbol, "avg_volume")
    market_cap = yahoo_float(quote.get("marketCap")) or baseline_float(symbol, "market_cap")
    pe = yahoo_float(quote.get("trailingPE")) or baseline_float(symbol, "pe")
    forward_pe = yahoo_float(quote.get("forwardPE")) or baseline_float(symbol, "forward_pe")
    pb = yahoo_float(quote.get("priceToBook")) or baseline_float(symbol, "pb")
    dividend_yield = yahoo_float(quote.get("dividendYield")) or baseline_float(symbol, "dividend_yield")
    beta = yahoo_float(quote.get("beta")) or baseline_float(symbol, "beta")
    return_1y = baseline_float(symbol, "return_1y")
    return_3y = baseline_float(symbol, "return_3y")
    volatility = baseline_float(symbol, "volatility")
    drawdown = baseline_float(symbol, "drawdown")
    name = first_text(quote.get("longName"), quote.get("shortName"), quote.get("shortname"), quote.get("longname"), profile.get("name"), symbol)
    exchange = first_text(quote.get("fullExchangeName"), quote.get("exchDisp"), quote.get("exchange"), profile.get("exchange"))
    currency = first_text(quote.get("currency"), profile.get("currency"))
    sector = first_text(quote.get("sector"), profile.get("sector"))
    industry = first_text(quote.get("industry"), profile.get("industry"))
    products = first_text(quote.get("longBusinessSummary"), profile.get("products"))
    source_url = f"https://finance.yahoo.com/quote/{urllib.parse.quote(symbol)}"
    verified_fields = []
    for label, value in [
        ("price", price),
        ("P/E", pe),
        ("P/B", pb),
        ("market cap", market_cap),
        ("average volume", avg_volume),
    ]:
        if value is not None:
            verified_fields.append(label)
    if baseline and not quote:
        status = "Live Yahoo was unavailable or rate-limited; fallback baseline profile filled this row for first-pass screening. Verify all values before investment use."
    elif baseline:
        status = "Yahoo returned partial data; fallback baseline profile filled missing fields. Verify all values before investment use."
    else:
        status = yahoo_data_status(quote, verified_fields)
    fallback_score, fallback_reasons = fallback_overvaluation_read(symbol, pe, pb, volatility, drawdown)
    return {
        "Country": country,
        "Ticker": symbol,
        "Name": name,
        "Exchange": exchange,
        "Currency": currency,
        "Sector": sector,
        "Industry": industry,
        "Products / What They Sell": products[:700],
        "Price": price,
        "Market Cap": market_cap,
        "Avg Volume": avg_volume,
        "Liquidity Value": price * avg_volume if price is not None and avg_volume is not None else None,
        "Trailing PE": pe,
        "Forward PE": forward_pe,
        "Price To Book": pb,
        "P/E and P/B Available": "Yes" if pe and pe > 0 and pb and pb > 0 else "No",
        "Dividend Yield %": dividend_yield,
        "Beta": beta,
        "1Y Return %": return_1y,
        "3Y Return %": return_3y,
        "CAGR %": return_3y,
        "Volatility %": volatility,
        "Max Drawdown %": drawdown,
        "SPY Comparison": "Fallback baseline used; compare manually against live SPY/QQQ/VGK data before use." if baseline else "Price history unavailable; benchmark comparison not calculated.",
        "Overvaluation Score": fallback_score if baseline else 0,
        "Overvaluation Reasons": fallback_reasons if baseline else "P/E and/or P/B unavailable from Yahoo public data; cannot rank as overvalued without verified valuation fields.",
        "Shortability / Liquidity Read": shortability_read(StockSeries(symbol, currency, [StockPoint(datetime.now(), price or 0.0)], source_url, regular_market_price=price, average_volume=avg_volume)) if baseline else "Liquidity unknown; verify borrow availability, local shares/ADR access, spreads, and broker short inventory.",
        "Western Market Comparison": "Compared to SPY when enough Yahoo price history is available.",
        "S&P Global Placeholder": "Requires licensed S&P Global Market Intelligence / Capital IQ / Compustat access; not scraped by this app.",
        "Yahoo Fact Check": status,
        "Yahoo URL": source_url,
    }


def run_country_screener(country: str, years: int = 3) -> list[dict[str, object]]:
    country = country if country in COUNTRY_MARKET_UNIVERSES else ""
    if not country:
        return []
    spy = fetch_stock_series("SPY", years)
    spy_metrics = compute_metrics(spy) if spy else None
    rows: list[dict[str, object]] = []
    for symbol in COUNTRY_MARKET_UNIVERSES[country]:
        profile = market_profile(symbol)
        series = fetch_stock_series(symbol, years)
        if not series:
            rows.append(yahoo_fact_check_symbol(symbol, country, years))
            continue
        metrics = compute_metrics(series)
        universe_fit = "Consumer/consumables"
        if not is_consumable_stock(series):
            summary_text = f"{series.name} {series.industry}".lower()
            if not any(term in summary_text for term in ["retail", "food", "beverage", "consumer", "auto", "apparel", "tobacco", "grocery"]):
                universe_fit = "Included for full country screen; Yahoo sector text is not clearly consumer/consumables."
        has_pe_pb = bool(series.trailing_pe and series.trailing_pe > 0 and series.price_to_book and series.price_to_book > 0)
        if has_pe_pb:
            score, reasons = overvaluation_score(series, metrics, spy_metrics)
        else:
            score, reasons = 0, ["P/E and/or P/B unavailable from Yahoo public data; cannot rank as overvalued without licensed fundamentals."]
        rows.append(
            {
                "Country": country,
                "Ticker": series.ticker,
                "Name": first_text(series.name, profile.get("name"), series.ticker),
                "Exchange": first_text(series.exchange, profile.get("exchange")),
                "Currency": first_text(series.currency, profile.get("currency")),
                "Sector": first_text(series.sector, profile.get("sector")),
                "Industry": first_text(series.industry, profile.get("industry")),
                "Products / What They Sell": first_text(series.business_summary, profile.get("products"))[:700],
                "Price": series.regular_market_price,
                "Market Cap": series.market_cap,
                "Avg Volume": series.average_volume,
                "Liquidity Value": liquidity_value(series),
                "Trailing PE": series.trailing_pe,
                "Forward PE": series.forward_pe,
                "Price To Book": series.price_to_book,
                "P/E and P/B Available": "Yes" if has_pe_pb else "No",
                "Dividend Yield %": series.dividend_yield_pct,
                "Beta": series.beta,
                "1Y Return %": period_return(series.points, 1),
                "3Y Return %": period_return(series.points, 3),
                "CAGR %": metrics.cagr_pct,
                "Volatility %": metrics.volatility_pct,
                "Max Drawdown %": metrics.max_drawdown_pct,
                "SPY Comparison": f"{metrics.total_return_pct - spy_metrics.total_return_pct:+.1f} pp vs SPY" if spy_metrics else "SPY unavailable",
                "Overvaluation Score": score,
                "Overvaluation Reasons": "; ".join(reasons),
                "Shortability / Liquidity Read": shortability_read(series),
                "Universe Fit": universe_fit,
                "Western Market Comparison": "Compared to SPY as U.S. western-market benchmark; add QQQ/VGK manually for deeper cross-market work.",
                "S&P Global Placeholder": "Requires licensed S&P Global Market Intelligence / Capital IQ / Compustat access; not scraped by this app.",
                "Yahoo Fact Check": series_data_status(
                    series,
                    [
                        label for label, value in [
                            ("price", series.regular_market_price),
                            ("P/E", series.trailing_pe),
                            ("P/B", series.price_to_book),
                            ("market cap", series.market_cap),
                            ("average volume", series.average_volume),
                        ]
                        if value is not None
                    ],
                ),
                "Yahoo URL": series.source_url,
            }
        )
    rows.sort(key=lambda row: (float(row.get("Overvaluation Score") or 0), float(row.get("Liquidity Value") or 0)), reverse=True)
    SCREENER_STATE["country"] = country
    SCREENER_STATE["rows"] = rows
    return rows


def run_country_market_screen(country: str, years: int = 3) -> list[dict[str, object]]:
    country = country if country in COUNTRY_MARKET_BENCHMARKS else ""
    symbols: list[tuple[str, str]] = []
    if country:
        symbols.extend(COUNTRY_MARKET_BENCHMARKS[country])
    symbols.extend((symbol, symbol) for symbol in WESTERN_BENCHMARKS)
    rows: list[dict[str, object]] = []
    seen: set[str] = set()
    for symbol, label in symbols:
        if symbol in seen:
            continue
        seen.add(symbol)
        series = fetch_stock_series(symbol, years)
        if not series:
            fact = yahoo_fact_check_symbol(symbol, country, years)
            baseline = market_baseline(symbol)
            rows.append(
                {
                    "Market": label,
                    "Ticker": symbol,
                    "Region": country if symbol not in WESTERN_BENCHMARKS else "Western benchmark",
                    "Currency": fact.get("Currency"),
                    "Latest Price": fact.get("Price") or baseline_float(symbol, "price"),
                    "1Y Return %": baseline_float(symbol, "return_1y"),
                    "3Y Return %": baseline_float(symbol, "return_3y"),
                    f"{years}Y Return %": baseline_float(symbol, "return_period") or baseline_float(symbol, "return_3y"),
                    "Volatility %": baseline_float(symbol, "volatility"),
                    "Max Drawdown %": baseline_float(symbol, "drawdown"),
                    "Trend": "Fallback baseline used because live Yahoo/Stooq market history was unavailable." if baseline else "Market data unavailable from Yahoo/Stooq in this run.",
                    "Fact Check": "Fallback baseline profile used; verify against live Yahoo, exchange data, or broker market data." if baseline else fact.get("Yahoo Fact Check"),
                    "Yahoo URL": fact.get("Yahoo URL"),
                }
            )
            continue
        metrics = compute_metrics(series)
        rows.append(
            {
                "Market": label if label != symbol else (series.name or symbol),
                "Ticker": symbol,
                "Region": country if symbol not in WESTERN_BENCHMARKS else "Western benchmark",
                "Currency": series.currency,
                "Latest Price": series.regular_market_price,
                "1Y Return %": period_return(series.points, 1),
                "3Y Return %": period_return(series.points, 3),
                f"{years}Y Return %": metrics.total_return_pct,
                "Volatility %": metrics.volatility_pct,
                "Max Drawdown %": metrics.max_drawdown_pct,
                "Trend": metrics.trend_label,
                "Fact Check": series_data_status(series, ["price", "price history"]),
                "Yahoo URL": series.source_url,
            }
        )
    SCREENER_STATE["market_rows"] = rows
    return rows


def screener_rows_markdown(rows: list[dict[str, object]], limit: int = 12) -> str:
    if not rows:
        return "| - | No Yahoo fact-check rows were produced. Select a supported country and try again. | - | - | - | - | - |"
    out = []
    for row in rows[:limit]:
        out.append(
            f"| {md_cell(row.get('Ticker'), 80)} | {md_cell(row.get('Name'), 180)} | {md_cell(row.get('Exchange'), 120)} | {md_cell(fmt_number(yahoo_float(row.get('Price'))), 80)} | {compact_money(yahoo_float(row.get('Market Cap')), str(row.get('Currency') or ''))} | {fmt_ratio(yahoo_float(row.get('Trailing PE')))} | {fmt_ratio(yahoo_float(row.get('Price To Book')))} | {md_cell(row.get('Overvaluation Score'), 40)} | {md_cell(row.get('P/E and P/B Available'), 40)} |"
        )
    return "\n".join(out)


def stock_detail_markdown(rows: list[dict[str, object]], limit: int = 12) -> str:
    if not rows:
        return "- No stock details available."
    details = []
    for row in rows[:limit]:
        details.append(
            f"### {md_cell(row.get('Name'), 180)} ({md_cell(row.get('Ticker'), 80)})\n"
            f"- Products / what they sell: {md_cell(row.get('Products / What They Sell'), 500)}\n"
            f"- Sector / industry: {md_cell(row.get('Sector'), 120)} / {md_cell(row.get('Industry'), 160)}\n"
            f"- Trading details: exchange {md_cell(row.get('Exchange'), 160)}, currency {md_cell(row.get('Currency'), 40)}, price {md_cell(fmt_number(yahoo_float(row.get('Price'))), 80)}, market cap {compact_money(yahoo_float(row.get('Market Cap')), str(row.get('Currency') or ''))}, average volume {md_cell(fmt_number(yahoo_float(row.get('Avg Volume'))), 80)}.\n"
            f"- Valuation: trailing P/E {fmt_ratio(yahoo_float(row.get('Trailing PE')))}, forward P/E {fmt_ratio(yahoo_float(row.get('Forward PE')))}, P/B {fmt_ratio(yahoo_float(row.get('Price To Book')))}, dividend yield {fmt_pct_plain(yahoo_float(row.get('Dividend Yield %')))}, beta {fmt_ratio(yahoo_float(row.get('Beta'))).replace('x', '')}.\n"
            f"- Performance / risk: 1Y return {fmt_pct_plain(yahoo_float(row.get('1Y Return %')))}, 3Y return {fmt_pct_plain(yahoo_float(row.get('3Y Return %')))}, volatility {fmt_pct_plain(yahoo_float(row.get('Volatility %')))}, max drawdown {fmt_pct_plain(yahoo_float(row.get('Max Drawdown %')))}, SPY comparison {md_cell(row.get('SPY Comparison'), 180)}.\n"
            f"- Overvaluation read: score {md_cell(row.get('Overvaluation Score'), 40)}; {md_cell(row.get('Overvaluation Reasons'), 420)}\n"
            f"- Shortability / liquidity: {md_cell(row.get('Shortability / Liquidity Read'), 320)}\n"
            f"- Data audit: {md_cell(row.get('Yahoo Fact Check'), 320)} Source: [Yahoo Finance]({md_cell(row.get('Yahoo URL'), 220)})"
        )
    return "\n\n".join(details)


def market_rows_markdown(rows: list[dict[str, object]], years: int) -> str:
    if not rows:
        return "| - | No market benchmark rows were produced. | - | - | - | - | - | - | - |"
    out = []
    for row in rows:
        out.append(
            f"| {md_cell(row.get('Market'), 160)} | {md_cell(row.get('Ticker'), 80)} | {md_cell(row.get('Region'), 120)} | {md_cell(fmt_number(yahoo_float(row.get('Latest Price'))), 90)} | {fmt_pct_plain(yahoo_float(row.get('1Y Return %')))} | {fmt_pct_plain(yahoo_float(row.get('3Y Return %')))} | {fmt_pct_plain(yahoo_float(row.get(f'{years}Y Return %')))} | {fmt_pct_plain(yahoo_float(row.get('Volatility %')))} | {fmt_pct_plain(yahoo_float(row.get('Max Drawdown %')))} | {md_cell(row.get('Trend'), 140)} |"
        )
    return "\n".join(out)


def market_detail_markdown(rows: list[dict[str, object]]) -> str:
    if not rows:
        return "- No market detail available."
    return "\n".join(
        f"- {md_cell(row.get('Market'), 160)} ({md_cell(row.get('Ticker'), 80)}): {md_cell(row.get('Fact Check'), 260)} Source: [Yahoo Finance]({md_cell(row.get('Yahoo URL'), 220)})"
        for row in rows
    )


def screener_section(country: str, rows: list[dict[str, object]], market_rows: list[dict[str, object]] | None = None, years: int = 3) -> str:
    if not country:
        return ""
    market_rows = market_rows or []
    return f"""
## Country Overvaluation / Short Candidate Screener
Country selected: {country}

This screen focuses on liquid consumer/consumables stocks and requires both P/E and P/B to rank a stock. It uses Yahoo Finance public fields first, Stooq historical price data as a no-key fallback, and clearly labeled baseline profiles when free live endpoints are rate-limited. S&P Global Market Intelligence, Capital IQ, and Compustat data are licensed datasets, so this app does not scrape them or invent S&P values.

Western-market comparison: SPY is used as the default U.S. benchmark. For a fuller western comparison, review SPY, QQQ, and VGK against local market benchmarks.

Excel export: [Download screener workbook](/download_screener.xlsx)

### Stocks
Compact ranked view:

| Ticker | Company | Exchange | Price | Market Cap | P/E | P/B | Overvaluation Score | P/E + P/B? |
| --- | --- | --- | --- | --- | --- | --- | --- | --- |
{screener_rows_markdown(rows)}

### Stocks: Full Details
{stock_detail_markdown(rows)}

### Markets
Country and western benchmark comparison. Live Yahoo/Stooq data is preferred; fallback baseline rows are labeled in the data audit:

| Market | Ticker | Region | Latest Price | 1Y Return | 3Y Return | Pulled-Period Return | Volatility | Max Drawdown | Trend |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
{market_rows_markdown(market_rows, years)}

### Markets: Data Audit
{market_detail_markdown(market_rows)}
"""


def screener_ai_summary(country: str, rows: list[dict[str, object]], years: int) -> str:
    if not rows:
        return "Yahoo did not return usable fact-check rows for this country run. Try a different country or ticker universe, then verify manually on Yahoo Finance and the local exchange disclosure site."
    verified = [row for row in rows if row.get("P/E and P/B Available") == "Yes"]
    missing = [row for row in rows if row.get("P/E and P/B Available") != "Yes"]
    top = rows[:5]
    packet = [
        {
            "ticker": row.get("Ticker"),
            "name": row.get("Name"),
            "price": row.get("Price"),
            "pe": row.get("Trailing PE"),
            "pb": row.get("Price To Book"),
            "liquidity": row.get("Liquidity Value"),
            "return_3y": row.get("3Y Return %"),
            "volatility": row.get("Volatility %"),
            "drawdown": row.get("Max Drawdown %"),
            "score": row.get("Overvaluation Score"),
            "reasons": row.get("Overvaluation Reasons"),
            "fact_check": row.get("Yahoo Fact Check"),
        }
        for row in top
    ]
    api_key = os.getenv("OPENAI_API_KEY")
    if api_key:
        try:
            from openai import OpenAI  # type: ignore

            prompt = f"""
You are writing the analysis summary for an overvalued-country stock screener.
Use only the Yahoo-checked screener rows below. Do not invent S&P Global, broker borrow, or filing data.
Explain which names look most overvalued, what Yahoo verified, what Yahoo did not verify, how this compares to western markets through SPY, and what diligence should be checked next.
Keep it practical and investor-style.

Country: {country}
Period: {years} years
Verified P/E and P/B rows: {len(verified)}
Missing P/E or P/B rows: {len(missing)}
Rows:
{json.dumps(packet, indent=2)}
"""
            client = OpenAI(api_key=api_key)
            response = client.responses.create(model=os.getenv("OPENAI_MODEL", "gpt-4.1-mini"), input=prompt)
            return clean_text(response.output_text)
        except Exception as exc:
            return f"AI screener summary unavailable ({exc}). " + deterministic_screener_summary(country, rows, years, verified, missing)
    return deterministic_screener_summary(country, rows, years, verified, missing)


def deterministic_screener_summary(country: str, rows: list[dict[str, object]], years: int, verified: list[dict[str, object]], missing: list[dict[str, object]]) -> str:
    top = rows[:5]
    leader = top[0] if top else {}
    parts = [
        f"{country} screen completed across the supported liquid universe for a {years}-year period.",
        f"Yahoo verified both P/E and P/B for {len(verified)} of {len(rows)} rows; {len(missing)} rows need manual valuation verification before ranking.",
    ]
    if leader:
        parts.append(
            f"The highest-ranked current candidate is {leader.get('Ticker')} ({leader.get('Name')}) with score {leader.get('Overvaluation Score')}; main reason: {leader.get('Overvaluation Reasons')}."
        )
    parts.append("Use the Yahoo fact-check column as the first audit trail, then verify local filings, borrow availability, local market liquidity, currency exposure, and any licensed S&P Global fundamentals you can access.")
    return " ".join(parts)


def screener_workbook_bytes(rows: list[dict[str, object]], country: str) -> bytes:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except Exception as exc:
        csv_lines = ["Openpyxl unavailable; install requirements.txt. Error: " + str(exc)]
        return "\n".join(csv_lines).encode("utf-8")

    def safe_rows(value: object) -> list[dict[str, object]]:
        return [row for row in value if isinstance(row, dict)] if isinstance(value, list) else []

    def append_table(ws, headers: list[str], table_rows: list[dict[str, object]], start_row: int = 1) -> int:
        ws.append(headers)
        for row in table_rows:
            ws.append([row.get(header, "") for header in headers])
        end_row = start_row + len(table_rows)
        style_sheet(ws, headers, start_row, end_row)
        return end_row

    def style_sheet(ws, headers: list[str], header_row: int = 1, end_row: int | None = None) -> None:
        header_fill = PatternFill("solid", fgColor="DDEFE9")
        dark_fill = PatternFill("solid", fgColor="143F36")
        thin = Side(style="thin", color="D8DFDC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        end_row = end_row or ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="16211F")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = border
        for row in ws.iter_rows(min_row=header_row + 1, max_row=end_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
        for idx, header in enumerate(headers, start=1):
            width = 14
            if header in {"Name", "Company", "Products / What They Sell", "Overvaluation Reasons", "Shortability / Liquidity Read", "Yahoo Fact Check", "Fact Check", "Interpretation", "What It Means"}:
                width = 34
            elif header in {"Ticker", "Currency", "P/E + P/B?", "Rank"}:
                width = 12
            elif header in {"Market Cap", "Liquidity Value", "Avg Volume"}:
                width = 18
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = f"A{header_row + 1}"
        if ws.max_column and ws.max_row >= header_row:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{end_row}"
        if ws.max_row >= header_row:
            ws.row_dimensions[header_row].height = 28
        if ws.title in {"Read Me", "Executive Summary"}:
            for cell in ws[1]:
                cell.fill = dark_fill
                cell.font = Font(bold=True, color="FFFFFF")

    stock_rows = safe_rows(rows)
    market_rows = safe_rows(SCREENER_STATE.get("market_rows", []))
    verified_count = sum(1 for row in stock_rows if row.get("P/E and P/B Available") == "Yes")
    fallback_count = sum(1 for row in stock_rows if "fallback" in clean_text(str(row.get("Yahoo Fact Check", ""))).lower())
    top_rows = sorted(stock_rows, key=lambda row: float(row.get("Overvaluation Score") or 0), reverse=True)[:5]

    wb = Workbook()
    ws = wb.active
    ws.title = "Read Me"
    readme_rows = [
        ["Workbook", "Advanced Country Overvaluation Screener"],
        ["Country", country or "Not selected"],
        ["How to use", "Start with Executive Summary, then review Stock Screen for ranked ideas, Stock Detail for business/valuation context, and Market Screen for country vs western-market comparison."],
        ["Data order", "Yahoo live data first, Stooq historical fallback second, then clearly labeled baseline fallback when free endpoints are rate-limited."],
        ["Important", "Fallback baseline rows are useful for first-pass screening only. Verify every market, valuation, and shortability field before using it in an investment memo."],
    ]
    for row in readme_rows:
        ws.append(row)
    style_sheet(ws, ["Topic", "Explanation"], 1, len(readme_rows))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    summary = wb.create_sheet("Executive Summary")
    summary_rows = [
        ["Metric", "Value", "What It Means"],
        ["Country screened", country or "Not selected", "Selected country universe for the overvaluation screen."],
        ["Stocks screened", len(stock_rows), "Number of supported liquid consumer/consumables names reviewed."],
        ["Rows with P/E and P/B", verified_count, "Rows where valuation fields are available from live or fallback sources."],
        ["Rows using fallback baseline", fallback_count, "Rows that need manual verification because free live endpoints were blocked or incomplete."],
        ["Highest score", max([float(row.get("Overvaluation Score") or 0) for row in stock_rows], default=0), "Higher score means more overvaluation / short-candidate warning signals."],
        ["Market benchmarks", len(market_rows), "Country and western-market benchmark rows included in Market Screen."],
    ]
    for row in summary_rows:
        summary.append(row)
    style_sheet(summary, ["Metric", "Value", "What It Means"], 1, len(summary_rows))
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 20
    summary.column_dimensions["C"].width = 75

    top = wb.create_sheet("Top Candidates")
    top_headers = ["Rank", "Ticker", "Company", "Price", "P/E", "P/B", "Volatility %", "Max Drawdown %", "Score", "Why It Matters", "Data Status"]
    top_table = []
    for idx, row in enumerate(top_rows, 1):
        top_table.append(
            {
                "Rank": idx,
                "Ticker": row.get("Ticker"),
                "Company": row.get("Name"),
                "Price": row.get("Price"),
                "P/E": row.get("Trailing PE"),
                "P/B": row.get("Price To Book"),
                "Volatility %": row.get("Volatility %"),
                "Max Drawdown %": row.get("Max Drawdown %"),
                "Score": row.get("Overvaluation Score"),
                "Why It Matters": row.get("Overvaluation Reasons"),
                "Data Status": row.get("Yahoo Fact Check"),
            }
        )
    append_table(top, top_headers, top_table)

    stock = wb.create_sheet("Stock Screen")
    stock_headers = [
        "Ticker", "Name", "Exchange", "Currency", "Price", "Market Cap", "Avg Volume", "Liquidity Value",
        "Trailing PE", "Forward PE", "Price To Book", "Dividend Yield %", "Beta", "1Y Return %",
        "3Y Return %", "Volatility %", "Max Drawdown %", "Overvaluation Score", "P/E and P/B Available",
        "Overvaluation Reasons", "Shortability / Liquidity Read", "Yahoo Fact Check", "Yahoo URL",
    ]
    append_table(stock, stock_headers, stock_rows)

    detail = wb.create_sheet("Stock Detail")
    detail_headers = ["Ticker", "Company", "Sector", "Industry", "Products / What They Sell", "Valuation Read", "Risk Read", "Next Checks"]
    detail_rows = []
    for row in stock_rows:
        detail_rows.append(
            {
                "Ticker": row.get("Ticker"),
                "Company": row.get("Name"),
                "Sector": row.get("Sector"),
                "Industry": row.get("Industry"),
                "Products / What They Sell": row.get("Products / What They Sell"),
                "Valuation Read": f"P/E {fmt_ratio(yahoo_float(row.get('Trailing PE')))}, P/B {fmt_ratio(yahoo_float(row.get('Price To Book')))}, score {row.get('Overvaluation Score')}. {row.get('Overvaluation Reasons')}",
                "Risk Read": f"Volatility {fmt_pct_plain(yahoo_float(row.get('Volatility %')))}, drawdown {fmt_pct_plain(yahoo_float(row.get('Max Drawdown %')))}. {row.get('Shortability / Liquidity Read')}",
                "Next Checks": "Verify local filings, current Yahoo/market data, borrow availability, currency exposure, and licensed fundamentals if available.",
            }
        )
    append_table(detail, detail_headers, detail_rows)

    market = wb.create_sheet("Market Screen")
    market_headers = list(market_rows[0].keys()) if market_rows else [
        "Market", "Ticker", "Region", "Currency", "Latest Price", "1Y Return %", "3Y Return %",
        "Volatility %", "Max Drawdown %", "Trend", "Fact Check", "Yahoo URL",
    ]
    append_table(market, market_headers, market_rows)

    dictionary = wb.create_sheet("Data Dictionary")
    dictionary_rows = [
        ["Column", "Plain-English Meaning"],
        ["Overvaluation Score", "A first-pass warning score based on P/E, P/B, volatility, drawdown, and benchmark underperformance when available."],
        ["P/E", "Price divided by earnings. High P/E can mean optimism, quality, scarcity, or overvaluation."],
        ["P/B", "Price divided by book value. Most useful for asset-heavy businesses; less decisive for asset-light brands."],
        ["Liquidity Value", "Approximate dollar value traded, calculated as price times average volume when available."],
        ["Max Drawdown", "Largest peak-to-trough decline in the pulled period. Bigger negative numbers mean larger historical downside."],
        ["Volatility", "Annualized price movement estimate. Higher volatility means a wider range of possible outcomes."],
        ["Fallback baseline", "A clearly labeled non-live fallback used only when free providers are rate-limited or incomplete."],
    ]
    for row in dictionary_rows:
        dictionary.append(row)
    style_sheet(dictionary, ["Column", "Plain-English Meaning"], 1, len(dictionary_rows))
    dictionary.column_dimensions["A"].width = 28
    dictionary.column_dimensions["B"].width = 95

    note = wb.create_sheet("Notes")
    note_rows = [
        ["Country", country or "Not selected"],
        ["Data source", "Yahoo Finance public endpoints first; Stooq historical fallback; labeled baseline fallback if free providers are blocked."],
        ["S&P Global", "Licensed source. Add your own S&P Global / Capital IQ / Compustat export or API integration if you have rights."],
        ["Shortability", "Liquidity is an approximation using price x average volume. Borrow availability, fees, locates, market hours, and local rules must be verified with broker."],
        ["Not investment advice", "Educational screener only."],
    ]
    for row in note_rows:
        note.append(row)
    style_sheet(note, ["Topic", "Note"], 1, len(note_rows))
    note.column_dimensions["A"].width = 24
    note.column_dimensions["B"].width = 95

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    header = str(sheet.cell(row=1, column=cell.column).value or "")
                    if "%" in header or "Return" in header or "Drawdown" in header or "Volatility" in header:
                        cell.number_format = '0.0'
                    elif "P/E" in header or "P/B" in header or "Beta" in header or "Score" in header:
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '#,##0.0'
        if sheet.max_row > 1 and sheet.max_column > 1:
            score_col = None
            for cell in sheet[1]:
                if cell.value in {"Overvaluation Score", "Score"}:
                    score_col = cell.column
                    break
            if score_col:
                col = get_column_letter(score_col)
                sheet.conditional_formatting.add(
                    f"{col}2:{col}{sheet.max_row}",
                    CellIsRule(operator="greaterThanOrEqual", formula=["4"], fill=PatternFill("solid", fgColor="F6D7D2")),
                )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def research_workbook_bytes(rows: list[dict[str, object]], assumptions: dict[str, object], parse_note: str, missing: list[str]) -> bytes:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import CellIsRule
    except Exception as exc:
        return f"Openpyxl unavailable; install requirements.txt. Error: {exc}".encode("utf-8")

    safe_rows = [row for row in rows if isinstance(row, dict)]
    safe_missing = [clean_text(str(item)) for item in missing if clean_text(str(item))]

    def style_sheet(ws, header_row: int = 1, end_row: int | None = None) -> None:
        header_fill = PatternFill("solid", fgColor="DDEFE9")
        dark_fill = PatternFill("solid", fgColor="143F36")
        warn_fill = PatternFill("solid", fgColor="F6D7D2")
        thin = Side(style="thin", color="D8DFDC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        end_row = end_row or ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="16211F")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = border
        if ws.title in {"Read Me", "Executive Summary"}:
            for cell in ws[header_row]:
                cell.fill = dark_fill
                cell.font = Font(bold=True, color="FFFFFF")
        for row in ws.iter_rows(min_row=header_row + 1, max_row=end_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
        for idx, cell in enumerate(ws[header_row], start=1):
            header = clean_text(str(cell.value or ""))
            width = 15
            if header in {"Company", "Data Source", "Interpretation", "Next Check", "Plain-English Meaning", "Note"}:
                width = 36
            elif header in {"Collateral Surplus / Shortfall", "Current Collateral", "Net After Loan", "Loan Payoff"}:
                width = 22
            elif header in {"Input Ticker", "Resolved Ticker", "Exchange"}:
                width = 18
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = f"A{header_row + 1}"
        if ws.max_column and ws.max_row >= header_row:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{end_row}"
        for row in ws.iter_rows(min_row=header_row + 1, max_row=end_row):
            for cell in row:
                header = clean_text(str(ws.cell(row=header_row, column=cell.column).value or ""))
                if isinstance(cell.value, (int, float)):
                    if "%" in header or "Drop" in header or "LTV" in header or "Volatility" in header:
                        cell.number_format = '0.0'
                    elif "Shares" in header:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'
                if header == "Plunged?" and cell.value == "Yes":
                    cell.fill = warn_fill

    def append_table(ws, headers: list[str], table_rows: list[dict[str, object]]) -> None:
        ws.append(headers)
        for row in table_rows:
            ws.append([row.get(header, "") for header in headers])
        style_sheet(ws, 1, ws.max_row)

    flagged = [row for row in safe_rows if row.get("flag") == "Yes"]
    worst = safe_rows[0] if safe_rows else {}
    wb = Workbook()
    readme = wb.active
    readme.title = "Read Me"
    for row in [
        ["Workbook", "Collateral Plunge Research Export"],
        ["Purpose", "Screens ticker/date rows for pledged-share collateral drawdowns, loan coverage, and call/put overlay economics."],
        ["Upload parsing", parse_note],
        ["Data order", "Yahoo Finance live data first, Stooq fallback second, labeled baseline fallback where available."],
        ["Important", "This is a screening workbook. Verify exchange prices, FX, pledge agreements, exact share count, option chains, and legal restrictions before use."],
    ]:
        readme.append(row)
    style_sheet(readme, 1, readme.max_row)
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 100

    summary = wb.create_sheet("Executive Summary")
    summary_rows = [
        ["Metric", "Value", "Interpretation"],
        ["Rows with usable market data", len(safe_rows), "Rows where the app produced collateral math."],
        ["Plunge flags", len(flagged), "Rows at or below the selected drop threshold."],
        ["Missing / unresolved tickers", len(safe_missing), "Tickers that need manual symbol or data verification."],
        ["Worst ticker", worst.get("resolved_ticker") or worst.get("ticker") or "None", "Largest price drop in the processed output."],
        ["Worst price drop %", worst.get("drop", ""), "Negative number means price declined from the loan-date week median."],
        ["Worst collateral surplus / shortfall", worst.get("shortfall", ""), "Negative number suggests the modeled collateral is below payoff before FX/legal checks."],
    ]
    for row in summary_rows:
        summary.append(row)
    style_sheet(summary, 1, summary.max_row)
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 24
    summary.column_dimensions["C"].width = 80

    assumptions_ws = wb.create_sheet("Assumptions")
    assumptions_ws.append(["Assumption", "Value", "Note"])
    assumption_notes = {
        "loan_amount": "Principal amount.",
        "ltv_pct": "Loan-to-value percentage used to infer pledged collateral value.",
        "collateral_value": "Loan amount divided by LTV.",
        "interest_rate_pct": "Simple annual interest assumption.",
        "loan_years": "Loan/option model term.",
        "loan_payoff": "Principal plus simple interest.",
        "volatility_pct": "Black-Scholes volatility input.",
        "call_strike_pct_of_today_price": "Buy-call strike as percent of current underlying price.",
        "put_strike_pct_of_today_price": "Sell-put strike as percent of current underlying price.",
        "market_hint": "Used to resolve plain international tickers to Yahoo suffixes.",
    }
    for key, value in assumptions.items():
        assumptions_ws.append([key, value, assumption_notes.get(key, "")])
    style_sheet(assumptions_ws, 1, assumptions_ws.max_row)
    assumptions_ws.column_dimensions["A"].width = 34
    assumptions_ws.column_dimensions["B"].width = 24
    assumptions_ws.column_dimensions["C"].width = 80

    screen = wb.create_sheet("Collateral Screen")
    screen_rows = []
    for row in safe_rows:
        screen_rows.append(
            {
                "Input Ticker": row.get("ticker"),
                "Resolved Ticker": row.get("resolved_ticker"),
                "Company": row.get("company"),
                "Exchange": row.get("exchange"),
                "Currency": row.get("currency"),
                "Loan Date": row.get("date"),
                "Week Median Price": row.get("start"),
                "Today Price": row.get("today"),
                "Price Drop %": row.get("drop"),
                "Pledged Shares": row.get("shares"),
                "Current Collateral": row.get("current_collateral"),
                "Loan Payoff": row.get("loan_payoff"),
                "Collateral Surplus / Shortfall": row.get("shortfall"),
                "Net Sale Payout / Share": row.get("net_payout"),
                "Net After Loan": row.get("net_after_loan"),
                "Data Source": row.get("data_status"),
                "Plunged?": row.get("flag"),
            }
        )
    append_table(screen, list(screen_rows[0].keys()) if screen_rows else ["Input Ticker", "Resolved Ticker", "Company", "Exchange", "Currency", "Loan Date", "Week Median Price", "Today Price", "Price Drop %", "Pledged Shares", "Current Collateral", "Loan Payoff", "Collateral Surplus / Shortfall", "Net Sale Payout / Share", "Net After Loan", "Data Source", "Plunged?"], screen_rows)

    overlay = wb.create_sheet("Option Overlay")
    overlay_rows = []
    for row in safe_rows:
        overlay_rows.append(
            {
                "Resolved Ticker": row.get("resolved_ticker"),
                "Currency": row.get("currency"),
                "Today Price": row.get("today"),
                "Buy Call Cost / Share": row.get("call"),
                "Sell Put Credit / Share": row.get("put"),
                "Net Option Cost / Share": row.get("net_option"),
                "Net Sale Payout / Share": row.get("net_payout"),
                "Pledged Shares": row.get("shares"),
                "Net After Loan": row.get("net_after_loan"),
                "Interpretation": "Negative net after loan suggests collateral/hedge economics need urgent review." if float(row.get("net_after_loan") or 0) < 0 else "Modeled net proceeds cover payoff before FX/legal verification.",
            }
        )
    append_table(overlay, list(overlay_rows[0].keys()) if overlay_rows else ["Resolved Ticker", "Currency", "Today Price", "Buy Call Cost / Share", "Sell Put Credit / Share", "Net Option Cost / Share", "Net Sale Payout / Share", "Pledged Shares", "Net After Loan", "Interpretation"], overlay_rows)

    missing_ws = wb.create_sheet("Missing Data")
    missing_ws.append(["Ticker / Attempt", "Next Check"])
    for item in safe_missing:
        missing_ws.append([item, "Confirm Yahoo ticker suffix, local exchange code, trading status, and whether shares had corporate actions."])
    if not safe_missing:
        missing_ws.append(["None", "All uploaded rows produced a screen row."])
    style_sheet(missing_ws, 1, missing_ws.max_row)
    missing_ws.column_dimensions["A"].width = 45
    missing_ws.column_dimensions["B"].width = 95

    dictionary = wb.create_sheet("Data Dictionary")
    dictionary_rows = [
        ["Column", "Plain-English Meaning"],
        ["Input Ticker", "Ticker value from the Excel upload."],
        ["Resolved Ticker", "Ticker the app used after applying international Yahoo suffix logic."],
        ["Week Median Price", "Median share price in the week around the loan/collateral date, or nearest prior/closest price if needed."],
        ["Current Collateral", "Modeled pledged shares multiplied by today's pulled share price."],
        ["Loan Payoff", "Loan principal plus simple interest for the entered term."],
        ["Collateral Surplus / Shortfall", "Current collateral less loan payoff. FX must be verified for non-USD shares."],
        ["Buy Call Cost / Share", "Black-Scholes screening estimate for bought call protection."],
        ["Sell Put Credit / Share", "Black-Scholes screening estimate for sold put financing."],
        ["Net Option Cost / Share", "Call cost less put credit."],
        ["Net Sale Payout / Share", "Today price less net option cost."],
        ["Plunged?", "Yes when the price drop is at or below the selected plunge threshold."],
    ]
    for row in dictionary_rows:
        dictionary.append(row)
    style_sheet(dictionary, 1, dictionary.max_row)
    dictionary.column_dimensions["A"].width = 32
    dictionary.column_dimensions["B"].width = 95

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        if sheet.max_row > 1 and sheet.max_column > 1:
            for cell in sheet[1]:
                if cell.value == "Price Drop %":
                    col = get_column_letter(cell.column)
                    sheet.conditional_formatting.add(f"{col}2:{col}{sheet.max_row}", CellIsRule(operator="lessThanOrEqual", formula=["-30"], fill=PatternFill("solid", fgColor="F6D7D2")))
                if cell.value in {"Collateral Surplus / Shortfall", "Net After Loan"}:
                    col = get_column_letter(cell.column)
                    sheet.conditional_formatting.add(f"{col}2:{col}{sheet.max_row}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="F6D7D2")))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_amount(value: str, default: float = 1000.0) -> float:
    cleaned = re.sub(r"[^0-9.]", "", value or "")
    try:
        amount = float(cleaned)
    except Exception:
        amount = default
    return min(max(amount, 0.0), 1_000_000_000.0)


def parse_float(value: str, default: float = 0.0, minimum: float | None = None, maximum: float | None = None) -> float:
    cleaned = re.sub(r"[^0-9.+-]", "", value or "")
    try:
        number = float(cleaned)
    except Exception:
        number = default
    if minimum is not None:
        number = max(number, minimum)
    if maximum is not None:
        number = min(number, maximum)
    return number


def future_value(amount: float, return_pct: float) -> float:
    return amount * (1 + return_pct / 100)


def parse_research_date(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and 1 <= float(value) <= 80000:
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(value))
        except Exception:
            return None
    if value is None:
        return None
    text = clean_text(str(value))
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    try:
        return datetime.fromisoformat(text[:10])
    except Exception:
        return None


def research_ticker_text(value: object) -> str:
    text = clean_text(str(value or "")).upper()
    text = text.replace(" ", "")
    if re.fullmatch(r"[A-Z]{1,6}(?:[-.][A-Z0-9]{1,4})?", text):
        return text
    if re.fullmatch(r"\d{1,6}(?:\.[A-Z]{1,4})?", text):
        return text
    return ""


def research_symbol_candidates(symbol: str, market_hint: str = "Global") -> list[str]:
    symbol = research_ticker_text(symbol)
    if not symbol:
        return []
    if "." in symbol or symbol.startswith("^"):
        return [symbol]
    hint = clean_text(market_hint) or "Global"
    candidates = [symbol]
    if re.fullmatch(r"\d{6}", symbol):
        if hint == "China":
            candidates.extend([symbol + ".SS", symbol + ".SZ"])
        elif hint == "South Korea":
            candidates.append(symbol + ".KS")
        elif hint == "Taiwan":
            candidates.extend([symbol + ".TW", symbol + ".TWO"])
        else:
            candidates.extend([symbol + ".SS", symbol + ".SZ", symbol + ".KS", symbol + ".TW", symbol + ".TWO"])
    if re.fullmatch(r"\d{4}", symbol):
        hk_symbol = symbol.zfill(4) + ".HK"
        if hint == "Japan":
            candidates.append(symbol + ".T")
        elif hint == "Hong Kong":
            candidates.append(hk_symbol)
        else:
            candidates.extend([hk_symbol, symbol + ".T"])
    if re.fullmatch(r"\d{1,3}", symbol):
        candidates.append(symbol.zfill(4) + ".HK")
    suffix_by_hint = {
        "United Kingdom": ".L",
        "Canada": ".TO",
        "Australia": ".AX",
        "Germany": ".DE",
        "France": ".PA",
        "India": ".NS",
    }
    suffix = suffix_by_hint.get(hint)
    if suffix and re.fullmatch(r"[A-Z]{1,8}", symbol):
        candidates.append(symbol + suffix)
    out: list[str] = []
    for candidate in candidates:
        if candidate not in out:
            out.append(candidate)
    return out


def fallback_research_series(symbol: str, loan_date: datetime) -> StockSeries | None:
    baseline = market_baseline(symbol)
    if not baseline:
        return None
    current = baseline_float(symbol, "price")
    if not current:
        return None
    lookback_return = baseline_float(symbol, "return_3y")
    if lookback_return is None:
        lookback_return = baseline_float(symbol, "return_period")
    start = current / (1 + lookback_return / 100) if lookback_return is not None and lookback_return > -95 else current
    profile = market_profile(symbol)
    return StockSeries(
        ticker=symbol,
        currency=first_text(profile.get("currency"), "USD"),
        points=[StockPoint(loan_date, start), StockPoint(datetime.now(), current)],
        source_url="fallback baseline profile; verify against live exchange data",
        name=first_text(profile.get("name"), symbol),
        exchange=first_text(profile.get("exchange"), "International market"),
        regular_market_price=current,
        average_volume=baseline_float(symbol, "avg_volume"),
        market_cap=baseline_float(symbol, "market_cap"),
        trailing_pe=baseline_float(symbol, "pe"),
        forward_pe=baseline_float(symbol, "forward_pe"),
        price_to_book=baseline_float(symbol, "pb"),
        dividend_yield_pct=baseline_float(symbol, "dividend_yield"),
        beta=baseline_float(symbol, "beta"),
        sector=first_text(profile.get("sector")),
        industry=first_text(profile.get("industry")),
        business_summary=first_text(profile.get("products")),
    )


def fetch_research_stock_series(symbol: str, years: int, loan_date: datetime, market_hint: str) -> tuple[StockSeries | None, list[str], str]:
    candidates = research_symbol_candidates(symbol, market_hint)
    attempted: list[str] = []
    for candidate in candidates:
        attempted.append(candidate)
        series = fetch_stock_series(candidate, years)
        if series:
            return series, attempted, "Live Yahoo/Stooq price history"
    for candidate in candidates:
        series = fallback_research_series(candidate, loan_date)
        if series:
            return series, attempted, "Fallback baseline profile; verify before use"
    return None, attempted, "No usable Yahoo/Stooq/baseline data"


def infer_research_columns(rows: list[tuple[object, ...]]) -> tuple[int | None, int | None, int | None, int]:
    if not rows:
        return None, None, None, 0
    header_values = [clean_text(str(cell or "")).lower() for cell in rows[0]]
    ticker_idx = next((i for i, header in enumerate(header_values) if header in {"ticker", "symbol", "stock", "stock ticker"}), None)
    date_idx = next((i for i, header in enumerate(header_values) if header in {"date", "loan date", "pledge date", "collateral date", "as of date"}), None)
    company_idx = next((i for i, header in enumerate(header_values) if header in {"company", "company name", "name"}), None)
    if ticker_idx is not None and date_idx is not None:
        return ticker_idx, date_idx, company_idx, 1
    max_cols = max((len(row) for row in rows[:25]), default=0)
    ticker_scores: dict[int, int] = {}
    date_scores: dict[int, int] = {}
    for row in rows[:25]:
        for idx in range(max_cols):
            cell = row[idx] if idx < len(row) else None
            if research_ticker_text(cell):
                ticker_scores[idx] = ticker_scores.get(idx, 0) + 1
            if parse_research_date(cell):
                date_scores[idx] = date_scores.get(idx, 0) + 1
    inferred_ticker = max(ticker_scores, key=ticker_scores.get) if ticker_scores else None
    inferred_date = max((idx for idx in date_scores if idx != inferred_ticker), key=lambda idx: date_scores[idx], default=None)
    inferred_company = None
    for idx in range(max_cols):
        if idx not in {inferred_ticker, inferred_date}:
            inferred_company = idx
            break
    header_like = any(value in {"ticker", "symbol", "date", "loan date", "company", "name"} for value in header_values)
    return inferred_ticker, inferred_date, inferred_company, 1 if header_like else 0


def normal_cdf(value: float) -> float:
    return 0.5 * (1.0 + math.erf(value / math.sqrt(2.0)))


def black_scholes_price(spot: float, strike: float, years: float, rate_pct: float, volatility_pct: float, option_type: str) -> float:
    if spot <= 0 or strike <= 0 or years <= 0 or volatility_pct <= 0:
        return 0.0
    rate = rate_pct / 100
    volatility = volatility_pct / 100
    d1 = (math.log(spot / strike) + (rate + 0.5 * volatility * volatility) * years) / (volatility * math.sqrt(years))
    d2 = d1 - volatility * math.sqrt(years)
    if option_type.lower() == "put":
        return strike * math.exp(-rate * years) * normal_cdf(-d2) - spot * normal_cdf(-d1)
    return spot * normal_cdf(d1) - strike * math.exp(-rate * years) * normal_cdf(d2)


def weekly_median_price(series: StockSeries, loan_date: datetime) -> float | None:
    if not series.points:
        return None
    window = [point.close for point in series.points if abs((point.date - loan_date).days) <= 7 and point.close > 0]
    if window:
        return statistics.median(window)
    earlier = [point for point in series.points if point.date <= loan_date and point.close > 0]
    if earlier:
        return earlier[-1].close
    closest = min(series.points, key=lambda point: abs((point.date - loan_date).days), default=None)
    return closest.close if closest else None


def excel_research_rows(files: list[tuple[str, bytes]]) -> tuple[list[dict[str, object]], str]:
    upload = next(((filename, raw) for filename, raw in files if filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")) and raw), None)
    if not upload:
        return [], "No Excel workbook was uploaded. Upload a spreadsheet with ticker/date columns. Headers are optional."
    filename, raw = upload
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        return [], f"Excel parsing unavailable because openpyxl is not installed: {exc}"
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        return [], f"Excel workbook could not be read: {exc}"
    if not rows:
        return [], f"{filename} was empty."
    ticker_idx, date_idx, company_idx, start_row = infer_research_columns(rows)
    if ticker_idx is None or date_idx is None:
        preview = []
        for row in rows[:5]:
            preview.append(" / ".join(clean_text(str(cell or "")) for cell in row[:6]))
        return [], f"{filename} needs ticker/date values. The app could not infer the columns from the first rows: {'; '.join(preview)}"
    parsed: list[dict[str, object]] = []
    for row in rows[start_row:151]:
        ticker = research_ticker_text(row[ticker_idx] if ticker_idx < len(row) else None)
        loan_date = parse_research_date(row[date_idx] if date_idx < len(row) else None)
        company = clean_text(str(row[company_idx] or "")) if company_idx is not None and company_idx < len(row) else ""
        if ticker and loan_date:
            parsed.append({"ticker": ticker, "date": loan_date, "company": company})
    column_note = "header row detected" if start_row else "columns inferred from values; no header row required"
    return parsed, f"Loaded {len(parsed)} ticker/date rows from {filename} ({column_note})."


def ai_collateral_research_read(rows: list[dict[str, object]], assumptions: dict[str, object], parse_note: str, missing: list[str]) -> str:
    if not rows:
        return (
            "No usable ticker/date rows were processed. Check that the spreadsheet has one column with ticker symbols and one column with dates. "
            "Headers are optional now; the app will infer the columns from values."
        )
    worst = rows[:5]
    flagged = [row for row in rows if row.get("flag") == "Yes"]
    packet = {
        "parse_note": parse_note,
        "assumptions": assumptions,
        "worst_rows": worst,
        "flagged_count": len(flagged),
        "missing": missing[:20],
    }
    api_key = os.getenv("OPENAI_API_KEY")
    if api_key:
        try:
            from openai import OpenAI  # type: ignore

            prompt = f"""
You are a credit hedge research analyst. Write a concise but useful readout for a pledged-share collateral screen.
Use only the data in this packet. Do not invent live prices or legal facts. Explain which tickers look most stressed, why the collateral may be impaired, how to interpret the call/put overlay, and what must be verified next.
Packet:
{json.dumps(packet, default=str, indent=2)}
"""
            client = OpenAI(api_key=api_key)
            response = client.responses.create(model=os.getenv("OPENAI_MODEL", "gpt-4.1-mini"), input=prompt)
            return clean_text(response.output_text)
        except Exception as exc:
            return f"AI readout unavailable ({exc}). Rules-based readout: {deterministic_collateral_research_read(rows, assumptions, missing)}"
    return deterministic_collateral_research_read(rows, assumptions, missing)


def deterministic_collateral_research_read(rows: list[dict[str, object]], assumptions: dict[str, object], missing: list[str]) -> str:
    flagged = [row for row in rows if row.get("flag") == "Yes"]
    worst = rows[0] if rows else None
    if not worst:
        return "No usable rows were processed."
    read = (
        f"The worst screened row is {worst.get('ticker')} with a price move of {pct(float(worst.get('drop') or 0))} from the loan-date week median to today's pulled price. "
        f"{len(flagged)} row(s) crossed the plunge threshold. "
    )
    if float(worst.get("shortfall") or 0) < 0:
        read += "The worst row shows a collateral shortfall versus the modeled loan payoff, so it should be prioritized for agreement review, margin-call rights, FX conversion, and liquidity checks. "
    else:
        read += "The worst row still shows modeled collateral surplus versus payoff, but the drawdown and option overlay should still be checked against live prices. "
    read += "The option overlay is a screening estimate, not a live options-chain quote. "
    if missing:
        read += f"Market data was missing for: {', '.join(missing[:8])}."
    return read


def build_collateral_plunge_research(fields: dict[str, str], files: list[tuple[str, bytes]]) -> str:
    entries, parse_note = excel_research_rows(files)
    loan_amount = parse_amount(fields.get("loan_amount", "100000000"), 100_000_000.0)
    ltv = parse_float(fields.get("ltv", "50"), 50.0, 1.0, 95.0)
    interest_rate = parse_float(fields.get("interest_rate", "5"), 5.0, 0.0, 100.0)
    loan_years = parse_float(fields.get("loan_years", "3"), 3.0, 0.1, 30.0)
    volatility = parse_float(fields.get("volatility", "30"), 30.0, 1.0, 200.0)
    drop_threshold = abs(parse_float(fields.get("drop_threshold", "30"), 30.0, 0.0, 95.0))
    put_strike_pct = parse_float(fields.get("put_strike_pct", "60"), 60.0, 1.0, 200.0)
    call_strike_pct = parse_float(fields.get("call_strike_pct", "100"), 100.0, 1.0, 300.0)
    market_hint = clean_text(fields.get("research_market", "Global")) or "Global"
    collateral_value = loan_amount / (ltv / 100) if ltv else loan_amount * 2
    loan_payoff = loan_amount * (1 + interest_rate / 100 * loan_years)
    rows = []
    flagged = 0
    missing = []
    for entry in entries[:80]:
        ticker = str(entry["ticker"])
        loan_date = entry["date"]
        company = str(entry.get("company") or ticker)
        if not isinstance(loan_date, datetime):
            missing.append(ticker)
            continue
        series, attempted_symbols, data_status = fetch_research_stock_series(ticker, max(int(math.ceil(loan_years)) + 1, 5), loan_date, market_hint)
        if not series:
            missing.append(f"{ticker} (tried {', '.join(attempted_symbols)})")
            continue
        start_price = weekly_median_price(series, loan_date) if isinstance(loan_date, datetime) else None
        today_price = series.regular_market_price or (series.points[-1].close if series.points else None)
        if not start_price or not today_price:
            missing.append(ticker)
            continue
        pledged_shares = collateral_value / start_price
        current_collateral = pledged_shares * today_price
        drop_pct = (today_price / start_price - 1) * 100
        collateral_shortfall = current_collateral - loan_payoff
        call_strike = today_price * call_strike_pct / 100
        put_strike = today_price * put_strike_pct / 100
        call_cost = black_scholes_price(today_price, call_strike, loan_years, interest_rate, volatility, "call")
        put_credit = black_scholes_price(today_price, put_strike, loan_years, interest_rate, volatility, "put")
        net_option_cost = call_cost - put_credit
        net_sale_payout_per_share = today_price - net_option_cost
        net_sale_payout_total = pledged_shares * net_sale_payout_per_share
        net_after_loan = net_sale_payout_total - loan_payoff
        is_plunged = drop_pct <= -drop_threshold
        if is_plunged:
            flagged += 1
        rows.append(
            {
                "ticker": ticker,
                "resolved_ticker": series.ticker,
                "company": series.name or company,
                "currency": series.currency or "USD",
                "exchange": series.exchange or "Not available",
                "data_status": data_status,
                "date": loan_date.strftime("%Y-%m-%d") if isinstance(loan_date, datetime) else "",
                "start": start_price,
                "today": today_price,
                "drop": drop_pct,
                "shares": pledged_shares,
                "current_collateral": current_collateral,
                "loan_payoff": loan_payoff,
                "shortfall": collateral_shortfall,
                "call": call_cost,
                "put": put_credit,
                "net_option": net_option_cost,
                "net_payout": net_sale_payout_per_share,
                "net_after_loan": net_after_loan,
                "flag": "Yes" if is_plunged else "No",
            }
        )
    rows.sort(key=lambda row: float(row["drop"]))
    assumptions = {
        "loan_amount": loan_amount,
        "ltv_pct": ltv,
        "collateral_value": collateral_value,
        "interest_rate_pct": interest_rate,
        "loan_years": loan_years,
        "loan_payoff": loan_payoff,
        "volatility_pct": volatility,
        "drop_threshold_pct": drop_threshold,
        "call_strike_pct_of_today_price": call_strike_pct,
        "put_strike_pct_of_today_price": put_strike_pct,
        "market_hint": market_hint,
    }
    research_read = ai_collateral_research_read(rows, assumptions, parse_note, missing)
    RESEARCH_STATE["rows"] = rows
    RESEARCH_STATE["assumptions"] = assumptions
    RESEARCH_STATE["parse_note"] = parse_note
    RESEARCH_STATE["missing"] = missing
    if rows:
        table_rows = "\n".join(
            f"| {md_cell(row['ticker'], 60)} | {md_cell(row['resolved_ticker'], 70)} | {md_cell(row['company'], 140)} | {md_cell(row['exchange'], 110)} | {md_cell(row['currency'], 40)} | {row['date']} | {money(float(row['start']), str(row['currency']))} | {money(float(row['today']), str(row['currency']))} | {pct(float(row['drop']))} | {fmt_number(float(row['shares']))} | {compact_money(float(row['current_collateral']), str(row['currency']))} | {compact_money(float(row['loan_payoff']), 'USD')} | {compact_money(float(row['shortfall']), str(row['currency']))} | {money(float(row['call']), str(row['currency']))} | {money(float(row['put']), str(row['currency']))} | {money(float(row['net_option']), str(row['currency']))} | {money(float(row['net_payout']), str(row['currency']))} | {compact_money(float(row['net_after_loan']), str(row['currency']))} | {md_cell(row['data_status'], 130)} | {row['flag']} |"
            for row in rows
        )
    else:
        table_rows = "| - | - | No usable rows | - | - | - | - | - | - | - | - | - | - | - | - | - | - | - | - | - |"
    brief = f"""# Collateral Plunge Research

## Executive Readout
This section screens an uploaded Excel list of ticker/date rows to find pledged-share collateral that has plummeted since the loan date.

{parse_note}

Loan case: {money(loan_amount, 'USD')} principal, {fmt_pct_plain(ltv)} LTV, implied pledged-share collateral of {money(collateral_value, 'USD')}, {loan_years:.1f}-year term, {fmt_pct_plain(interest_rate)} simple annual interest, and payoff of {money(loan_payoff, 'USD')}. Market hint: {market_hint}.

Rows reviewed with usable market data: {len(rows)}. Plunge flags at or below -{drop_threshold:.1f}%: {flagged}. Missing/unavailable tickers: {', '.join(missing[:12]) if missing else 'None'}.

Excel export: [Download research workbook](/download_research.xlsx)

## AI Research Readout
{research_read}

## Step 1: Collateral Share Review
The app estimates the number of pledged shares by dividing {money(collateral_value, 'USD')} of collateral by the median share price in the week around the spreadsheet date. It then marks those same shares to today's pulled price.

Currency note: share-price math uses each ticker's pulled trading currency. If the loan is USD and the shares trade in another currency, verify FX conversion before relying on collateral coverage. For plain numeric international tickers, the app tries common Yahoo suffixes based on the market hint, including China `.SS`/`.SZ`, Hong Kong `.HK`, Japan `.T`, Korea `.KS`, Taiwan `.TW`/`.TWO`, Canada `.TO`, Australia `.AX`, London `.L`, Germany `.DE`, France `.PA`, and India `.NS`.

## Step 2: Loan Coverage And Plunge Filter
The filter flags rows where the share price dropped at least {drop_threshold:.1f}% from the loan-date weekly median price. Current collateral value is compared with the loan payoff including simple interest.

## Step 3: Call / Put Overlay
The option overlay assumes {fmt_pct_plain(volatility)} volatility, {loan_years:.1f} years to maturity, a bought call at {fmt_pct_plain(call_strike_pct)} of today's underlying price, and a sold put at {fmt_pct_plain(put_strike_pct)} of today's underlying price. Premiums use a Black-Scholes estimate for screening only.

| Input Ticker | Resolved Ticker | Company | Exchange | Currency | Loan Date | Week Median Price | Today Price | Price Drop | Pledged Shares | Current Collateral | Loan Payoff | Collateral Surplus / Shortfall | Buy Call Cost / Share | Sell Put Credit / Share | Net Option Cost / Share | Net Sale Payout / Share | Net After Loan | Data Source | Plunged? |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
{table_rows}

## What To Do Next
- Verify each ticker/date against actual pledge agreements, exact share count, exchange prices, and corporate actions.
- Confirm whether the loan has covenants, margin-call rights, recourse, lockups, and forced-sale restrictions.
- Verify live options chains instead of relying on Black-Scholes estimates.
- Check borrow, liquidity, trading halts, FX, withholding tax, local-market settlement, and legal restrictions before acting.
"""
    CHAT_STATE["company"] = "Collateral plunge research"
    CHAT_STATE["brief"] = brief
    CHAT_STATE["sources"] = []
    CHAT_STATE["figures"] = []
    (OUTPUT_DIR / "latest_brief.md").write_text(brief, encoding="utf-8")
    return brief


def investment_simulator_rows(market: MarketAnalysis | None, amount: float) -> str:
    if not market:
        return "| - | No public ticker was available, so the simulator could not calculate historical-trend outcomes. | - | - | - | - |"
    rows = []
    horizons = [1, 3, 5, 10]
    for scenario in market.scenarios:
        values = []
        for years in horizons:
            ret = scenario.projected_returns.get(years)
            values.append(f"{money(future_value(amount, ret), market.target.currency)} ({pct(ret)})" if ret is not None else "Not available")
        rows.append(f"| {md_cell(scenario.label)} | {pct(scenario.annual_rate_pct)} | {' | '.join(values)} |")
    return "\n".join(rows)


def monthly_allocation(market: MarketAnalysis | None, monthly_amount: float = 1000.0) -> list[tuple[str, float, str]]:
    if not market:
        return [
            ("Cash / wait for ticker data", monthly_amount, "No market data was available, so the model cannot allocate to the target."),
        ]
    target_weight = 0.50
    benchmark_weight = 0.35
    cash_weight = 0.15
    m = market.target_metrics
    valuation_text = " ".join(valuation_assessment(market)).lower()
    if "potentially overvalued" in valuation_text:
        target_weight -= 0.25
        cash_weight += 0.15
        benchmark_weight += 0.10
    elif "demanding" in valuation_text:
        target_weight -= 0.15
        cash_weight += 0.10
        benchmark_weight += 0.05
    elif "does not show obvious overvaluation" in valuation_text:
        target_weight += 0.10
        benchmark_weight -= 0.05
        cash_weight -= 0.05
    if m.volatility_pct >= 35 or m.max_drawdown_pct <= -40:
        target_weight -= 0.15
        cash_weight += 0.10
        benchmark_weight += 0.05
    if m.cagr_pct < 0:
        target_weight -= 0.10
        cash_weight += 0.05
        benchmark_weight += 0.05
    if market.benchmark:
        target_1y = period_return(market.target.points, 1)
        benchmark_1y = period_return(market.benchmark.points, 1)
        if target_1y is not None and benchmark_1y is not None and target_1y < benchmark_1y:
            target_weight -= 0.10
            benchmark_weight += 0.10
    target_weight = min(max(target_weight, 0.10), 0.75)
    benchmark_weight = min(max(benchmark_weight, 0.10), 0.75)
    cash_weight = max(cash_weight, 0.05)
    total = target_weight + benchmark_weight + cash_weight
    target_weight, benchmark_weight, cash_weight = target_weight / total, benchmark_weight / total, cash_weight / total
    benchmark_label = market.benchmark.ticker if market.benchmark else "broad benchmark ETF / index fund"
    return [
        (market.target.ticker, monthly_amount * target_weight, "Target stock sleeve based on valuation, trend, volatility, and drawdown screen."),
        (benchmark_label, monthly_amount * benchmark_weight, "Diversifier/core market sleeve to reduce single-stock risk."),
        ("Cash / watchlist reserve", monthly_amount * cash_weight, "Reserve for volatility, better entry points, or missing diligence confirmation."),
    ]


def monthly_allocation_rows(market: MarketAnalysis | None, monthly_amount: float = 1000.0) -> str:
    rows = []
    for label, amount, reason in monthly_allocation(market, monthly_amount):
        rows.append(f"| {md_cell(label, 120)} | {money(amount, market.target.currency if market else 'USD')} | {md_cell(reason, 360)} |")
    return "\n".join(rows)


def monthly_allocation_read(market: MarketAnalysis | None, monthly_amount: float = 1000.0) -> str:
    if not market:
        return "No ticker data was available, so the sample monthly allocation keeps the amount in cash/watchlist until market data is provided."
    allocation = monthly_allocation(market, monthly_amount)
    target = allocation[0]
    return (
        f"For an educational {money(monthly_amount, market.target.currency)} monthly model, the screen allocates "
        f"{money(target[1], market.target.currency)} to {target[0]} and diversifies the rest based on valuation, volatility, drawdown, and benchmark evidence. "
        "This is not personalized investment advice."
    )


def market_section(market: MarketAnalysis | None, investment_amount: float = 1000.0) -> str:
    if not market:
        return """
## Market & Competitors
No public ticker was provided, so market performance could not be analyzed.
"""
    m = market.target_metrics
    returns = {
        "1-year return": period_return(market.target.points, 1),
        "3-year return": period_return(market.target.points, 3),
        "5-year return": period_return(market.target.points, 5),
    }
    benchmark_read = "No benchmark comparison was available."
    benchmark_row = ""
    if market.benchmark:
        bm = compute_metrics(market.benchmark)
        target_1y = returns["1-year return"]
        benchmark_1y = period_return(market.benchmark.points, 1)
        if target_1y is not None and benchmark_1y is not None:
            benchmark_read = f"{market.target.ticker} {'outperformed' if target_1y >= benchmark_1y else 'underperformed'} {market.benchmark.ticker} by {abs(target_1y - benchmark_1y):.1f} percentage points over the last year."
        else:
            benchmark_read = f"Benchmark {market.benchmark.ticker} was pulled, but a full 1-year comparison was not available."
        benchmark_row = f"| Benchmark: {md_cell(bm.ticker)} | {pct(bm.total_return_pct)} | {pct(bm.cagr_pct)} | {bm.volatility_pct:.1f}% | {bm.max_drawdown_pct:.1f}% | {bm.sharpe_like:.2f} |"
    annual_rows = "\n".join(f"| {year} | {pct(ret)} |" for year, ret in sorted(m.annual_returns.items())[-8:])
    peer_rows = "\n".join(
        f"| Peer: {md_cell(cm.ticker)} | {pct(cm.total_return_pct)} | {pct(cm.cagr_pct)} | {cm.volatility_pct:.1f}% | {cm.max_drawdown_pct:.1f}% | {cm.sharpe_like:.2f} |"
        for cm in market.peer_metrics
    ) or "| - | No peer tickers were provided or peer data could not be pulled. | - | - | - | - |"
    scenario_rows = "\n".join(
        f"| {md_cell(s.label)} | {pct(s.annual_rate_pct)} | {money(s.projected_prices[1], m.currency)} | {pct(s.projected_returns[1])} | {money(s.projected_prices[3], m.currency)} | {pct(s.projected_returns[3])} | {money(s.projected_prices[5], m.currency)} | {pct(s.projected_returns[5])} | {money(s.projected_prices[10], m.currency)} | {pct(s.projected_returns[10])} |"
        for s in market.scenarios
    )
    news_rows = "\n".join(
        f"| {md_cell(item.published.strftime('%Y-%m-%d') if item.published else 'Date not available', 80)} | {md_cell(item.title, 260)} | {md_cell(item.publisher or 'Source not available', 120)} | {f'[Link]({item.link})' if item.link.startswith('http') else '-'} |"
        for item in market.market_news[:8]
    ) or "| - | No recent market news headlines were returned. | - | - |"
    catalyst_block = bullet_block(
        [catalyst.explanation for catalyst in market.move_catalysts],
        "No unusually large weekly move was detected in the pulled period, or news search did not return enough context for a catalyst read.",
    )
    valuation_block = bullet_block(valuation_assessment(market), "Valuation data was not available.")
    return f"""
## Market & Competitors: {market.target.ticker}
Data note: {market.source_note} Source: [Yahoo Finance]({market.target.source_url}).

Symbol resolution: {market.symbol_note}

International / Yahoo note: {market.international_note}

Trend summary: {m.trend_label} {benchmark_read} Potential upside/downside should be framed against annualized volatility of {m.volatility_pct:.1f}% and historical max drawdown of {m.max_drawdown_pct:.1f}%.

![{market.target.ticker} price trend]({market.price_chart_path})

### Trading Snapshot
This is the current / latest available Yahoo Finance trading picture for the public instrument.

| Field | Value |
| --- | --- |
{trading_snapshot_rows(market.target)}

### Yahoo Finance Company & Quote Details
For non-U.S. listings, this section uses Yahoo Finance as the primary market-data source because SEC filings may not exist.

| Field | Yahoo Finance Data |
| --- | --- |
{yahoo_detail_rows(market.target)}

{f"Business summary from Yahoo Finance: {market.target.business_summary}" if market.target.business_summary else "Business summary from Yahoo Finance: Not available."}

### Public Disclosure Search Pack
Use these links to verify local exchange filings, annual/interim reports, announcements, and issuer disclosures. The app cannot guarantee every market has a free machine-readable filings API, so it provides the best public disclosure starting points by listing/exchange.

| Source | Link |
| --- | --- |
{disclosure_link_rows(market.disclosure_links)}

### Valuation Read: P/E, P/B, And Overvaluation Check
This is a valuation screen, not investment advice. It uses Yahoo Finance fields where available and compares against peers when peer tickers are entered.

| Metric | Value | Read |
| --- | --- | --- |
{valuation_rows(market)}

{valuation_block}

| Metric | Value |
| --- | --- |
| Latest weekly close | {money(m.latest_price, m.currency)} |
| 1-year return | {pct(returns["1-year return"]) if returns["1-year return"] is not None else "Not available from pulled history"} |
| 3-year return | {pct(returns["3-year return"]) if returns["3-year return"] is not None else "Not available from pulled history"} |
| 5-year return | {pct(returns["5-year return"]) if returns["5-year return"] is not None else "Not available from pulled history"} |
| Starting weekly close in pulled period | {money(m.start_price, m.currency)} |
| High weekly close | {money(m.high_price, m.currency)} |
| Low weekly close | {money(m.low_price, m.currency)} |
| Total return | {pct(m.total_return_pct)} |
| CAGR / annualized return | {pct(m.cagr_pct)} |
| Annualized volatility estimate | {m.volatility_pct:.1f}% |
| Sharpe-like return/risk ratio | {m.sharpe_like:.2f} |
| Max drawdown | {m.max_drawdown_pct:.1f}% |
| Current drawdown from high | {m.current_drawdown_pct:.1f}% |
| Best week | {pct(m.best_week_pct)} |
| Worst week | {pct(m.worst_week_pct)} |
| Positive week rate | {m.positive_week_pct:.1f}% |
| 40-week moving average | {money(m.ma_40, m.currency)} |
| Trend read | {m.trend_label} |

### Investment Simulator
Educational model only. This does not know your objectives, taxes, liquidity needs, time horizon, or risk tolerance. It uses historical CAGR and volatility stress cases from Yahoo price history, so it is not a prediction.

Starting amount entered: {money(investment_amount, m.currency)}

| Scenario | Annual rate used | 1Y value | 3Y value | 5Y value | 10Y value |
| --- | --- | --- | --- | --- | --- |
{investment_simulator_rows(market, investment_amount)}

### Sample $1,000 Monthly Allocation
This is a rule-based example for the current month, not a personal recommendation. It shifts money away from the target when valuation looks stretched, volatility/drawdown is high, or the stock underperforms the benchmark.

{monthly_allocation_read(market, 1000.0)}

| Sleeve | Amount | Why |
| --- | --- | --- |
{monthly_allocation_rows(market, 1000.0)}

![Indexed performance comparison]({market.indexed_chart_path})

| Comparable | Total Return | CAGR | Volatility | Max Drawdown | Return/Risk |
| --- | --- | --- | --- | --- | --- |
| Target: {md_cell(m.ticker)} | {pct(m.total_return_pct)} | {pct(m.cagr_pct)} | {m.volatility_pct:.1f}% | {m.max_drawdown_pct:.1f}% | {m.sharpe_like:.2f} |
{benchmark_row}
{peer_rows}

![Potential growth / loss scenarios]({market.forecast_chart_path})

### Potential Growth / Loss Scenarios
These are not predictions. They are simple historical-trend stress cases based on the pulled period's CAGR and volatility.

{scenario_read(m, market.scenarios)}

| Scenario | Annual rate used | 1Y price | 1Y gain/loss | 3Y price | 3Y gain/loss | 5Y price | 5Y gain/loss | 10Y price | 10Y gain/loss |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
{scenario_rows}

![{market.target.ticker} drawdown]({market.drawdown_chart_path})

![{market.target.ticker} annual returns]({market.annual_chart_path})

| Year | Return |
| --- | --- |
{annual_rows}

### News & Large-Move Catalyst Review
This section searches market news and uses it to form hypotheses for large price moves. It is not proof of causation; verify against filings, earnings releases, management commentary, and full articles.

{catalyst_block}

| Date | Headline | Publisher | Link |
| --- | --- | --- | --- |
{news_rows}
"""


def quality_score(
    figures: list[Figure],
    market: MarketAnalysis | None,
    uploaded_count: int,
    pasted_notes: bool,
    peer_tickers: str,
    benchmark_ticker: str,
    website_only: bool,
    sec_quarterly_count: int = 0,
) -> tuple[int, str]:
    score = 10
    verified_count = len(financial_snapshot_figures(figures))
    possible_operating_count = len(operating_figures(figures))
    if sec_quarterly_count:
        score += min(16 + sec_quarterly_count * 6, 34)
    if market:
        score += 18
    if uploaded_count:
        score += min(12 + uploaded_count * 8, 28)
    if pasted_notes:
        score += 12
    if verified_count:
        score += min(verified_count * 7, 28)
    if possible_operating_count:
        score += min(possible_operating_count * 3, 10)
    if peer_tickers:
        score += 7
    if benchmark_ticker:
        score += 5
    if website_only:
        score -= 18
    score = min(score, 100)
    if score >= 75:
        label = "strong source packet for a first-pass diligence brief"
    elif score >= 50:
        label = "usable first-pass packet with meaningful gaps"
    else:
        label = "thin source packet; add financial statements, SEC excerpts, deck text, or analyst notes"
    return score, label


def build_risk_flags(sources: list[Source], figures: list[Figure], market: MarketAnalysis | None, uploaded_count: int) -> list[str]:
    flags: list[str] = []
    verified = financial_snapshot_figures(figures)
    corpus = " ".join(fig.context.lower() for fig in verified + operating_figures(figures))
    if not verified:
        flags.append("Missing financials: no verified revenue, margin, cash, debt, or cash-flow data was found.")
    if "customer" not in corpus and "customers" not in corpus and "concentration" not in corpus:
        flags.append("Customer concentration unknown: current source packet does not disclose top-customer exposure.")
    if "margin" not in corpus and "ebitda" not in corpus and "operating income" not in corpus:
        flags.append("No margin data: gross margin, EBITDA, or operating income is not disclosed based on current source packet.")
    if "cash" not in corpus and "debt" not in corpus:
        flags.append("No debt/cash data: liquidity, leverage, runway, and debt maturity risk cannot be assessed.")
    if uploaded_count == 0 and not has_pasted_notes(sources) and not has_sec_quarterly_sources(sources):
        flags.append("Weak source quality: current packet relies primarily on website text rather than diligence materials.")
    if market:
        m = market.target_metrics
        if m.volatility_pct >= 35:
            flags.append(f"High volatility: {market.target.ticker} annualized volatility is {m.volatility_pct:.1f}%.")
        if m.max_drawdown_pct <= -35:
            flags.append(f"Major drawdown: {market.target.ticker} experienced a {m.max_drawdown_pct:.1f}% max drawdown in the pulled period.")
        if market.benchmark:
            target_1y = period_return(market.target.points, 1)
            benchmark_1y = period_return(market.benchmark.points, 1)
            if target_1y is not None and benchmark_1y is not None and target_1y < benchmark_1y:
                flags.append(f"Underperformance vs benchmark: {market.target.ticker} trailed {market.benchmark.ticker} by {benchmark_1y - target_1y:.1f} percentage points over the last year.")
    if detect_legal_or_regulatory_language(sources):
        flags.append("Legal/regulatory language found: review litigation, compliance, data security, and regulatory disclosures for materiality.")
    if detect_conflicting_numbers(figures):
        flags.append("Conflicting numbers in uploaded/pasted materials: reconcile duplicate metric references before relying on them.")
    return flags


def sec_driver_sentences(sources: list[Source], limit: int = 10) -> list[str]:
    driver_terms = [
        "decrease", "decreased", "decline", "declined", "lower", "down", "increase", "increased", "higher", "growth",
        "primarily due", "driven by", "offset by", "because", "resulted from", "attributable to", "quarter", "three months",
        "six months", "nine months", "revenue", "sales", "gross margin", "operating income", "net income", "cash flow",
    ]
    items: list[str] = []
    seen: set[str] = set()
    for source in sec_quarterly_sources(sources):
        for sentence in re.split(r"(?<=[.!?])\s+", source.text):
            sentence = clean_text(sentence)
            lowered = sentence.lower()
            if len(sentence) < 45 or len(sentence) > 520:
                continue
            if not any(term in lowered for term in driver_terms):
                continue
            signature = sentence[:150].lower()
            if signature in seen:
                continue
            seen.add(signature)
            items.append(f"{sentence} Source: {source.label}.")
            if len(items) >= limit:
                return items
    return items


def verified_fact_bullets(figures: list[Figure], limit: int = 10) -> list[str]:
    facts = []
    for fig in financial_snapshot_figures(figures)[:limit]:
        facts.append(f"{fig.value}: {fig.context} Source: {fig.source_label}.")
    return facts


def market_summary_bullets(market: MarketAnalysis | None) -> list[str]:
    if not market:
        return ["No public ticker was provided or resolved, so market performance, volatility, drawdown, peer comparison, and catalyst analysis could not be completed."]
    m = market.target_metrics
    returns = {
        "1-year": period_return(market.target.points, 1),
        "3-year": period_return(market.target.points, 3),
        "5-year": period_return(market.target.points, 5),
    }
    bullets = [
        f"{market.target.ticker} trades in {market.target.currency} on {market.target.exchange or 'an exchange not disclosed by Yahoo'}; Yahoo resolved it as {market.target.name or market.target.ticker}.",
        f"Latest weekly close was {money(m.latest_price, m.currency)}; pulled-period total return was {pct(m.total_return_pct)}, CAGR was {pct(m.cagr_pct)}, annualized volatility was {m.volatility_pct:.1f}%, and max drawdown was {m.max_drawdown_pct:.1f}%.",
        f"Return windows: 1-year {pct(returns['1-year']) if returns['1-year'] is not None else 'not available'}, 3-year {pct(returns['3-year']) if returns['3-year'] is not None else 'not available'}, 5-year {pct(returns['5-year']) if returns['5-year'] is not None else 'not available'}.",
        f"Trend read: {m.trend_label}",
    ]
    if market.benchmark:
        benchmark_1y = period_return(market.benchmark.points, 1)
        target_1y = returns["1-year"]
        if benchmark_1y is not None and target_1y is not None:
            bullets.append(f"Benchmark comparison: {market.target.ticker} {'outperformed' if target_1y >= benchmark_1y else 'underperformed'} {market.benchmark.ticker} by {abs(target_1y - benchmark_1y):.1f} percentage points over the last year.")
    if market.target.fifty_two_week_low is not None and market.target.fifty_two_week_high is not None:
        bullets.append(f"Yahoo 52-week / derived range: {money(market.target.fifty_two_week_low, market.target.currency)} to {money(market.target.fifty_two_week_high, market.target.currency)}.")
    if market.target.market_cap is not None or market.target.trailing_pe is not None or market.target.dividend_yield_pct is not None:
        bullets.append(
            f"Yahoo valuation snapshot: market cap {compact_money(market.target.market_cap, market.target.currency)}, trailing P/E {fmt_ratio(market.target.trailing_pe)}, dividend yield {fmt_pct_plain(market.target.dividend_yield_pct)}."
        )
    bullets.extend(valuation_assessment(market)[:4])
    return bullets


def catalyst_summary_bullets(market: MarketAnalysis | None) -> list[str]:
    if not market:
        return ["No market data was available, so the app could not explain stock drops or spikes."]
    if not market.move_catalysts:
        return ["No unusually large weekly move was detected, or relevant headlines were not available. Do not infer a catalyst without more evidence."]
    bullets = [catalyst.explanation for catalyst in market.move_catalysts]
    if not market.market_news:
        bullets.append("Relevant Yahoo headline search returned no company-specific headlines, so large-move explanations should be treated as price-action observations rather than confirmed causes.")
    return bullets


def build_deep_summary(
    company: str,
    sources: list[Source],
    figures: list[Figure],
    market: MarketAnalysis | None,
    risk_flags: list[str],
    sec_summary: str,
    verified_summary: str,
    operating_summary: str,
    profile_items: list[str] | None = None,
    transaction_items: list[str] | None = None,
    investment_amount: float = 1000.0,
) -> str:
    verified_facts = verified_fact_bullets(figures)
    sec_drivers = sec_driver_sentences(sources)
    market_bullets = market_summary_bullets(market)
    catalysts = catalyst_summary_bullets(market)
    profile_items = profile_items or company_profile_bullets(sources, market)
    transaction_items = transaction_items or transaction_bullets(sources, market)
    missing_items = [
        "Revenue by segment/geography and quarter-over-quarter bridge",
        "Gross margin, operating margin, EBITDA, cash flow, cash balance, and debt bridge",
        "Management explanation for any quarter-over-quarter weakness",
        "Customer concentration and churn / retention",
        "Full earnings call transcript, investor presentation, and latest annual report",
    ]
    conclusion = (
        f"The current packet for {company} is strong enough for an initial read when SEC/Yahoo data is available, "
        "but not enough for a final investment decision unless the missing diligence items are resolved."
    )
    all_learned = [
        f"Business/products learned: {profile_items[0]}" if profile_items else "Business/products learned: not disclosed based on current source packet.",
        f"Major transactions/corporate actions learned: {transaction_items[0]}" if transaction_items else "Major transactions/corporate actions learned: none found in the current source packet.",
        f"Verified financials learned: {verified_facts[0]}" if verified_facts else "Verified financials learned: no verified financial metrics were extracted.",
        f"Quarterly drivers learned: {sec_drivers[0]}" if sec_drivers else "Quarterly drivers learned: no source-backed explanation for quarterly increases/decreases was found.",
        f"Market performance learned: {market_bullets[0]}" if market_bullets else "Market performance learned: no public market signal was available.",
        f"Valuation learned: {valuation_assessment(market)[-2]}" if market else "Valuation learned: no P/E or P/B market valuation data was available.",
        f"Investment simulator learned: {monthly_allocation_read(market, 1000.0)}" if market else "Investment simulator learned: no ticker data was available, so no modeled stock allocation was produced.",
        f"Stock move/catalyst learned: {catalysts[0]}" if catalysts else "Stock move/catalyst learned: no source-backed catalyst was available.",
        f"Primary risk learned: {risk_flags[0]}" if risk_flags else "Primary risk learned: no automated red flags were found, but analyst review is still required.",
    ]
    return f"""
## Summary: Everything You Need To Know

### All Information Learned In This Run
{bullet_block(all_learned, 'No summary facts were available from the current source packet.')}

### Bottom Line
- {conclusion}
- Source position: {sec_summary} {verified_summary} {operating_summary}
- The app separates confirmed source-backed facts from market-data signals and from catalyst hypotheses. Any statement about why a stock dropped or rose should be treated as a hypothesis unless it is supported by filings, earnings releases, transcripts, or multiple relevant headlines.

### What Is Verified
{bullet_block(verified_facts, 'No verified financial metrics were extracted. Not disclosed based on current source packet.')}

### What The Company Does / Products Sold
{bullet_block(profile_items, 'Product and business model details were not disclosed based on current source packet. Add an investor presentation, annual report, or clean website/company profile text.')}

### Major Transactions / Corporate Actions
{bullet_block(transaction_items, 'No major acquisitions, divestitures, buybacks, dividends, debt offerings, capital raises, restructurings, major contracts, or partnerships were found in the current source packet.')}

### Quarterly / Operating Drivers
These are the filing-backed or source-backed lines most relevant to why quarters may be up or down.

{bullet_block(sec_drivers, 'No source-backed quarter-over-quarter drivers were found. Upload/paste the latest earnings release, 10-Q/6-K/annual report, or transcript to explain why the quarter was down or up.')}

### Stock Performance And Market Signal
{bullet_block(market_bullets, 'No public market signal was available.')}

### Valuation / Overvaluation Read
{bullet_block(valuation_assessment(market), 'No P/E, P/B, or peer valuation data was available.')}

### Investment Simulator / Monthly Allocation
- Starting amount modeled: {money(investment_amount, market.target.currency if market else 'USD')}.
- {monthly_allocation_read(market, 1000.0)}
- The simulator is educational only; it uses historical trend stress cases and cannot account for your personal finances, taxes, time horizon, liquidity needs, or risk tolerance.

### Why The Stock Dropped Or Rose
{bullet_block(catalysts, 'No catalyst explanation was available.')}

### Biggest Risks And Open Questions
{bullet_block(risk_flags, 'No automated red flags were found, but analyst review is still required.')}

### What Is Still Missing
{bullet_block([f'Not disclosed based on current source packet: {item}.' for item in missing_items], 'No missing items were identified.')}

### Next Analyst Actions
- Read the most recent earnings release and call transcript to confirm management's explanation for revenue, margin, and guidance changes.
- Reconcile Yahoo price moves against earnings dates, guidance changes, analyst revisions, regulatory events, macro moves, and sector/peer performance.
- For non-U.S. companies, pull the local annual/interim report or exchange filing because SEC 10-Q data may not exist.
- Verify every key number against primary sources before using this as an investment memo.
"""


def make_fallback_brief(
    company: str,
    website: str,
    strategy: str,
    sources: list[Source],
    figures: list[Figure],
    market: MarketAnalysis | None,
    uploaded_count: int,
    peer_tickers: str,
    benchmark_ticker: str,
    sec_company: SECCompany | None = None,
    investment_amount: float = 1000.0,
) -> str:
    corpus = "\n".join(source.text for source in sources)
    overview = clean_business_sentences(sources)
    verified_financials = financial_snapshot_figures(figures)
    possible_operating = operating_figures(figures)
    financial_table = make_figure_table(verified_financials, None, 22)
    operating_table = make_figure_table(possible_operating, None, 14)
    review_table = make_figure_table(figures, None, 40)
    source_rows = "\n".join(f"- {source.label}: {source.url if source.url else 'provided input'}" for source in sources)
    profile_items = company_profile_bullets(sources, market)
    transaction_items = transaction_bullets(sources, market)
    pasted_notes = has_pasted_notes(sources)
    sec_sources = sec_quarterly_sources(sources)
    sec_count = len(sec_sources)
    website_only = has_website_source(sources) and not uploaded_count and not pasted_notes and market is None and not sec_count
    score, score_label = quality_score(figures, market, uploaded_count, pasted_notes, peer_tickers, benchmark_ticker, website_only, sec_count)
    risk_flags = build_risk_flags(sources, figures, market, uploaded_count)
    verified_summary = "Verified financial metrics were found in diligence-grade source material." if verified_financials else "No verified financial data was found."
    operating_summary = "Possible operating metrics were identified for analyst review." if possible_operating else "No reliable operating metrics were disclosed based on current source packet."
    sec_summary = (
        f"SEC quarterly filings pulled: {sec_count} recent Form 10-Q filing(s) for {sec_company.title} ({sec_company.ticker}, CIK {sec_company.cik})."
        if sec_company and sec_count
        else "SEC quarterly filings: no recent Form 10-Q packet was pulled for this company."
    )
    sec_readout_items = sec_filing_readout(sources)
    market_summary = ""
    if market:
        tm = market.target_metrics
        market_summary = f" Public market data for {market.target.ticker} shows {pct(tm.total_return_pct)} total return, {pct(tm.cagr_pct)} CAGR, {tm.volatility_pct:.1f}% estimated annualized volatility, and {tm.max_drawdown_pct:.1f}% max drawdown over the pulled period."
    deep_summary = build_deep_summary(company, sources, figures, market, risk_flags, sec_summary, verified_summary, operating_summary, profile_items, transaction_items, investment_amount)

    return f"""# Advanced Due-Diligence Brief: {company}

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}
Target website: {website or 'Not provided'}
Strategy lens: {strategy}

## Executive Readout
- Source quality score: {score}/100 — {score_label}.
- Source quality: {'website-only packet; low diligence reliability' if website_only else 'mixed source packet; rely on tables below for what is actually supported'}.
- SEC packet: {sec_summary}
- Verified information: {verified_summary} {operating_summary}
- Missing information: revenue, margins, EBITDA, cash balance, debt, runway, funding, and customer concentration remain "Not disclosed based on current source packet" unless shown in the verified table below.
- Public market performance: {f'{market.target.ticker} market data was pulled and analyzed below.' if market else 'No public ticker was provided, so market performance could not be analyzed.'}
- Biggest diligence risks: {('; '.join(risk_flags[:3]) + '.') if risk_flags else 'No major automated red flags were found, but analyst review is still required.'}
- Clear next steps: upload financial statements, paste SEC filing excerpts, add a pitch deck, provide peer tickers, and reconcile any management-provided metrics against source documents.

## Business Overview
Website text was used only for company description, products, services, and basic positioning. Footer, legal, privacy, phone, product-promo, and navigation text was filtered out.

{bullet_block(overview, 'Business description is not disclosed based on current source packet. Add a clean company description, pitch deck text, or management notes to complete this section.')}

## Products & Business Model
This section explains what the company does, what it sells, and how Yahoo/company/public-source text describes the business.

{bullet_block(profile_items, 'Product and business model details were not disclosed based on current source packet. Add an annual report, investor presentation, or company profile text.')}

## Major Transactions / Corporate Actions
This section searches filings, public-source text, uploaded materials, pasted notes, and Yahoo headlines for acquisitions, mergers, divestitures, capital raises, debt, buybacks, dividends, restructuring, major contracts, and partnerships.

{bullet_block(transaction_items, 'No major transactions or corporate actions were found in the current source packet.')}

## Financial Snapshot
{'' if verified_financials else 'No verified financial data was found. Upload financial statements, SEC filings, or paste financial excerpts to complete this section.'}

### Quarterly SEC Filing Readout
{bullet_block(sec_readout_items, 'No quarterly SEC filing financial readout was available. If this company is public, enter the exact ticker or paste the latest 10-Q excerpt.')}

### Verified Financial Metrics
| ID | Figure | Classification | Source context | Source | Analyst note |
| --- | --- | --- | --- | --- | --- |
{financial_table}

### Possible Operating Metrics
These items have operating context but still require confirmation of definition, period, and source.

| ID | Figure | Classification | Source context | Source | Analyst note |
| --- | --- | --- | --- | --- | --- |
{operating_table}

{market_section(market, investment_amount)}

## Risk / Red Flags
{bullet_block(risk_flags, 'No automated red flags were found. Continue analyst review for source quality, omitted financials, customer concentration, legal exposure, and metric definitions.')}

## Advanced Diligence Workplan
| Area | What to Request | Why It Matters |
| --- | --- | --- |
| Financials | Monthly P&L, balance sheet, cash flow, revenue detail, and debt schedule | Confirms whether the story is supported by actual operating performance. |
| Revenue Quality | Customer cohort, retention, churn, pipeline, bookings, and contract renewal schedule | Separates recurring, sticky revenue from one-time or fragile revenue. |
| Customers | Top customer list, concentration, signed contracts, NPS or satisfaction data | Tests whether growth depends on a few relationships. |
| Product | Product roadmap, uptime, support tickets, usage data, security reports | Checks scalability and hidden technical risk. |
| Legal / Compliance | Litigation, regulatory correspondence, IP ownership, privacy/security policies | Finds risks that do not appear on the website. |
| Public Market Lens | Compare returns, volatility, drawdowns, and peer performance | Shows how public investors are pricing the company relative to alternatives. |

## Extracted Numeric Claims Review
This is a review queue, not a financial statement. The classifier only treats numbers as financial or operating candidates when nearby context contains diligence-relevant terms, and it rejects website footer, phone, legal, privacy, product, promo, and navigation noise.

| ID | Figure | Classification | Source context | Source | Analyst note |
| --- | --- | --- | --- | --- | --- |
{review_table}

## Analyst Follow-Up Questions
- What revenue, gross margin, EBITDA, cash burn, cash balance, debt, and net retention figures can management source from financial statements?
- Which customers account for the largest share of revenue, and what contract terms govern renewal, termination, exclusivity, and churn?
- Which numbers in the deck are audited, system-generated, management-estimated, or forward-looking?
- Who are the closest competitors, and where does the target win or lose on price, product depth, distribution, and switching costs?
- If this is a public company, how do stock performance, drawdowns, and volatility compare with revenue growth and margin performance?
- For a {strategy.lower()} lens, what would break the investment thesis in the first 100 days after close?

{deep_summary}

## Sources Used
{source_rows}
"""


def make_llm_brief(
    company: str,
    website: str,
    strategy: str,
    sources: list[Source],
    figures: list[Figure],
    market: MarketAnalysis | None,
    uploaded_count: int,
    peer_tickers: str,
    benchmark_ticker: str,
    sec_company: SECCompany | None = None,
    investment_amount: float = 1000.0,
) -> str | None:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None
    try:
        from openai import OpenAI  # type: ignore
    except Exception:
        return None
    source_packet = [{"label": s.label, "url": s.url, "text": s.text[:7000]} for s in sources]
    figure_packet = [
        {"id": f"F{i}", "value": f.value, "category": f.category, "context": f.context, "source": f.source_label, "url": f.source_url}
        for i, f in enumerate(figures, start=1)
    ]
    market_packet = None
    if market:
        tm = market.target_metrics
        benchmark_packet = None
        if market.benchmark:
            bm = compute_metrics(market.benchmark)
            benchmark_packet = {
                "ticker": market.benchmark.ticker,
                "total_return_pct": bm.total_return_pct,
                "cagr_pct": bm.cagr_pct,
                "volatility_pct": bm.volatility_pct,
                "max_drawdown_pct": bm.max_drawdown_pct,
                "one_year_return_pct": period_return(market.benchmark.points, 1),
            }
        market_packet = {
            "ticker": market.target.ticker,
            "latest_price": tm.latest_price,
            "one_year_return_pct": period_return(market.target.points, 1),
            "three_year_return_pct": period_return(market.target.points, 3),
            "five_year_return_pct": period_return(market.target.points, 5),
            "total_return_pct": tm.total_return_pct,
            "cagr_pct": tm.cagr_pct,
            "volatility_pct": tm.volatility_pct,
            "max_drawdown_pct": tm.max_drawdown_pct,
            "current_drawdown_pct": tm.current_drawdown_pct,
            "trend_label": tm.trend_label,
            "benchmark": benchmark_packet,
            "peers": [{"ticker": p.ticker, "total_return_pct": p.total_return_pct, "cagr_pct": p.cagr_pct, "volatility_pct": p.volatility_pct} for p in market.peer_metrics],
            "symbol_note": market.symbol_note,
            "international_note": market.international_note,
            "yahoo_quote_details": {
                "name": market.target.name,
                "symbol": market.target.ticker,
                "quote_type": market.target.quote_type,
                "exchange": market.target.exchange,
                "country": market.target.country,
                "currency": market.target.currency,
                "sector": market.target.sector,
                "industry": market.target.industry,
                "regular_market_price": market.target.regular_market_price,
                "previous_close": market.target.previous_close,
                "day_low": market.target.day_low,
                "day_high": market.target.day_high,
                "fifty_two_week_low": market.target.fifty_two_week_low,
                "fifty_two_week_high": market.target.fifty_two_week_high,
                "market_cap": market.target.market_cap,
                "trailing_pe": market.target.trailing_pe,
                "forward_pe": market.target.forward_pe,
                "price_to_book": market.target.price_to_book,
                "dividend_yield_pct": market.target.dividend_yield_pct,
                "beta": market.target.beta,
                "average_volume": market.target.average_volume,
                "shares_outstanding": market.target.shares_outstanding,
                "employees": market.target.employees,
                "website": market.target.website,
                "business_summary": market.target.business_summary,
            },
            "valuation_assessment": valuation_assessment(market),
            "investment_simulator": {
                "starting_amount": investment_amount,
                "currency": market.target.currency,
                "scenario_rows_markdown": investment_simulator_rows(market, investment_amount),
                "sample_monthly_allocation_1000": [
                    {"sleeve": label, "amount": amount, "reason": reason}
                    for label, amount, reason in monthly_allocation(market, 1000.0)
                ],
                "allocation_read": monthly_allocation_read(market, 1000.0),
            },
            "public_disclosure_links": [
                {"source": label, "url": url}
                for label, url in market.disclosure_links
            ],
            "recent_news": [
                {
                    "title": item.title,
                    "publisher": item.publisher,
                    "published": item.published.isoformat() if item.published else None,
                    "link": item.link,
                }
                for item in market.market_news[:10]
            ],
            "large_move_catalysts": [
                {
                    "date": catalyst.date.isoformat(),
                    "move_pct": catalyst.move_pct,
                    "explanation": catalyst.explanation,
                    "headlines": [item.title for item in catalyst.headlines],
                }
                for catalyst in market.move_catalysts
            ],
            "growth_loss_scenarios": [
                {
                    "label": scenario.label,
                    "annual_rate_pct": scenario.annual_rate_pct,
                    "projected_returns": scenario.projected_returns,
                    "projected_prices": scenario.projected_prices,
                }
                for scenario in market.scenarios
            ],
        }
    pasted_notes = has_pasted_notes(sources)
    sec_sources = sec_quarterly_sources(sources)
    website_only = has_website_source(sources) and not uploaded_count and not pasted_notes and market is None and not sec_sources
    score, score_label = quality_score(figures, market, uploaded_count, pasted_notes, peer_tickers, benchmark_ticker, website_only, len(sec_sources))
    risk_flags = build_risk_flags(sources, figures, market, uploaded_count)
    profile_items = company_profile_bullets(sources, market)
    transaction_items = transaction_bullets(sources, market)
    sec_packet = {
        "matched_company": None if not sec_company else {
            "title": sec_company.title,
            "ticker": sec_company.ticker,
            "cik": sec_company.cik,
        },
        "quarterly_filings_pulled": [
            {"label": source.label, "url": source.url, "excerpt": source.text[:5000]}
            for source in sec_sources
        ],
    }
    prompt = f"""
Create an advanced preliminary pre-investment due-diligence brief for {company}.
Website: {website}
Strategy lens: {strategy}
Source quality score: {score}/100 — {score_label}

Hard rules:
- Do not invent facts or numbers.
- Every numeric business claim must cite one of the provided figure IDs like [F3] or the market data packet.
- Use only figures categorized as "Verified financial metric" for the Financial Snapshot.
- Do not use company website numbers in the Financial Snapshot.
- Prefer recent SEC Form 10-Q filing excerpts for Financial Snapshot and Executive Readout when SEC data is available.
- If no verified financial metrics exist, write exactly: "No verified financial data was found. Upload financial statements, SEC filings, or paste financial excerpts to complete this section."
- Business Overview may use website text only for company description, products, services, and basic positioning.
- Do not include phone numbers, footer text, privacy policy text, Apple Card legal text, promo text, or navigation text in Business Overview.
- Market & Competitors must use market data if a ticker exists. If no ticker exists, write exactly: "No public ticker was provided, so market performance could not be analyzed."
- Risk / Red Flags must focus on missing financials, customer concentration unknown, no margin data, no debt/cash data, weak source quality, volatility, drawdown, benchmark underperformance, legal/regulatory language, or conflicting numbers.
- If data is missing, say "Not disclosed based on current source packet."
- Executive Readout must summarize source quality, verified information, missing information, market performance if any, biggest diligence risks, and clear next steps.
- Include a very detailed final section titled "Summary: Everything You Need To Know".
- The first subsection inside that final summary must be "All Information Learned In This Run" and must recap, in one place, every important thing learned about the company: what it does, products/services, major transactions, verified financials, quarterly drivers, market performance, why the stock moved, risks, missing information, and next actions.
- Include a "Valuation / Overvaluation Read" that discusses P/E, forward P/E, P/B, market cap, dividend yield, peer medians if available, and whether the stock looks cheap, fairly valued, expensive, or potentially overvalued.
- Never call a stock definitively overvalued based only on P/E or P/B. Explain the evidence, caveats, sector context, missing data, and what would confirm or disprove the valuation concern.
- Include an "Investment Simulator" section showing the entered starting amount across bear/base/bull 1-year, 3-year, 5-year, and 10-year outcomes. Say clearly this is educational, uses historical trend stress cases, and is not a prediction.
- Include a "Sample $1,000 Monthly Allocation" section showing how the rule-based model would split $1,000 this month between target stock, benchmark/core ETF, and cash/watchlist reserve. Say clearly it is not personalized investment advice.
- Include dedicated sections for "Products & Business Model" and "Major Transactions / Corporate Actions".
- In those sections, explain what products/services the company sells, what the company does, major business segments, and any public-source evidence of acquisitions, mergers, divestitures, capital raises, debt, buybacks, dividends, restructuring, major contracts, or partnerships.
- In the final summary, explain all available evidence on: verified financials, products/services, major transactions, quarterly performance, why quarters may be down or up, why the stock dropped or rose, drawdown, volatility, benchmark/peer performance, relevant news catalysts, risks, missing information, and exact next diligence actions.
- Do not claim a causal reason for a stock move unless the source packet, SEC filing, earnings release text, or relevant news headlines support it. If the app only has price action, call it a hypothesis and say what evidence is still needed.
- For non-U.S. companies, rely on Yahoo Finance market/profile data and clearly state when SEC 10-Q filings are not expected or were not found.
- Keep it investor-style and practical.

Market data, if available:
{json.dumps(market_packet, indent=2)}

Automated risk flags:
{json.dumps(risk_flags, indent=2)}

Products / business model findings:
{json.dumps(profile_items, indent=2)}

Major transactions / corporate actions findings:
{json.dumps(transaction_items, indent=2)}

SEC quarterly filing packet:
{json.dumps(sec_packet, indent=2)}

Figures:
{json.dumps(figure_packet, indent=2)}

Sources:
{json.dumps(source_packet, indent=2)}
"""
    client = OpenAI(api_key=api_key)
    response = client.responses.create(model=os.getenv("OPENAI_MODEL", "gpt-4.1-mini"), input=prompt)
    generated = response.output_text
    return generated + "\n\n" + market_section(market, investment_amount)


def build_brief(fields: dict[str, str], files: list[tuple[str, bytes]]) -> str:
    company = clean_text(fields.get("company", "")) or "Target Company"
    website = clean_text(fields.get("website", ""))
    ticker = clean_text(fields.get("ticker", ""))
    peer_tickers = clean_text(fields.get("peers", ""))
    benchmark_ticker = clean_text(fields.get("benchmark", "SPY"))
    investment_amount = parse_amount(fields.get("investment_amount", "1000"), 1000.0)
    screener_country = clean_text(fields.get("screener_country", ""))
    years_raw = clean_text(fields.get("years", "5")) or "5"
    try:
        years = min(max(int(years_raw), 1), 10)
    except Exception:
        years = 5
    strategy = clean_text(fields.get("strategy", "Venture & Growth"))
    pasted = clean_text(fields.get("notes", ""))

    sec_company, sec_sources = fetch_quarterly_sec_filings(company, ticker)
    resolved_ticker = ticker or (sec_company.ticker if sec_company and sec_company.ticker else "")

    sources = [fetch_website(company, website)]
    sources.extend(sec_sources)
    sources.append(fetch_sec_company_search(company))
    if pasted:
        sources.append(Source("Pasted analyst notes / articles", "provided by user", pasted[:MAX_TEXT_CHARS]))
    upload_sources = []
    for filename, raw in files:
        if raw:
            upload_sources.append(source_from_upload(filename, raw))
    sources.extend(upload_sources)

    figures = extract_figures(sources)
    market_lookup_value = resolved_ticker or company
    market = fetch_market_analysis(company, market_lookup_value, peer_tickers, benchmark_ticker, years) if market_lookup_value else None
    screener_years = min(max(years, 3), 10)
    screener_rows = run_country_screener(screener_country, screener_years) if screener_country else []
    screener_market_rows = run_country_market_screen(screener_country, screener_years) if screener_country else []
    brief = make_llm_brief(company, website, strategy, sources, figures, market, len(upload_sources), peer_tickers, benchmark_ticker, sec_company, investment_amount)
    if not brief:
        brief = make_fallback_brief(company, website, strategy, sources, figures, market, len(upload_sources), peer_tickers, benchmark_ticker, sec_company, investment_amount)
    if screener_country:
        brief += "\n\n" + screener_section(screener_country, screener_rows, screener_market_rows, screener_years)
    CHAT_STATE["company"] = company
    CHAT_STATE["brief"] = brief
    CHAT_STATE["sources"] = sources
    CHAT_STATE["figures"] = figures
    (OUTPUT_DIR / "latest_brief.md").write_text(brief, encoding="utf-8")
    return brief


def parse_years(value: str, default: int = 5, minimum: int = 1) -> int:
    try:
        return min(max(int(clean_text(value) or default), minimum), 10)
    except Exception:
        return default


def build_standalone_simulator(fields: dict[str, str]) -> str:
    ticker = clean_text(fields.get("ticker", ""))
    company = clean_text(fields.get("company", "")) or ticker or "Selected Ticker"
    benchmark = clean_text(fields.get("benchmark", "SPY")) or "SPY"
    peers = clean_text(fields.get("peers", ""))
    years = parse_years(fields.get("years", "5"), 5, 1)
    amount = parse_amount(fields.get("investment_amount", "1000"), 1000.0)
    market = fetch_market_analysis(company, ticker or company, peers, benchmark, years) if (ticker or company) else None
    brief = f"""# Standalone Investment Simulator

## Executive Readout
Ticker / company entered: {ticker or company}. Starting amount: {money(amount, market.target.currency if market else 'USD')}. Benchmark: {benchmark}. Period pulled: {years} years.

This simulator runs without the Pre-Investment Brief. It uses the market engine directly, with Yahoo first, Stooq historical fallback second, and clearly labeled baseline fallback where free endpoints are unavailable.

## Simulator
| Scenario | Annual Rate | 1Y Value | 3Y Value | 5Y Value | 10Y Value |
| --- | --- | --- | --- | --- | --- |
{investment_simulator_rows(market, amount)}

## Market Context
{market_section(market, amount)}

## What This Means
{monthly_allocation_read(market, amount)}

## Next Checks
- Verify latest price, P/E, P/B, market cap, and volume before relying on the model.
- Compare the target against peers and benchmark performance.
- Treat the simulator as educational scenario analysis, not a prediction or personalized investment advice.
"""
    CHAT_STATE["company"] = company
    CHAT_STATE["brief"] = brief
    CHAT_STATE["sources"] = []
    CHAT_STATE["figures"] = []
    (OUTPUT_DIR / "latest_brief.md").write_text(brief, encoding="utf-8")
    return brief


def allocation_candidate_tickers(fields: dict[str, str]) -> tuple[list[str], str]:
    typed = [clean_text(item).upper() for item in re.split(r"[,\s]+", fields.get("tickers", "")) if clean_text(item)]
    if typed:
        return typed[:10], "User-entered ticker list"
    market = clean_text(fields.get("allocation_market", "United States")) or "United States"
    if market == "Global":
        return ["SPY", "QQQ", "VGK", "PG", "KO", "COST", "600519.SS", "000333.SZ"], "Default global opportunity set"
    if market == "Western Markets":
        return ["SPY", "QQQ", "VGK", "PG", "KO", "COST", "WMT", "MSFT"], "Default western-market opportunity set"
    if market in COUNTRY_MARKET_UNIVERSES:
        base = COUNTRY_MARKET_UNIVERSES[market][:8]
        return (["SPY", "QQQ", "VGK"] + base)[:10], f"Default {market} opportunity set"
    return ["SPY", "QQQ", "VGK"], "Default broad-market opportunity set"


def ai_allocation_read(amount: float, ranked: list[tuple[float, MarketAnalysis]], candidate_source: str, risk_style: str) -> str:
    packet = [
        {
            "ticker": market.target.ticker,
            "name": market.target.name,
            "score": score,
            "price": market.target.regular_market_price,
            "pe": market.target.trailing_pe,
            "pb": market.target.price_to_book,
            "one_year_return": period_return(market.target.points, 1),
            "volatility": market.target_metrics.volatility_pct,
            "drawdown": market.target_metrics.max_drawdown_pct,
            "trend": market.target_metrics.trend_label,
        }
        for score, market in ranked[:6]
    ]
    api_key = os.getenv("OPENAI_API_KEY")
    if api_key:
        try:
            from openai import OpenAI  # type: ignore

            prompt = f"""
You are an educational investing assistant. The user asks what the AI would invest in this month with {amount}.
Use only this candidate packet. Do not claim personalized advice. Explain the choice, why the top names were selected, what could go wrong, and what to verify next.
Risk style: {risk_style}
Candidate source: {candidate_source}
Candidates:
{json.dumps(packet, indent=2)}
"""
            client = OpenAI(api_key=api_key)
            response = client.responses.create(model=os.getenv("OPENAI_MODEL", "gpt-4.1-mini"), input=prompt)
            return clean_text(response.output_text)
        except Exception as exc:
            return f"AI wording unavailable ({exc}). The rules-based model selected the highest-scoring mix below using trend, volatility, drawdown, and valuation penalties."
    return "Rules-based AI fallback: the model selected the highest-scoring mix below using trend, volatility, drawdown, and valuation penalties. This is educational and must be verified before any real trade."


def build_standalone_allocation(fields: dict[str, str]) -> str:
    tickers, candidate_source = allocation_candidate_tickers(fields)
    benchmark = clean_text(fields.get("benchmark", "SPY")) or "SPY"
    years = parse_years(fields.get("years", "3"), 3, 1)
    amount = parse_amount(fields.get("monthly_amount", "1000"), 1000.0)
    risk_style = clean_text(fields.get("risk_style", "Balanced")) or "Balanced"
    analyses = []
    for ticker in tickers[:8]:
        market = fetch_market_analysis(ticker, ticker, "", benchmark, years)
        if market:
            analyses.append(market)
    ranked = []
    for market in analyses:
        m = market.target_metrics
        valuation_penalty = 0
        if market.target.trailing_pe and market.target.trailing_pe > 30:
            valuation_penalty += 2
        if market.target.price_to_book and market.target.price_to_book > 6:
            valuation_penalty += 2
        style_adjustment = 0
        if risk_style == "Conservative":
            style_adjustment = -m.volatility_pct * 0.20 + max(m.max_drawdown_pct, -80) * 0.20
        elif risk_style == "Aggressive":
            style_adjustment = (period_return(market.target.points, 1) or 0) * 0.30 + m.cagr_pct * 0.20
        score = m.cagr_pct + (period_return(market.target.points, 1) or 0) * 0.35 - m.volatility_pct * 0.35 + max(m.max_drawdown_pct, -80) * 0.20 - valuation_penalty * 8 + style_adjustment
        ranked.append((score, market))
    ranked.sort(key=lambda item: item[0], reverse=True)
    picks = ranked[:4]
    if picks:
        positive_scores = [max(score, 5.0) for score, _ in picks]
        total_score = sum(positive_scores)
        rows = []
        for weight_score, (score, market) in zip(positive_scores, picks):
            allocation = amount * (weight_score / total_score)
            rows.append(
                f"| {md_cell(market.target.ticker, 80)} | {md_cell(market.target.name or market.target.ticker, 160)} | {money(allocation, market.target.currency)} | {pct(period_return(market.target.points, 1))} | {pct(market.target_metrics.volatility_pct)} | {pct(market.target_metrics.max_drawdown_pct)} | P/E {fmt_ratio(market.target.trailing_pe)}, P/B {fmt_ratio(market.target.price_to_book)} | {md_cell(market.target_metrics.trend_label, 160)} |"
            )
        allocation_table = "\n".join(rows)
        top_read = ai_allocation_read(amount, ranked, candidate_source, risk_style)
    else:
        allocation_table = f"| Cash / watchlist reserve | No usable tickers | {money(amount, 'USD')} | Not available | Not available | Not available | Not available | No market data was available. |"
        top_read = "No usable market data was available, so the educational model keeps the amount in cash/watchlist reserve."
    brief = f"""# What AI Would Invest In With $1,000 This Month

## Executive Readout
Monthly amount modeled: {money(amount, 'USD')}. Candidate source: {candidate_source}. Candidates reviewed: {', '.join(tickers)}. Risk style: {risk_style}. Benchmark: {benchmark}. Period pulled: {years} years.

This runs independently from the Pre-Investment Brief. If you leave tickers blank, the app chooses from a default opportunity set. If your OpenAI key is configured, the readout is written by AI using only the scored market packet; otherwise a rules-based fallback is used.

## $1,000 This Month Allocation
{top_read}

| Ticker | Company | Allocation | 1Y Return | Volatility | Max Drawdown | Valuation | Trend |
| --- | --- | --- | --- | --- | --- | --- | --- |
{allocation_table}

## How To Read This
- Higher allocation means the ticker screened better versus the other entered tickers on this educational model.
- Lower allocation or exclusion usually means weaker trend, higher volatility, bigger drawdown, missing data, or valuation concerns.
- This is not personalized investment advice. Verify current prices, fundamentals, risk, taxes, liquidity, and your own time horizon before any real decision.
"""
    CHAT_STATE["company"] = "Standalone allocation model"
    CHAT_STATE["brief"] = brief
    CHAT_STATE["sources"] = []
    CHAT_STATE["figures"] = []
    (OUTPUT_DIR / "latest_brief.md").write_text(brief, encoding="utf-8")
    return brief


def build_gross_margin_calculator(fields: dict[str, str]) -> str:
    revenue = parse_amount(fields.get("revenue", "0"), 0.0)
    cogs = parse_amount(fields.get("cogs", "0"), 0.0)
    gross_profit_input = parse_amount(fields.get("gross_profit", "0"), 0.0)
    operating_expenses = parse_amount(fields.get("operating_expenses", "0"), 0.0)
    target_margin = parse_float(fields.get("target_margin", "40"), 40.0, -100.0, 100.0)
    currency = clean_text(fields.get("currency", "USD")) or "USD"
    if gross_profit_input and not cogs:
        gross_profit = gross_profit_input
        cogs = max(revenue - gross_profit, 0.0)
    else:
        gross_profit = revenue - cogs
    gross_margin = (gross_profit / revenue * 100) if revenue else None
    operating_income = gross_profit - operating_expenses
    operating_margin = (operating_income / revenue * 100) if revenue else None
    target_gross_profit = revenue * target_margin / 100
    cogs_at_target = revenue - target_gross_profit
    improvement_needed = target_gross_profit - gross_profit
    revenue_at_same_cogs = cogs / (1 - target_margin / 100) if target_margin < 100 else None
    read = []
    if gross_margin is None:
        read.append("Gross margin cannot be calculated until revenue is entered.")
    elif gross_margin < 20:
        read.append("Gross margin is thin; pricing power, input costs, product mix, and scale economics need close review.")
    elif gross_margin < 45:
        read.append("Gross margin is moderate; compare it against direct peers and watch whether operating expenses consume the gross profit.")
    else:
        read.append("Gross margin is strong on the entered numbers, but it still needs verification against filings or financial statements.")
    if operating_expenses and operating_margin is not None:
        read.append(f"After entered operating expenses, operating margin is {fmt_pct_plain(operating_margin)}, so operating leverage should be reviewed.")
    brief = f"""# Gross Margin Calculator

## Executive Readout
This section calculates gross profit, gross margin, operating income, and the gap to a target gross margin using only the numbers you entered.

{' '.join(read)}

## Inputs And Formula
| Item | Value |
| --- | --- |
| Revenue | {money(revenue, currency)} |
| COGS | {money(cogs, currency)} |
| Gross profit input | {money(gross_profit_input, currency)} |
| Operating expenses | {money(operating_expenses, currency)} |
| Target gross margin | {fmt_pct_plain(target_margin)} |

Formula: gross margin = `(revenue - COGS) / revenue`.

## Output
| Metric | Result | What It Means |
| --- | --- | --- |
| Gross profit | {money(gross_profit, currency)} | Money left after direct cost of goods/services. |
| Gross margin | {fmt_pct_plain(gross_margin)} | Pricing power and direct cost efficiency. |
| Operating income after entered opex | {money(operating_income, currency)} | Gross profit less entered operating expenses. |
| Operating margin after entered opex | {fmt_pct_plain(operating_margin)} | Operating profit as a percent of revenue. |
| Gross profit needed at target margin | {money(target_gross_profit, currency)} | Required gross profit at the target margin. |
| COGS allowed at target margin | {money(cogs_at_target, currency)} | Maximum COGS at the target margin. |
| Gross profit improvement needed | {money(improvement_needed, currency)} | Positive means margin must improve to hit target. |
| Revenue needed at same COGS for target margin | {money(revenue_at_same_cogs, currency) if revenue_at_same_cogs is not None else "Not available"} | Revenue required if COGS stays flat. |

## Diligence Questions
- Is the margin based on verified financial statements, management notes, or an estimate?
- Did margin move because of price, volume, mix, freight, labor, commodity costs, promotions, or write-downs?
- Are gross margins comparable to peers after adjusting for accounting differences?
- Is the target margin realistic for this product category and country?
"""
    CHAT_STATE["company"] = "Gross margin calculator"
    CHAT_STATE["brief"] = brief
    CHAT_STATE["sources"] = []
    CHAT_STATE["figures"] = []
    (OUTPUT_DIR / "latest_brief.md").write_text(brief, encoding="utf-8")
    return brief


def build_risk_calculator(fields: dict[str, str]) -> str:
    ticker = clean_text(fields.get("ticker", ""))
    benchmark = clean_text(fields.get("benchmark", "SPY")) or "SPY"
    years = parse_years(fields.get("years", "3"), 3, 1)
    position = parse_amount(fields.get("position_amount", "10000"), 10000.0)
    stop_loss_pct = abs(parse_float(fields.get("stop_loss_pct", "15"), 15.0, 0.0, 95.0))
    debt_cash = parse_float(fields.get("debt_cash_score", "5"), 5.0, 1.0, 10.0)
    margin_quality = parse_float(fields.get("margin_quality_score", "5"), 5.0, 1.0, 10.0)
    source_quality = parse_float(fields.get("source_quality_score", "5"), 5.0, 1.0, 10.0)
    concentration = parse_float(fields.get("concentration_score", "5"), 5.0, 1.0, 10.0)
    regulatory = parse_float(fields.get("regulatory_score", "3"), 3.0, 1.0, 10.0)
    market = fetch_market_analysis(ticker, ticker, "", benchmark, years) if ticker else None
    volatility = market.target_metrics.volatility_pct if market else parse_float(fields.get("volatility_pct", "30"), 30.0, 0.0, 200.0)
    drawdown = abs(market.target_metrics.max_drawdown_pct) if market else abs(parse_float(fields.get("max_drawdown_pct", "35"), 35.0, 0.0, 100.0))
    one_year = period_return(market.target.points, 1) if market else None
    market_risk = min(10.0, volatility / 8 + drawdown / 12)
    data_risk = (11 - source_quality) * 0.55
    fundamental_risk = debt_cash * 0.18 + margin_quality * 0.15 + concentration * 0.16 + regulatory * 0.12
    momentum_risk = 1.0 if one_year is not None and one_year < 0 else 0.0
    score = min(10.0, max(1.0, market_risk * 0.42 + fundamental_risk + data_risk + momentum_risk))
    category = "Low" if score < 4 else "Moderate" if score < 7 else "High"
    stop_loss_value = position * stop_loss_pct / 100
    drawdown_value = position * drawdown / 100
    volatility_value = position * volatility / 100
    brief = f"""# Risk Calculator

## Executive Readout
Risk category: **{category}**. Composite risk score: **{score:.1f} / 10**.

This calculator blends market risk, drawdown, source quality, debt/cash uncertainty, margin quality, customer concentration, and regulatory risk. If a ticker is entered, market risk is pulled from Yahoo/Stooq/baseline market data; otherwise it uses your manual volatility and drawdown inputs.

## Risk Scorecard
| Risk Component | Input / Result | Risk Read |
| --- | --- | --- |
| Ticker | {ticker or "Not entered"} | {'Market data included.' if market else 'Manual market-risk inputs used.'} |
| Annualized volatility | {fmt_pct_plain(volatility)} | Higher volatility means wider outcome range. |
| Max drawdown | {fmt_pct_plain(drawdown)} | Larger drawdown means larger historical downside stress. |
| 1Y return | {fmt_pct_plain(one_year)} | Negative 1Y return adds momentum risk. |
| Debt / cash risk | {debt_cash:.1f} / 10 | Higher means weaker balance-sheet comfort. |
| Margin risk | {margin_quality:.1f} / 10 | Higher means margins are weak, falling, or unknown. |
| Source quality | {source_quality:.1f} / 10 | Higher means stronger source packet; lower increases risk. |
| Customer concentration risk | {concentration:.1f} / 10 | Higher means revenue may depend on fewer customers. |
| Legal / regulatory risk | {regulatory:.1f} / 10 | Higher means more regulatory or legal exposure. |
| Composite risk score | {score:.1f} / 10 | {category} risk. |

## Position Stress
| Stress Case | Loss / Swing On {money(position, 'USD')} |
| --- | --- |
| Stop-loss case at {fmt_pct_plain(stop_loss_pct)} | {money(stop_loss_value, 'USD')} |
| Historical max drawdown case | {money(drawdown_value, 'USD')} |
| One-volatility annual swing | {money(volatility_value, 'USD')} |

## Next Risk Checks
- Verify cash, debt, margins, customer concentration, legal disclosures, and liquidity from filings or audited financials.
- For short ideas, check borrow availability, hard-to-borrow fees, options liquidity, and local short-sale rules.
- For long ideas, define max position size, stop level, time horizon, and what evidence would make you exit.
"""
    CHAT_STATE["company"] = "Risk calculator"
    CHAT_STATE["brief"] = brief
    CHAT_STATE["sources"] = []
    CHAT_STATE["figures"] = []
    (OUTPUT_DIR / "latest_brief.md").write_text(brief, encoding="utf-8")
    return brief


def build_short_collar_calculator(fields: dict[str, str]) -> str:
    ticker = clean_text(fields.get("ticker", ""))
    entry = parse_float(fields.get("entry_price", "100"), 100.0, 0.01)
    shares = parse_float(fields.get("shares", "100"), 100.0, 0.0)
    put_strike = parse_float(fields.get("put_strike", str(entry * 0.9)), entry * 0.9, 0.0)
    put_premium = parse_float(fields.get("put_premium", "2"), 2.0, 0.0)
    call_strike = parse_float(fields.get("call_strike", str(entry * 1.1)), entry * 1.1, 0.0)
    call_premium = parse_float(fields.get("call_premium", "2"), 2.0, 0.0)
    borrow_rate = parse_float(fields.get("borrow_rate", "0"), 0.0, 0.0, 200.0)
    months = parse_float(fields.get("months", "3"), 3.0, 0.1, 120.0)
    borrow_cost = entry * shares * borrow_rate / 100 * months / 12
    net_option_credit = (put_premium - call_premium) * shares

    def pnl(final_price: float) -> float:
        short_stock = (entry - final_price) * shares
        short_put = (put_premium - max(put_strike - final_price, 0.0)) * shares
        long_call = (max(final_price - call_strike, 0.0) - call_premium) * shares
        return short_stock + short_put + long_call - borrow_cost

    test_prices = [entry * 0.7, put_strike, entry, call_strike, entry * 1.3]
    rows = "\n".join(
        f"| {money(price, 'USD')} | {money(pnl(price), 'USD')} | {'Gain' if pnl(price) >= 0 else 'Loss'} |"
        for price in test_prices
    )
    max_gain = (entry - put_strike + put_premium - call_premium) * shares - borrow_cost
    max_loss = (entry - call_strike + put_premium - call_premium) * shares - borrow_cost
    brief = f"""# Short Collar Hedging Calculator

## Executive Readout
This models a **short stock collar**: short stock, short put, and long call. The long call helps cap upside risk on the short; the short put helps finance the hedge but gives up extra profit if the stock falls below the put strike.

Ticker: {ticker or "Not entered"}. Short entry: {money(entry, 'USD')}. Shares: {fmt_number(shares)}. Estimated borrow cost: {money(borrow_cost, 'USD')}.

## Strategy Legs
| Leg | Input | Effect |
| --- | --- | --- |
| Short stock | {fmt_number(shares)} shares at {money(entry, 'USD')} | Profits if price falls; loses if price rises. |
| Short put | Strike {money(put_strike, 'USD')}, premium received {money(put_premium, 'USD')} | Finances hedge; caps downside profit below put strike. |
| Long call | Strike {money(call_strike, 'USD')}, premium paid {money(call_premium, 'USD')} | Caps loss if stock rises sharply. |
| Net option credit / debit | {money(net_option_credit, 'USD')} | Positive means credit; negative means hedge costs premium. |

## Expiration Payoff Table
| Final Stock Price | Net P&L After Options And Borrow | Result |
| --- | --- | --- |
{rows}

## Key Payoff Read
| Metric | Result |
| --- | --- |
| Approx. max gain below put strike | {money(max_gain, 'USD')} |
| Approx. capped loss above call strike | {money(max_loss, 'USD')} |
| Borrow rate used | {fmt_pct_plain(borrow_rate)} annualized |
| Time to expiration | {months:.1f} months |

## Diligence Notes
- Confirm borrow availability and borrow fee before entering a short.
- Check option bid/ask spreads, contract multiplier, assignment risk, and liquidity.
- This hedge reduces tail risk but can also cap upside/downside economics in ways that make the trade unattractive.
"""
    CHAT_STATE["company"] = "Short collar calculator"
    CHAT_STATE["brief"] = brief
    CHAT_STATE["sources"] = []
    CHAT_STATE["figures"] = []
    (OUTPUT_DIR / "latest_brief.md").write_text(brief, encoding="utf-8")
    return brief


def build_forwards_futures_calculator(fields: dict[str, str]) -> str:
    asset = clean_text(fields.get("asset", "")) or "Underlying"
    spot = parse_float(fields.get("spot_price", "100"), 100.0, 0.01)
    contract = parse_float(fields.get("contract_price", "100"), 100.0, 0.01)
    units = parse_float(fields.get("units", "100"), 100.0, 0.0)
    months = parse_float(fields.get("months", "6"), 6.0, 0.1, 240.0)
    risk_free = parse_float(fields.get("risk_free_rate", "5"), 5.0, -50.0, 100.0)
    carry = parse_float(fields.get("carry_rate", "0"), 0.0, -50.0, 100.0)
    income = parse_float(fields.get("income_yield", "0"), 0.0, -50.0, 100.0)
    direction = clean_text(fields.get("direction", "Long")).title()
    margin_pct = parse_float(fields.get("initial_margin", "10"), 10.0, 0.0, 100.0)
    t = months / 12
    theoretical = spot * math.exp(((risk_free + carry - income) / 100) * t)
    basis = contract - spot
    fair_value_gap = contract - theoretical
    multiplier = 1 if direction == "Long" else -1
    notional = contract * units
    initial_margin_value = notional * margin_pct / 100

    def contract_pnl(final_price: float) -> float:
        return (final_price - contract) * units * multiplier

    test_prices = [spot * 0.8, spot * 0.9, spot, spot * 1.1, spot * 1.2]
    rows = "\n".join(
        f"| {money(price, 'USD')} | {money(contract_pnl(price), 'USD')} | {pct((price / spot - 1) * 100)} |"
        for price in test_prices
    )
    if fair_value_gap > spot * 0.02:
        valuation_read = "Contract price is above the cost-of-carry fair value estimate; a long contract is paying a premium versus the model."
    elif fair_value_gap < -spot * 0.02:
        valuation_read = "Contract price is below the cost-of-carry fair value estimate; a long contract is buying at a discount versus the model."
    else:
        valuation_read = "Contract price is close to the cost-of-carry fair value estimate on entered assumptions."
    brief = f"""# Forwards And Futures Calculator

## Executive Readout
Underlying: {asset}. Direction: {direction}. Contract price: {money(contract, 'USD')}. Units: {fmt_number(units)}. Notional: {money(notional, 'USD')}.

{valuation_read}

## Fair Value / Carry Model
| Input / Metric | Value |
| --- | --- |
| Spot price | {money(spot, 'USD')} |
| Contract price | {money(contract, 'USD')} |
| Time to maturity | {months:.1f} months |
| Risk-free rate | {fmt_pct_plain(risk_free)} |
| Storage / financing carry | {fmt_pct_plain(carry)} |
| Income / dividend yield | {fmt_pct_plain(income)} |
| Theoretical forward price | {money(theoretical, 'USD')} |
| Basis, contract minus spot | {money(basis, 'USD')} |
| Contract minus theoretical value | {money(fair_value_gap, 'USD')} |
| Initial margin estimate | {money(initial_margin_value, 'USD')} |

Formula: theoretical forward price = `spot * e^((risk-free + carry - income) * time)`.

## Expiration P&L
| Final Underlying Price | Contract P&L | Spot Move |
| --- | --- | --- |
{rows}

## Forwards Vs Futures Notes
- A forward is private/OTC and concentrates counterparty and settlement risk.
- A futures contract is exchange-traded, marked to market, and requires margin.
- Both can create leverage: small price moves can produce large gains or losses compared with posted margin.
- Confirm contract multiplier, tick size, settlement method, expiry, liquidity, margin rules, and tax treatment before using real capital.
"""
    CHAT_STATE["company"] = "Forwards and futures calculator"
    CHAT_STATE["brief"] = brief
    CHAT_STATE["sources"] = []
    CHAT_STATE["figures"] = []
    (OUTPUT_DIR / "latest_brief.md").write_text(brief, encoding="utf-8")
    return brief


def source_packet_for_chat(sources: list[Source]) -> str:
    chunks = []
    for source in sources:
        chunks.append(f"Source: {source.label}\nURL: {source.url or 'provided input'}\nText: {source.text[:6000]}")
    return "\n\n".join(chunks)[:CHAT_CONTEXT_LIMIT]


def local_chat_answer(question: str, brief: str, sources: object, note: str = "") -> str:
    corpus_parts = [brief]
    if isinstance(sources, list):
        corpus_parts.extend(f"{source.label}: {source.text}" for source in sources if isinstance(source, Source))
    sentences = []
    for sentence in re.split(r"(?<=[.!?])\s+", clean_text(" ".join(corpus_parts))):
        if 45 <= len(sentence) <= 420:
            sentences.append(sentence)
    terms = [term for term in re.findall(r"[A-Za-z0-9$%.-]+", question.lower()) if len(term) > 2]
    scored: list[tuple[int, str]] = []
    for sentence in sentences:
        lowered = sentence.lower()
        score = sum(1 for term in terms if term in lowered)
        if score:
            scored.append((score, sentence))
    scored.sort(key=lambda item: item[0], reverse=True)
    prefix = f"{note} " if note else ""
    if not scored:
        return prefix + "I could not find support for that in the current brief/source packet. Not disclosed based on current source packet."
    top = []
    seen: set[str] = set()
    for _, sentence in scored:
        signature = sentence[:120].lower()
        if signature in seen:
            continue
        seen.add(signature)
        top.append(sentence)
        if len(top) >= 4:
            break
    return prefix + "Local source-search answer: " + " ".join(top)


def extract_tickers_for_fact_check(text: str, limit: int = 8) -> list[str]:
    ignore = {
        "AI", "API", "ARR", "SEC", "CEO", "CFO", "DD", "EBITDA", "EPS", "FCF", "FY", "GAAP", "GDP",
        "IPO", "LLC", "NYSE", "NASDAQ", "PE", "PB", "Q", "Q1", "Q2", "Q3", "Q4", "ROI", "SPY",
        "USD", "US", "USA",
    }
    candidates = re.findall(r"\b[A-Z0-9]{1,6}(?:\.[A-Z]{1,4})?\b", text.upper())
    symbols: list[str] = []
    for candidate in candidates:
        if candidate in ignore or candidate.isdigit():
            continue
        if candidate not in symbols:
            symbols.append(candidate)
        if len(symbols) >= limit:
            break
    return symbols


def yahoo_fact_check_context(question: str, brief: str = "") -> str:
    symbols = extract_tickers_for_fact_check(question, limit=6)
    rows = SCREENER_STATE.get("rows", [])
    if not symbols and isinstance(rows, list):
        for row in rows[:6]:
            ticker = clean_text(str(row.get("Ticker", ""))) if isinstance(row, dict) else ""
            if ticker and ticker not in symbols:
                symbols.append(ticker)
    if not symbols:
        symbols = extract_tickers_for_fact_check(brief[:3000], limit=4)
    if not symbols:
        return ""
    facts = [yahoo_fact_check_symbol(symbol) for symbol in symbols[:8]]
    lines = [
        "Yahoo live fact-check packet. Use this to verify market facts before answering:",
        "| Ticker | Yahoo status | Price | P/E | P/B | Market cap | Avg volume | Exchange |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for fact in facts:
        lines.append(
            f"| {md_cell(fact.get('Ticker'), 40)} | {md_cell(fact.get('Yahoo Fact Check'), 220)} | {md_cell(fmt_number(yahoo_float(fact.get('Price'))), 80)} | {md_cell(fmt_ratio(yahoo_float(fact.get('Trailing PE'))), 80)} | {md_cell(fmt_ratio(yahoo_float(fact.get('Price To Book'))), 80)} | {md_cell(compact_money(yahoo_float(fact.get('Market Cap')), str(fact.get('Currency') or '')), 120)} | {md_cell(fmt_number(yahoo_float(fact.get('Avg Volume'))), 90)} | {md_cell(fact.get('Exchange'), 120)} |"
        )
    return "\n".join(lines)


def answer_chat_question(question: str) -> str:
    question = clean_text(question)
    if not question:
        return "Ask a question about the generated brief, SEC filings, market data, risks, or missing diligence items."

    company = str(CHAT_STATE.get("company") or "the target company")
    brief = str(CHAT_STATE.get("brief") or "")
    sources = CHAT_STATE.get("sources") or []
    if not brief:
        return "Generate a DD brief first, then I can answer questions from that brief and its source packet."

    yahoo_context = yahoo_fact_check_context(question, brief)
    api_key = os.getenv("OPENAI_API_KEY")
    if api_key:
        try:
            from openai import OpenAI  # type: ignore

            prompt = f"""
You are a diligence Q&A assistant for {company}.
Answer the user's question using the generated brief, source packet, and Yahoo live fact-check packet below.
For market facts such as price, P/E, P/B, market cap, volume, returns, volatility, drawdown, benchmark comparison, or whether a number is available, prioritize the Yahoo fact-check packet.
If the answer is not supported, say what is not disclosed based on the current source packet.
Be concise, cite the source label when useful, and call out uncertainty. Do not invent missing Yahoo, SEC, or S&P Global values.

Question:
{question}

Generated brief:
{brief[:18000]}

Yahoo live fact-check:
{yahoo_context or "No tickers were detected for live Yahoo fact-checking."}

Source packet:
{source_packet_for_chat(sources if isinstance(sources, list) else [])}
"""
            client = OpenAI(api_key=api_key)
            response = client.responses.create(model=os.getenv("OPENAI_MODEL", "gpt-4.1-mini"), input=prompt)
            return clean_text(response.output_text)
        except ImportError:
            return local_chat_answer(
                question,
                brief,
                sources,
                "OpenAI is configured, but the Python package is not installed. Run `python3 -m pip install -r requirements.txt` to enable AI chat.",
            )
        except Exception as exc:
            fallback = local_chat_answer(question, brief, sources, f"AI chat failed ({exc}).")
            return fallback + (("\n\n" + yahoo_context) if yahoo_context else "")

    fallback = local_chat_answer(question, brief, sources, "Set OPENAI_API_KEY in your hosting platform environment variables to enable fuller AI chat.")
    return fallback + (("\n\n" + yahoo_context) if yahoo_context else "")


def convert_inline(text: str) -> str:
    escaped = html.escape(text)
    escaped = re.sub(r"\[([^\]]+)\]\((https?://[^)]+)\)", r'<a href="\2" target="_blank" rel="noreferrer">\1</a>', escaped)
    escaped = re.sub(r"\[([^\]]+)\]\((/[^)]+)\)", r'<a href="\2" target="_blank" rel="noreferrer">\1</a>', escaped)
    return escaped


def split_md_row(row: str) -> list[str]:
    text = row.strip().strip("|")
    return [cell.strip() for cell in text.split("|")]


def markdown_table_to_html(rows: list[str]) -> str:
    raw_rows = [split_md_row(row) for row in rows]
    if not raw_rows:
        return ""
    has_header = len(raw_rows) > 1 and all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) for cell in raw_rows[1])
    max_cols = max(len(r) for r in raw_rows)
    parsed = []
    for row in raw_rows:
        if len(row) < max_cols:
            row = row + [""] * (max_cols - len(row))
        parsed.append([convert_inline(cell) for cell in row[:max_cols]])
    html_rows = []
    start = 0
    if has_header:
        html_rows.append("<thead><tr>" + "".join(f"<th>{cell}</th>" for cell in parsed[0]) + "</tr></thead>")
        start = 2
    body = ["<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>" for row in parsed[start:]]
    return '<div class="table-wrap"><table>' + "".join(html_rows) + "<tbody>" + "".join(body) + "</tbody></table></div>"


def render_markdown(md: str) -> str:
    lines = md.splitlines()
    out: list[str] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if not stripped:
            i += 1
            continue
        if stripped.startswith("|"):
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_lines.append(lines[i])
                i += 1
            out.append(markdown_table_to_html(table_lines))
            continue
        if stripped.startswith("# "):
            out.append(f"<h1>{convert_inline(stripped[2:])}</h1>")
        elif stripped.startswith("## "):
            title = stripped[3:]
            anchor = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
            out.append(f'<h2 id="{html.escape(anchor)}">{convert_inline(title)}</h2>')
        elif stripped.startswith("### "):
            title = stripped[4:]
            anchor = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
            out.append(f'<h3 id="{html.escape(anchor)}">{convert_inline(title)}</h3>')
        elif stripped.startswith("!["):
            match = re.match(r"!\[([^\]]*)\]\(([^)]+)\)", stripped)
            if match:
                alt, src = html.escape(match.group(1)), html.escape(match.group(2))
                out.append(f'<img class="chart" src="{src}" alt="{alt}">')
            else:
                out.append(f"<p>{convert_inline(stripped)}</p>")
        elif stripped.startswith("- "):
            items = []
            while i < len(lines) and lines[i].strip().startswith("- "):
                items.append(f"<li>{convert_inline(lines[i].strip()[2:])}</li>")
                i += 1
            out.append("<ul>" + "".join(items) + "</ul>")
            continue
        else:
            out.append(f"<p>{convert_inline(stripped)}</p>")
        i += 1
    return "\n".join(out)


def chat_panel() -> str:
    return """
<section class="chat-panel">
  <div class="chat-head">
    <div>
      <p class="eyebrow">Brief Q&A</p>
      <h2>Ask questions about this DD brief</h2>
    </div>
  </div>
  <div id="chat-log" class="chat-log">
    <div class="chat-message assistant">Ask about revenue, margins, SEC filing support, market performance, red flags, missing diligence, or next steps.</div>
  </div>
  <form id="chat-form" class="chat-form">
    <label class="sr-only" for="chat-question">Question</label>
    <input id="chat-question" name="question" placeholder="Ask a question about the brief..." autocomplete="off">
    <button type="submit">Ask</button>
  </form>
  <script>
    const chatForm = document.getElementById("chat-form");
    const chatInput = document.getElementById("chat-question");
    const chatLog = document.getElementById("chat-log");
    function addChatMessage(role, text) {
      const item = document.createElement("div");
      item.className = "chat-message " + role;
      item.textContent = text;
      chatLog.appendChild(item);
      chatLog.scrollTop = chatLog.scrollHeight;
    }
    chatForm.addEventListener("submit", async (event) => {
      event.preventDefault();
      const question = chatInput.value.trim();
      if (!question) return;
      addChatMessage("user", question);
      chatInput.value = "";
      const thinking = document.createElement("div");
      thinking.className = "chat-message assistant";
      thinking.textContent = "Thinking...";
      chatLog.appendChild(thinking);
      chatLog.scrollTop = chatLog.scrollHeight;
      try {
        const response = await fetch("/chat", {
          method: "POST",
          headers: {"Content-Type": "application/x-www-form-urlencoded"},
          body: new URLSearchParams({question})
        });
        const data = await response.json();
        thinking.textContent = data.answer || "No answer returned.";
      } catch (error) {
        thinking.textContent = "Chat request failed. Check the server terminal and try again.";
      }
      chatLog.scrollTop = chatLog.scrollHeight;
    });
  </script>
</section>
"""


def result_shell(brief: str) -> str:
    sections = [
        ("executive-readout", "Executive"),
        ("products-business-model", "Products"),
        ("major-transactions-corporate-actions", "Transactions"),
        ("financial-snapshot", "Financials"),
        ("market-competitors", "Market"),
        ("country-overvaluation-short-candidate-screener", "Overvalued"),
        ("stocks", "Stocks"),
        ("markets", "Markets"),
        ("simulator", "Simulator"),
        ("1-000-this-month-allocation", "$1K Allocation"),
        ("risk-red-flags", "Risks"),
        ("summary-everything-you-need-to-know", "Summary"),
    ]
    links = "".join(f'<a href="#{anchor}">{label}</a>' for anchor, label in sections)
    return f"""
<main class="result-shell">
  <aside class="result-nav">
    <p class="eyebrow">Brief Map</p>
    <nav>{links}</nav>
    <a class="ghost full" href="/download">Download markdown</a>
  </aside>
  <div class="result-main">
    <section class="brief">{render_markdown(brief)}</section>
    {chat_panel()}
  </div>
</main>
"""


FORM = """
<main class="app-shell">
  <aside class="app-sidebar">
    <div class="brand-mark">DD</div>
    <nav>
      <a href="/preinvestment">Pre-Investment</a>
      <a href="/overvalued">Overvalued</a>
      <a href="/simulator">Simulator</a>
      <a href="/allocate">$1K Month</a>
      <a href="/gross-margin">Gross Margin</a>
      <a href="/risk-calculator">Risk Calculator</a>
      <a href="/short-collar">Short Collar</a>
      <a href="/forwards-futures">Forwards/Futures</a>
      <a href="/collateral-research">Research</a>
      <a href="#deployment">Deployment</a>
    </nav>
  </aside>
  <section class="app-main">
    <header class="app-header">
      <div>
        <p class="eyebrow">Advanced diligence app</p>
        <h1>Investment DD Workbench</h1>
        <p class="subhead">Focused workflows for pre-investment diligence, country valuation screens, simulation, monthly allocation ideas, margin math, risk scoring, hedging, derivatives, and AI Q&A.</p>
      </div>
      <a class="ghost" href="/download">Download latest markdown</a>
    </header>

    <section class="metric-strip" id="market-engine">
      <div><strong>Pre-Investment</strong><span>Company, filings, products, risks, catalysts</span></div>
      <div><strong>Overvalued</strong><span>Select country only; P/E, P/B, liquidity</span></div>
      <div><strong>Simulator</strong><span>1Y, 3Y, 5Y, 10Y outcomes</span></div>
      <div><strong>$1K Month</strong><span>Model allocation from the latest run</span></div>
      <div><strong>Calculators</strong><span>Gross margin, risk, collars, futures</span></div>
      <div><strong>Research</strong><span>Excel ticker/date collateral plunge screen</span></div>
    </section>

    <section class="module-grid">
      <form class="panel builder-panel module-card module-card-wide" id="preinvestment" action="/generate" method="post" enctype="multipart/form-data">
        <div class="panel-head">
          <div>
            <p class="eyebrow">Section 1</p>
            <h2>Pre-Investment Brief</h2>
            <p class="module-copy">Build the main source-backed DD brief from company inputs, public ticker data, SEC quarterly filings when available, uploads, and pasted notes.</p>
          </div>
          <button type="submit">Generate Advanced DD Brief</button>
        </div>
        <div class="grid">
          <label>Company name
            <input name="company" required placeholder="e.g. Apple">
          </label>
          <label>Website
            <input name="website" required placeholder="https://apple.com">
          </label>
          <label>Public stock ticker
            <input name="ticker" placeholder="Optional, e.g. AAPL, 7203.T, TSCO.L, SHOP.TO, BHP.AX">
          </label>
          <label>Peer tickers
            <input name="peers" placeholder="e.g. MSFT, GOOGL, META">
          </label>
          <label>Benchmark ticker
            <input name="benchmark" value="SPY" placeholder="e.g. SPY, QQQ, ^N225">
          </label>
          <label>Market period
            <select name="years">
              <option value="1">1 year</option>
              <option value="3">3 years</option>
              <option value="5" selected>5 years</option>
              <option value="10">10 years</option>
            </select>
          </label>
          <label>Strategy lens
            <select name="strategy">
              <option>Venture & Growth</option>
              <option>Private Equity / Roll-Up</option>
              <option>Public Equity Review</option>
              <option>Distressed / Special Situations</option>
              <option>Real Assets</option>
              <option>Funds & Co-Investments</option>
              <option>Civic Engagement</option>
            </select>
          </label>
          <label>Pitch deck / financials
            <input name="document" type="file" accept=".txt,.md,.csv,.pdf">
          </label>
        </div>
        <label>Paste notes, article text, deck text, SEC filing excerpt, or financials
          <textarea name="notes" rows="9" placeholder="Paste real financial excerpts here. The app will extract numeric claims, classify them, and keep source context attached."></textarea>
        </label>
      </form>

      <form class="panel module-card" id="overvalued" action="/screen" method="post">
        <p class="eyebrow">Section 2</p>
        <h2>Overvalued Country Screener</h2>
        <p class="module-copy">Select a country and the app pulls the supported liquid consumer universe, P/E, P/B, liquidity, return history, SPY comparison, and shortability notes.</p>
        <div class="grid grid-tight">
          <label>Country
            <select name="screener_country">
              <option>United States</option>
              <option>China</option>
              <option>Japan</option>
              <option>United Kingdom</option>
              <option>Canada</option>
              <option>Australia</option>
              <option>Germany</option>
              <option>France</option>
              <option>India</option>
              <option>South Korea</option>
              <option>Hong Kong</option>
            </select>
          </label>
          <label>Screen period
            <select name="years">
              <option value="3" selected>3 years</option>
              <option value="5">5 years</option>
              <option value="10">10 years</option>
            </select>
          </label>
        </div>
        <button type="submit">Run Country Screener</button>
      </form>

      <form class="panel module-card" id="simulator" action="/simulate" method="post">
        <p class="eyebrow">Section 3</p>
        <h2>Simulator</h2>
        <p class="module-copy">Run a standalone investment simulator from just a ticker and amount. No Pre-Investment Brief required.</p>
        <div class="grid grid-tight">
          <label>Ticker
            <input name="ticker" required placeholder="e.g. AAPL, SPY, 600519.SS">
          </label>
          <label>Amount
            <input name="investment_amount" value="1000" placeholder="e.g. 1000">
          </label>
          <label>Benchmark
            <input name="benchmark" value="SPY" placeholder="e.g. SPY, QQQ, ^N225">
          </label>
          <label>Period
            <select name="years">
              <option value="1">1 year</option>
              <option value="3">3 years</option>
              <option value="5" selected>5 years</option>
              <option value="10">10 years</option>
            </select>
          </label>
        </div>
        <label>Optional peer tickers
          <input name="peers" placeholder="e.g. MSFT, GOOGL, QQQ">
        </label>
        <button type="submit">Run Simulator</button>
      </form>

      <form class="panel module-card" id="monthly-allocation" action="/allocate" method="post">
        <p class="eyebrow">Section 4</p>
        <h2>$1,000 This Month</h2>
        <p class="module-copy">Ask what the AI would invest in this month with $1,000. Tickers are optional; leave them blank and the app chooses from a default opportunity set.</p>
        <div class="grid grid-tight">
          <label>Optional tickers
            <input name="tickers" placeholder="Leave blank for AI-picked default set, or enter AAPL, MSFT, SPY">
          </label>
          <label>Monthly amount
            <input name="monthly_amount" value="1000" placeholder="e.g. 1000">
          </label>
          <label>Opportunity set
            <select name="allocation_market">
              <option>United States</option>
              <option>Global</option>
              <option>Western Markets</option>
              <option>China</option>
              <option>Japan</option>
              <option>United Kingdom</option>
              <option>India</option>
            </select>
          </label>
          <label>Risk style
            <select name="risk_style">
              <option>Balanced</option>
              <option>Conservative</option>
              <option>Aggressive</option>
            </select>
          </label>
          <label>Benchmark
            <input name="benchmark" value="SPY" placeholder="e.g. SPY">
          </label>
          <label>Period
            <select name="years">
              <option value="1">1 year</option>
              <option value="3" selected>3 years</option>
              <option value="5">5 years</option>
              <option value="10">10 years</option>
            </select>
          </label>
        </div>
        <button type="submit">Build $1K Allocation</button>
      </form>

      <form class="panel module-card" id="gross-margin" action="/gross-margin" method="post">
        <p class="eyebrow">Section 5</p>
        <h2>Gross Margin Calculator</h2>
        <p class="module-copy">Calculate gross profit, gross margin, operating margin, and the gap to a target margin.</p>
        <div class="grid grid-tight">
          <label>Revenue <input name="revenue" value="1000000" placeholder="e.g. 1000000"></label>
          <label>COGS <input name="cogs" value="600000" placeholder="e.g. 600000"></label>
          <label>Gross profit, optional <input name="gross_profit" placeholder="Use if COGS is unknown"></label>
          <label>Operating expenses <input name="operating_expenses" value="250000" placeholder="e.g. 250000"></label>
          <label>Target gross margin % <input name="target_margin" value="45" placeholder="e.g. 45"></label>
          <label>Currency <input name="currency" value="USD" placeholder="USD, CNY, EUR"></label>
        </div>
        <button type="submit">Calculate Gross Margin</button>
      </form>

      <form class="panel module-card" id="risk-calculator" action="/risk-calculator" method="post">
        <p class="eyebrow">Section 6</p>
        <h2>Risk Calculator</h2>
        <p class="module-copy">Score market, source, balance-sheet, margin, concentration, and regulatory risk.</p>
        <div class="grid grid-tight">
          <label>Ticker, optional <input name="ticker" placeholder="e.g. AAPL, SPY, 600519.SS"></label>
          <label>Position amount <input name="position_amount" value="10000" placeholder="e.g. 10000"></label>
          <label>Benchmark <input name="benchmark" value="SPY"></label>
          <label>Period <select name="years"><option value="1">1 year</option><option value="3" selected>3 years</option><option value="5">5 years</option><option value="10">10 years</option></select></label>
          <label>Manual volatility % <input name="volatility_pct" value="30"></label>
          <label>Manual max drawdown % <input name="max_drawdown_pct" value="35"></label>
          <label>Stop-loss % <input name="stop_loss_pct" value="15"></label>
          <label>Source quality 1-10 <input name="source_quality_score" value="5"></label>
          <label>Debt/cash risk 1-10 <input name="debt_cash_score" value="5"></label>
          <label>Margin risk 1-10 <input name="margin_quality_score" value="5"></label>
          <label>Customer concentration risk 1-10 <input name="concentration_score" value="5"></label>
          <label>Legal/regulatory risk 1-10 <input name="regulatory_score" value="3"></label>
        </div>
        <button type="submit">Calculate Risk</button>
      </form>

      <form class="panel module-card" id="short-collar" action="/short-collar" method="post">
        <p class="eyebrow">Section 7</p>
        <h2>Short Collar Hedge</h2>
        <p class="module-copy">Model a short-stock hedge using short stock, short put, and long call payoff economics.</p>
        <div class="grid grid-tight">
          <label>Ticker, optional <input name="ticker" placeholder="e.g. AAPL"></label>
          <label>Short entry price <input name="entry_price" value="100"></label>
          <label>Shares <input name="shares" value="100"></label>
          <label>Short put strike <input name="put_strike" value="90"></label>
          <label>Put premium received <input name="put_premium" value="2"></label>
          <label>Long call strike <input name="call_strike" value="110"></label>
          <label>Call premium paid <input name="call_premium" value="2"></label>
          <label>Borrow rate % <input name="borrow_rate" value="0"></label>
          <label>Months to expiration <input name="months" value="3"></label>
        </div>
        <button type="submit">Model Short Collar</button>
      </form>

      <form class="panel module-card" id="forwards-futures" action="/forwards-futures" method="post">
        <p class="eyebrow">Section 8</p>
        <h2>Forwards And Futures</h2>
        <p class="module-copy">Calculate theoretical forward value, basis, margin, and long/short contract payoff.</p>
        <div class="grid grid-tight">
          <label>Underlying asset <input name="asset" value="Index / commodity / stock"></label>
          <label>Direction <select name="direction"><option>Long</option><option>Short</option></select></label>
          <label>Spot price <input name="spot_price" value="100"></label>
          <label>Contract price <input name="contract_price" value="102"></label>
          <label>Units / multiplier <input name="units" value="100"></label>
          <label>Months to maturity <input name="months" value="6"></label>
          <label>Risk-free rate % <input name="risk_free_rate" value="5"></label>
          <label>Carry / storage % <input name="carry_rate" value="0"></label>
          <label>Income / dividend yield % <input name="income_yield" value="0"></label>
          <label>Initial margin % <input name="initial_margin" value="10"></label>
        </div>
        <button type="submit">Calculate Contract</button>
      </form>

      <form class="panel module-card module-card-wide" id="collateral-research" action="/collateral-research" method="post" enctype="multipart/form-data">
        <div class="panel-head">
          <div>
            <p class="eyebrow">Section 9</p>
            <h2>Collateral Plunge Research</h2>
            <p class="module-copy">Upload Excel with just ticker and date. Headers are optional; the app fills the company/profile, collateral math, option overlay, and AI research readout.</p>
          </div>
          <button type="submit">Run Research Screen</button>
        </div>
        <div class="grid">
          <label>Excel spreadsheet <input name="document" type="file" accept=".xlsx,.xlsm,.xltx,.xltm" required></label>
          <label>Market hint
            <select name="research_market">
              <option>Global</option><option>United States</option><option>China</option><option>Hong Kong</option><option>Japan</option><option>United Kingdom</option><option>Canada</option><option>Australia</option><option>Germany</option><option>France</option><option>India</option><option>South Korea</option><option>Taiwan</option>
            </select>
          </label>
          <label>Loan amount <input name="loan_amount" value="100000000"></label>
          <label>LTV % <input name="ltv" value="50"></label>
          <label>Loan years <input name="loan_years" value="3"></label>
          <label>Interest rate % <input name="interest_rate" value="5"></label>
          <label>Plunge filter % <input name="drop_threshold" value="30"></label>
          <label>Volatility % <input name="volatility" value="30"></label>
          <label>Buy call strike % of current price <input name="call_strike_pct" value="100"></label>
          <label>Sell put strike % of current price <input name="put_strike_pct" value="60"></label>
        </div>
      </form>
    </section>

    <section class="rules" id="deployment">
      <strong>Deployment note:</strong> for hosted AI chat, set <code>OPENAI_API_KEY</code> in your hosting platform environment variables. The app uses no-key Yahoo Finance public endpoints for global market data where available and adds exchange/regulator disclosure links for international issuers.
    </section>
  </section>
</main>
"""


def nav_link(path: str, label: str, active: str) -> str:
    active_class = ' class="active"' if active == path else ""
    return f'<a href="{path}"{active_class}>{label}</a>'


def app_page(active: str, title: str, subhead: str, content: str) -> str:
    nav = "".join(
        [
            nav_link("/", "Workbench", active),
            nav_link("/preinvestment", "Pre-Investment", active),
            nav_link("/overvalued", "Overvalued", active),
            nav_link("/simulator", "Simulator", active),
            nav_link("/allocate", "$1K Month", active),
            nav_link("/gross-margin", "Gross Margin", active),
            nav_link("/risk-calculator", "Risk Calculator", active),
            nav_link("/short-collar", "Short Collar", active),
            nav_link("/forwards-futures", "Forwards/Futures", active),
            nav_link("/collateral-research", "Research", active),
        ]
    )
    return f"""
<main class="app-shell">
  <aside class="app-sidebar">
    <div class="brand-mark">DD</div>
    <nav>{nav}</nav>
  </aside>
  <section class="app-main">
    <header class="app-header">
      <div>
        <p class="eyebrow">Advanced diligence app</p>
        <h1>{html.escape(title)}</h1>
        <p class="subhead">{html.escape(subhead)}</p>
      </div>
      <a class="ghost" href="/download">Download latest markdown</a>
    </header>
    <section class="single-tool">{content}</section>
  </section>
</main>
"""


PREINVESTMENT_PAGE = app_page(
    "/preinvestment",
    "Pre-Investment Brief",
    "Build the full DD brief from company, ticker, SEC filings, uploads, and source notes.",
    """
<form class="panel builder-panel module-card module-card-wide" id="preinvestment" action="/generate" method="post" enctype="multipart/form-data">
  <div class="panel-head">
    <div>
      <p class="eyebrow">Section 1</p>
      <h2>Pre-Investment Brief</h2>
      <p class="module-copy">Use this when you want the full diligence memo, not just a market screen.</p>
    </div>
    <button type="submit">Generate Advanced DD Brief</button>
  </div>
  <div class="grid">
    <label>Company name <input name="company" required placeholder="e.g. Apple"></label>
    <label>Website <input name="website" required placeholder="https://apple.com"></label>
    <label>Public stock ticker <input name="ticker" placeholder="Optional, e.g. AAPL, 7203.T, TSCO.L"></label>
    <label>Peer tickers <input name="peers" placeholder="e.g. MSFT, GOOGL, META"></label>
    <label>Benchmark ticker <input name="benchmark" value="SPY" placeholder="e.g. SPY, QQQ, ^N225"></label>
    <label>Market period
      <select name="years"><option value="1">1 year</option><option value="3">3 years</option><option value="5" selected>5 years</option><option value="10">10 years</option></select>
    </label>
    <label>Strategy lens
      <select name="strategy"><option>Venture & Growth</option><option>Private Equity / Roll-Up</option><option>Public Equity Review</option><option>Distressed / Special Situations</option><option>Real Assets</option><option>Funds & Co-Investments</option><option>Civic Engagement</option></select>
    </label>
    <label>Pitch deck / financials <input name="document" type="file" accept=".txt,.md,.csv,.pdf"></label>
  </div>
  <label>Paste notes, article text, deck text, SEC filing excerpt, or financials
    <textarea name="notes" rows="10" placeholder="Paste real financial excerpts here."></textarea>
  </label>
</form>
""",
)


OVERVALUED_PAGE = app_page(
    "/overvalued",
    "Overvalued Country Screener",
    "Select a country and run the stock and market overvaluation screen.",
    """
<form class="panel module-card" id="overvalued" action="/screen" method="post">
  <p class="eyebrow">Section 2</p>
  <h2>Overvalued Country Screener</h2>
  <p class="module-copy">Runs stocks and markets as separate sections, with Excel export after generation.</p>
  <div class="grid grid-tight">
    <label>Country
      <select name="screener_country">
        <option>United States</option><option>China</option><option>Japan</option><option>United Kingdom</option><option>Canada</option><option>Australia</option><option>Germany</option><option>France</option><option>India</option><option>South Korea</option><option>Hong Kong</option>
      </select>
    </label>
    <label>Screen period
      <select name="years"><option value="3" selected>3 years</option><option value="5">5 years</option><option value="10">10 years</option></select>
    </label>
  </div>
  <button type="submit">Run Country Screener</button>
</form>
""",
)


SIMULATOR_PAGE = app_page(
    "/simulator",
    "Simulator",
    "Run the investment simulator from a ticker and amount. No pre-investment brief required.",
    """
<form class="panel module-card" id="simulator" action="/simulate" method="post">
  <p class="eyebrow">Section 3</p>
  <h2>Simulator</h2>
  <p class="module-copy">Model bull, base, bear, and drawdown scenarios from the market engine.</p>
  <div class="grid grid-tight">
    <label>Ticker <input name="ticker" required placeholder="e.g. AAPL, SPY, 600519.SS"></label>
    <label>Amount <input name="investment_amount" value="1000" placeholder="e.g. 1000"></label>
    <label>Benchmark <input name="benchmark" value="SPY" placeholder="e.g. SPY, QQQ, ^N225"></label>
    <label>Period
      <select name="years"><option value="1">1 year</option><option value="3">3 years</option><option value="5" selected>5 years</option><option value="10">10 years</option></select>
    </label>
  </div>
  <label>Optional peer tickers <input name="peers" placeholder="e.g. MSFT, GOOGL, QQQ"></label>
  <button type="submit">Run Simulator</button>
</form>
""",
)


ALLOCATE_PAGE = app_page(
    "/allocate",
    "What AI Would Invest In With $1,000 This Month",
    "Let the app choose from a default opportunity set, or enter tickers if you want to compare your own list.",
    """
<form class="panel module-card" id="monthly-allocation" action="/allocate" method="post">
  <p class="eyebrow">Section 4</p>
  <h2>$1,000 This Month</h2>
  <p class="module-copy">Tickers are optional. Leave them blank and the AI/rules engine chooses from the selected opportunity set.</p>
  <div class="grid grid-tight">
    <label>Optional tickers <input name="tickers" placeholder="Leave blank for AI-picked set, or enter AAPL, MSFT, SPY"></label>
    <label>Monthly amount <input name="monthly_amount" value="1000" placeholder="e.g. 1000"></label>
    <label>Opportunity set
      <select name="allocation_market"><option>United States</option><option>Global</option><option>Western Markets</option><option>China</option><option>Japan</option><option>United Kingdom</option><option>India</option></select>
    </label>
    <label>Risk style
      <select name="risk_style"><option>Balanced</option><option>Conservative</option><option>Aggressive</option></select>
    </label>
    <label>Benchmark <input name="benchmark" value="SPY" placeholder="e.g. SPY"></label>
    <label>Period
      <select name="years"><option value="1">1 year</option><option value="3" selected>3 years</option><option value="5">5 years</option><option value="10">10 years</option></select>
    </label>
  </div>
  <button type="submit">Ask AI What To Invest In</button>
</form>
""",
)


GROSS_MARGIN_PAGE = app_page(
    "/gross-margin",
    "Gross Margin Calculator",
    "Calculate gross profit, margin, operating margin, and target-margin gap.",
    """
<form class="panel module-card" id="gross-margin" action="/gross-margin" method="post">
  <p class="eyebrow">Section 5</p>
  <h2>Gross Margin Calculator</h2>
  <p class="module-copy">Use this for unit economics, financial statement review, or target-margin planning.</p>
  <div class="grid grid-tight">
    <label>Revenue <input name="revenue" value="1000000" placeholder="e.g. 1000000"></label>
    <label>COGS <input name="cogs" value="600000" placeholder="e.g. 600000"></label>
    <label>Gross profit, optional <input name="gross_profit" placeholder="Use if COGS is unknown"></label>
    <label>Operating expenses <input name="operating_expenses" value="250000" placeholder="e.g. 250000"></label>
    <label>Target gross margin % <input name="target_margin" value="45" placeholder="e.g. 45"></label>
    <label>Currency <input name="currency" value="USD" placeholder="USD, CNY, EUR"></label>
  </div>
  <button type="submit">Calculate Gross Margin</button>
</form>
""",
)


RISK_CALCULATOR_PAGE = app_page(
    "/risk-calculator",
    "Risk Calculator",
    "Score market, data, financial, concentration, and legal/regulatory risk.",
    """
<form class="panel module-card" id="risk-calculator" action="/risk-calculator" method="post">
  <p class="eyebrow">Section 6</p>
  <h2>Risk Calculator</h2>
  <p class="module-copy">Enter a ticker for market-driven risk, or leave it blank and use manual volatility/drawdown assumptions.</p>
  <div class="grid grid-tight">
    <label>Ticker, optional <input name="ticker" placeholder="e.g. AAPL, SPY, 600519.SS"></label>
    <label>Position amount <input name="position_amount" value="10000" placeholder="e.g. 10000"></label>
    <label>Benchmark <input name="benchmark" value="SPY"></label>
    <label>Period <select name="years"><option value="1">1 year</option><option value="3" selected>3 years</option><option value="5">5 years</option><option value="10">10 years</option></select></label>
    <label>Manual volatility % <input name="volatility_pct" value="30"></label>
    <label>Manual max drawdown % <input name="max_drawdown_pct" value="35"></label>
    <label>Stop-loss % <input name="stop_loss_pct" value="15"></label>
    <label>Source quality 1-10 <input name="source_quality_score" value="5"></label>
    <label>Debt/cash risk 1-10 <input name="debt_cash_score" value="5"></label>
    <label>Margin risk 1-10 <input name="margin_quality_score" value="5"></label>
    <label>Customer concentration risk 1-10 <input name="concentration_score" value="5"></label>
    <label>Legal/regulatory risk 1-10 <input name="regulatory_score" value="3"></label>
  </div>
  <button type="submit">Calculate Risk</button>
</form>
""",
)


SHORT_COLLAR_PAGE = app_page(
    "/short-collar",
    "Short Collar Hedge",
    "Model short-stock hedge payoff using a long call and short put collar.",
    """
<form class="panel module-card" id="short-collar" action="/short-collar" method="post">
  <p class="eyebrow">Section 7</p>
  <h2>Short Collar Hedge</h2>
  <p class="module-copy">Use this to see capped-loss and capped-gain economics for a hedged short setup.</p>
  <div class="grid grid-tight">
    <label>Ticker, optional <input name="ticker" placeholder="e.g. AAPL"></label>
    <label>Short entry price <input name="entry_price" value="100"></label>
    <label>Shares <input name="shares" value="100"></label>
    <label>Short put strike <input name="put_strike" value="90"></label>
    <label>Put premium received <input name="put_premium" value="2"></label>
    <label>Long call strike <input name="call_strike" value="110"></label>
    <label>Call premium paid <input name="call_premium" value="2"></label>
    <label>Borrow rate % <input name="borrow_rate" value="0"></label>
    <label>Months to expiration <input name="months" value="3"></label>
  </div>
  <button type="submit">Model Short Collar</button>
</form>
""",
)


FORWARDS_FUTURES_PAGE = app_page(
    "/forwards-futures",
    "Forwards And Futures",
    "Calculate theoretical forward price, basis, contract P&L, and margin.",
    """
<form class="panel module-card" id="forwards-futures" action="/forwards-futures" method="post">
  <p class="eyebrow">Section 8</p>
  <h2>Forwards And Futures</h2>
  <p class="module-copy">Compare contract price to cost-of-carry fair value and stress long/short payoff at expiry.</p>
  <div class="grid grid-tight">
    <label>Underlying asset <input name="asset" value="Index / commodity / stock"></label>
    <label>Direction <select name="direction"><option>Long</option><option>Short</option></select></label>
    <label>Spot price <input name="spot_price" value="100"></label>
    <label>Contract price <input name="contract_price" value="102"></label>
    <label>Units / multiplier <input name="units" value="100"></label>
    <label>Months to maturity <input name="months" value="6"></label>
    <label>Risk-free rate % <input name="risk_free_rate" value="5"></label>
    <label>Carry / storage % <input name="carry_rate" value="0"></label>
    <label>Income / dividend yield % <input name="income_yield" value="0"></label>
    <label>Initial margin % <input name="initial_margin" value="10"></label>
  </div>
  <button type="submit">Calculate Contract</button>
</form>
""",
)


COLLATERAL_RESEARCH_PAGE = app_page(
    "/collateral-research",
    "Collateral Plunge Research",
    "Upload Excel ticker/date rows and screen pledged-share collateral drawdowns.",
    """
<form class="panel builder-panel module-card module-card-wide" id="collateral-research" action="/collateral-research" method="post" enctype="multipart/form-data">
  <div class="panel-head">
    <div>
      <p class="eyebrow">Section 9</p>
      <h2>Collateral Plunge Research</h2>
      <p class="module-copy">Upload an Excel file with just ticker and date. Headers are optional; optional company/name column is supported.</p>
    </div>
    <button type="submit">Run Research Screen</button>
  </div>
  <div class="grid">
    <label>Excel spreadsheet <input name="document" type="file" accept=".xlsx,.xlsm,.xltx,.xltm" required></label>
    <label>Market hint
      <select name="research_market">
        <option>Global</option><option>United States</option><option>China</option><option>Hong Kong</option><option>Japan</option><option>United Kingdom</option><option>Canada</option><option>Australia</option><option>Germany</option><option>France</option><option>India</option><option>South Korea</option><option>Taiwan</option>
      </select>
    </label>
    <label>Loan amount <input name="loan_amount" value="100000000"></label>
    <label>LTV % <input name="ltv" value="50"></label>
    <label>Loan years <input name="loan_years" value="3"></label>
    <label>Interest rate % <input name="interest_rate" value="5"></label>
    <label>Plunge filter % <input name="drop_threshold" value="30"></label>
    <label>Volatility % <input name="volatility" value="30"></label>
    <label>Buy call strike % of current price <input name="call_strike_pct" value="100"></label>
    <label>Sell put strike % of current price <input name="put_strike_pct" value="60"></label>
  </div>
  <div class="mini-list">
    <span>Step 1: infer ticker/date columns and find the median price in the week around the spreadsheet date.</span>
    <span>Step 2: mark the pledged shares to today's pulled price and flag big drawdowns.</span>
    <span>Step 3: estimate a call/put overlay using 30% default volatility and a 60% put strike.</span>
  </div>
</form>
""",
)


def page_for_path(path: str) -> str:
    if path == "/preinvestment":
        return PREINVESTMENT_PAGE
    if path == "/overvalued":
        return OVERVALUED_PAGE
    if path == "/simulator":
        return SIMULATOR_PAGE
    if path == "/allocate":
        return ALLOCATE_PAGE
    if path == "/gross-margin":
        return GROSS_MARGIN_PAGE
    if path == "/risk-calculator":
        return RISK_CALCULATOR_PAGE
    if path == "/short-collar":
        return SHORT_COLLAR_PAGE
    if path == "/forwards-futures":
        return FORWARDS_FUTURES_PAGE
    if path == "/collateral-research":
        return COLLATERAL_RESEARCH_PAGE
    return FORM


CSS = """
:root {
  color-scheme: light;
  --ink: #16211f;
  --muted: #5e6c68;
  --line: #d8dfdc;
  --paper: #f7faf8;
  --panel: #ffffff;
  --accent: #146c5f;
  --accent-2: #8a4a2f;
  --soft: #eef5f2;
  --nav: #102421;
  --warning: #8a4a2f;
}
* { box-sizing: border-box; }
html { scroll-behavior: smooth; }
body {
  margin: 0;
  background: #f3f7f5;
  color: var(--ink);
  font: 15px/1.5 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
}
.app-shell { min-height: 100vh; display: grid; grid-template-columns: 250px minmax(0, 1fr); }
.app-sidebar {
  background: var(--nav);
  color: #fff;
  min-height: 100vh;
  padding: 24px 18px;
  position: sticky;
  top: 0;
  align-self: start;
}
.brand-mark {
  align-items: center;
  background: #e8f4ef;
  border-radius: 8px;
  color: var(--accent);
  display: inline-flex;
  font-weight: 900;
  height: 46px;
  justify-content: center;
  margin-bottom: 24px;
  width: 46px;
}
.app-sidebar nav { display: grid; gap: 8px; }
.app-sidebar a {
  border-radius: 8px;
  color: rgba(255,255,255,.82);
  font-weight: 800;
  padding: 10px 12px;
  text-decoration: none;
}
.app-sidebar a:hover { background: rgba(255,255,255,.10); color: #fff; }
.app-sidebar a.active { background: rgba(255,255,255,.14); color: #fff; }
.app-main { min-width: 0; padding: 28px; }
.app-header {
  align-items: end;
  display: flex;
  gap: 24px;
  justify-content: space-between;
  margin: 0 auto 18px;
  max-width: 1260px;
}
.eyebrow { color: var(--accent-2); font-size: 13px; font-weight: 800; margin: 0 0 6px; text-transform: uppercase; letter-spacing: .04em; }
h1 { font-size: 38px; line-height: 1.05; margin: 0; letter-spacing: -0.03em; }
.subhead { color: var(--muted); margin: 8px 0 0; }
h2 { font-size: 22px; margin-top: 34px; padding-top: 22px; border-top: 1px solid var(--line); }
h3 { font-size: 17px; margin-top: 22px; }
.panel, .brief, .rules, .chat-panel, .metric-strip, .feature-grid > div {
  background: rgba(255,255,255,.94);
  border: 1px solid var(--line);
  border-radius: 8px;
  box-shadow: 0 16px 34px rgba(32, 46, 42, 0.07);
}
.panel, .metric-strip, .feature-grid, .rules { margin-left: auto; margin-right: auto; max-width: 1260px; }
.panel { padding: 22px; }
.module-grid {
  display: grid;
  gap: 18px;
  grid-template-columns: repeat(2, minmax(0, 1fr));
  margin: 0 auto;
  max-width: 1260px;
}
.single-tool { margin: 0 auto; max-width: 1260px; }
.single-tool .panel { margin: 0; max-width: none; }
.module-card { margin: 0; max-width: none; }
.module-card-wide { grid-column: 1 / -1; }
.module-card h2 { border: 0; margin: 0; padding: 0; }
.module-copy { color: var(--muted); margin: 6px 0 0; max-width: 820px; }
.panel-head { align-items: end; display: flex; gap: 16px; justify-content: space-between; margin-bottom: 16px; }
.panel-head h2 { border: 0; margin: 0; padding: 0; }
.grid { display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 16px; }
.grid-tight { grid-template-columns: repeat(2, minmax(0, 1fr)); margin-top: 14px; }
.mini-list {
  border-top: 1px solid var(--line);
  display: grid;
  gap: 8px;
  margin-top: 16px;
  padding-top: 14px;
}
.mini-list span {
  color: var(--muted);
  font-weight: 700;
}
label { display: grid; gap: 7px; color: var(--muted); font-weight: 700; }
input, select, textarea {
  width: 100%;
  border: 1px solid var(--line);
  border-radius: 8px;
  padding: 12px 13px;
  color: var(--ink);
  background: #fff;
  font: inherit;
  outline: none;
}
input:focus, select:focus, textarea:focus { border-color: var(--accent); box-shadow: 0 0 0 3px rgba(20,108,95,.10); }
textarea { resize: vertical; margin-top: 16px; }
button, .ghost {
  appearance: none;
  border: 1px solid var(--accent);
  border-radius: 8px;
  background: var(--accent);
  color: #fff;
  cursor: pointer;
  display: inline-flex;
  align-items: center;
  min-height: 44px;
  padding: 0 18px;
  margin-top: 16px;
  font-weight: 800;
  text-decoration: none;
}
button:hover { filter: brightness(.96); }
.ghost { background: transparent; color: var(--accent); margin: 0; }
.ghost.full { justify-content: center; margin-top: 16px; width: 100%; }
.metric-strip {
  display: grid;
  gap: 0;
  grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
  margin-bottom: 18px;
  overflow: hidden;
}
.metric-strip > div { border-right: 1px solid var(--line); padding: 16px; }
.metric-strip > div:last-child { border-right: 0; }
.metric-strip strong, .feature-grid strong { display: block; font-size: 15px; margin-bottom: 4px; }
.metric-strip span, .feature-grid span { color: var(--muted); display: block; font-size: 13px; }
.feature-grid { display: grid; gap: 16px; grid-template-columns: repeat(3, minmax(0, 1fr)); margin-top: 18px; }
.feature-grid > div { padding: 16px; }
.rules { color: var(--muted); margin-top: 18px; padding: 15px 18px; }
.result-shell {
  display: grid;
  gap: 18px;
  grid-template-columns: 220px minmax(0, 1fr);
  margin: 22px auto 52px;
  max-width: 1480px;
  padding: 0 22px;
}
.result-nav {
  align-self: start;
  background: #fff;
  border: 1px solid var(--line);
  border-radius: 8px;
  box-shadow: 0 16px 34px rgba(32, 46, 42, 0.07);
  padding: 16px;
  position: sticky;
  top: 18px;
}
.result-nav nav { display: grid; gap: 6px; }
.result-nav a {
  border-radius: 8px;
  color: var(--ink);
  font-weight: 800;
  padding: 8px 10px;
  text-decoration: none;
}
.result-nav a:hover { background: var(--soft); color: var(--accent); }
.result-main { min-width: 0; }
.brief { margin: 0 0 24px; padding: 28px; overflow-x: auto; }
.brief h1:first-child { border-bottom: 1px solid var(--line); padding-bottom: 18px; }
.chat-panel { margin: 22px 0 52px; padding: 22px; }
.chat-head { display: flex; justify-content: space-between; align-items: end; gap: 16px; margin-bottom: 14px; }
.chat-head h2 { border: 0; margin: 0; padding: 0; }
.chat-log {
  border: 1px solid var(--line);
  border-radius: 8px;
  background: #fbfdfc;
  min-height: 160px;
  max-height: 380px;
  overflow-y: auto;
  padding: 14px;
}
.chat-message {
  border: 1px solid var(--line);
  border-radius: 8px;
  margin: 0 0 10px;
  max-width: 86%;
  padding: 10px 12px;
  white-space: pre-wrap;
}
.chat-message.assistant { background: #fff; color: var(--ink); }
.chat-message.user { background: var(--accent); border-color: var(--accent); color: #fff; margin-left: auto; }
.chat-form { display: grid; grid-template-columns: 1fr auto; gap: 10px; align-items: end; margin-top: 12px; }
.chat-form input { min-height: 44px; }
.chat-form button { margin-top: 0; }
.sr-only { position: absolute; left: -10000px; width: 1px; height: 1px; overflow: hidden; }
.table-wrap { width: 100%; overflow-x: auto; margin: 12px 0 22px; border: 1px solid var(--line); border-radius: 8px; }
.brief table { border-collapse: collapse; width: 100%; min-width: 780px; font-size: 14px; background: #fff; }
.brief th { background: var(--soft); color: #233330; text-align: left; font-weight: 900; }
.brief th, .brief td { border-bottom: 1px solid var(--line); border-right: 1px solid var(--line); padding: 10px; vertical-align: top; }
.brief tr:last-child td { border-bottom: 0; }
.brief li { margin-bottom: 6px; }
.brief p { margin: 10px 0; }
.chart { display: block; max-width: 100%; height: auto; border: 1px solid var(--line); border-radius: 8px; margin: 16px 0 22px; background: white; box-shadow: 0 10px 30px rgba(32,46,42,.05); }
a { color: var(--accent); }
code { background: var(--soft); border-radius: 6px; padding: 2px 5px; }
@media (max-width: 780px) {
  .app-shell { grid-template-columns: 1fr; }
  .app-sidebar { min-height: auto; position: static; }
  .app-sidebar nav { grid-template-columns: repeat(2, minmax(0, 1fr)); }
  .app-main { padding: 20px 14px; }
  .app-header, .panel-head { align-items: start; flex-direction: column; }
  .metric-strip, .feature-grid, .result-shell, .module-grid, .grid-tight { grid-template-columns: 1fr; }
  .result-nav { position: static; }
  .grid { grid-template-columns: 1fr; }
  .chat-form { grid-template-columns: 1fr; }
  .chat-message { max-width: 100%; }
  h1 { font-size: 29px; }
}
"""


class Handler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        path = urllib.parse.urlparse(self.path).path
        if path == "/static/styles.css":
            self.send_response(200)
            self.send_header("Content-Type", "text/css")
            self.end_headers()
            self.wfile.write(CSS.encode("utf-8"))
            return
        if path.startswith("/outputs/"):
            name = Path(urllib.parse.unquote(path.split("/outputs/", 1)[1])).name
            file_path = OUTPUT_DIR / name
            if not file_path.exists():
                self.send_error(404, "Output file not found")
                return
            content_type = "image/svg+xml" if file_path.suffix.lower() == ".svg" else "application/octet-stream"
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.end_headers()
            self.wfile.write(file_path.read_bytes())
            return
        if path == "/download":
            path = OUTPUT_DIR / "latest_brief.md"
            if not path.exists():
                self.send_error(404, "No brief generated yet")
                return
            self.send_response(200)
            self.send_header("Content-Type", "text/markdown; charset=utf-8")
            self.send_header("Content-Disposition", 'attachment; filename="advanced_dd_brief.md"')
            self.end_headers()
            self.wfile.write(path.read_bytes())
            return
        if path == "/download_screener.xlsx":
            rows = SCREENER_STATE.get("rows", [])
            country = str(SCREENER_STATE.get("country") or "country_screener")
            if not isinstance(rows, list):
                rows = []
            payload = screener_workbook_bytes(rows, country)
            is_xlsx = payload[:2] == b"PK"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if is_xlsx else "text/csv; charset=utf-8")
            self.send_header("Content-Disposition", 'attachment; filename="country_overvaluation_screener.xlsx"' if is_xlsx else 'attachment; filename="country_overvaluation_screener.csv"')
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return
        if path == "/download_research.xlsx":
            rows = RESEARCH_STATE.get("rows", [])
            assumptions = RESEARCH_STATE.get("assumptions", {})
            parse_note = str(RESEARCH_STATE.get("parse_note") or "")
            missing = RESEARCH_STATE.get("missing", [])
            if not isinstance(rows, list):
                rows = []
            if not isinstance(assumptions, dict):
                assumptions = {}
            if not isinstance(missing, list):
                missing = []
            payload = research_workbook_bytes(rows, assumptions, parse_note, [str(item) for item in missing])
            is_xlsx = payload[:2] == b"PK"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if is_xlsx else "text/csv; charset=utf-8")
            self.send_header("Content-Disposition", 'attachment; filename="collateral_plunge_research.xlsx"' if is_xlsx else 'attachment; filename="collateral_plunge_research.csv"')
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.end_headers()
        self.wfile.write(html_page(page_for_path(path)))

    def do_POST(self) -> None:
        if self.path == "/chat":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length).decode("utf-8", errors="replace")
            fields = urllib.parse.parse_qs(body)
            question = fields.get("question", [""])[0]
            try:
                answer = answer_chat_question(question)
                payload = json.dumps({"answer": answer}).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            except Exception as exc:
                payload = json.dumps({"answer": f"Chat failed: {exc}"}).encode("utf-8")
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            return

        if self.path == "/screen":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length).decode("utf-8", errors="replace")
            fields = urllib.parse.parse_qs(body)
            country = clean_text(fields.get("screener_country", [""])[0])
            years_raw = clean_text(fields.get("years", ["3"])[0]) or "3"
            try:
                years = min(max(int(years_raw), 3), 10)
            except Exception:
                years = 3
            try:
                rows = run_country_screener(country, years)
                market_rows = run_country_market_screen(country, years)
                screen_summary = screener_ai_summary(country, rows, years)
                brief = f"""# Overvalued Country Screener

## Executive Readout
Country selected: {country or "Not selected"}.

This standalone screen looks for liquid consumer and consumables stocks that may be overvalued based on public Yahoo Finance fields. It emphasizes P/E, P/B, liquidity, volatility, drawdown, and comparison to SPY as a western-market reference point.

Fact-check process: the app first builds the screen from the selected country universe, then verifies each ticker against Yahoo Finance quote/search fields. If OpenAI chat is configured, the screen summary is written by AI using only those Yahoo-checked rows; otherwise the app creates a rules-based summary.

Screen analysis: {screen_summary}

{screener_section(country, rows, market_rows, years)}

## Summary / Everything You Need To Know
The screen is useful for first-pass idea generation only. Names with high P/E, high P/B, high volatility, large drawdowns, or weak performance versus SPY deserve deeper review. Rows where Yahoo cannot verify P/E or P/B should not be treated as overvalued until you manually confirm fundamentals. Before making any short or investment decision, verify filings, broker borrow availability, local exchange rules, ADR availability, currency exposure, and licensed fundamentals from providers such as S&P Global if you have access.
"""
                CHAT_STATE["company"] = f"{country} overvaluation screener"
                CHAT_STATE["brief"] = brief
                CHAT_STATE["sources"] = []
                CHAT_STATE["figures"] = []
                (OUTPUT_DIR / "latest_brief.md").write_text(brief, encoding="utf-8")
                body_html = OVERVALUED_PAGE + result_shell(brief)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html_page(body_html))
            except Exception as exc:
                self.send_error(500, f"Country screen failed: {exc}")
            return

        if self.path == "/simulate":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length).decode("utf-8", errors="replace")
            parsed = urllib.parse.parse_qs(body)
            fields = {key: values[0] if values else "" for key, values in parsed.items()}
            try:
                brief = build_standalone_simulator(fields)
                body_html = SIMULATOR_PAGE + result_shell(brief)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html_page(body_html))
            except Exception as exc:
                self.send_error(500, f"Simulator failed: {exc}")
            return

        if self.path == "/allocate":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length).decode("utf-8", errors="replace")
            parsed = urllib.parse.parse_qs(body)
            fields = {key: values[0] if values else "" for key, values in parsed.items()}
            try:
                brief = build_standalone_allocation(fields)
                body_html = ALLOCATE_PAGE + result_shell(brief)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html_page(body_html))
            except Exception as exc:
                self.send_error(500, f"Allocation model failed: {exc}")
            return

        if self.path == "/gross-margin":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length).decode("utf-8", errors="replace")
            parsed = urllib.parse.parse_qs(body)
            fields = {key: values[0] if values else "" for key, values in parsed.items()}
            try:
                brief = build_gross_margin_calculator(fields)
                body_html = GROSS_MARGIN_PAGE + result_shell(brief)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html_page(body_html))
            except Exception as exc:
                self.send_error(500, f"Gross margin calculator failed: {exc}")
            return

        if self.path == "/risk-calculator":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length).decode("utf-8", errors="replace")
            parsed = urllib.parse.parse_qs(body)
            fields = {key: values[0] if values else "" for key, values in parsed.items()}
            try:
                brief = build_risk_calculator(fields)
                body_html = RISK_CALCULATOR_PAGE + result_shell(brief)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html_page(body_html))
            except Exception as exc:
                self.send_error(500, f"Risk calculator failed: {exc}")
            return

        if self.path == "/short-collar":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length).decode("utf-8", errors="replace")
            parsed = urllib.parse.parse_qs(body)
            fields = {key: values[0] if values else "" for key, values in parsed.items()}
            try:
                brief = build_short_collar_calculator(fields)
                body_html = SHORT_COLLAR_PAGE + result_shell(brief)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html_page(body_html))
            except Exception as exc:
                self.send_error(500, f"Short collar calculator failed: {exc}")
            return

        if self.path == "/forwards-futures":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length).decode("utf-8", errors="replace")
            parsed = urllib.parse.parse_qs(body)
            fields = {key: values[0] if values else "" for key, values in parsed.items()}
            try:
                brief = build_forwards_futures_calculator(fields)
                body_html = FORWARDS_FUTURES_PAGE + result_shell(brief)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html_page(body_html))
            except Exception as exc:
                self.send_error(500, f"Forwards/futures calculator failed: {exc}")
            return

        if self.path == "/collateral-research":
            length = int(self.headers.get("Content-Length", "0"))
            content_type = self.headers.get("Content-Type", "")
            body = self.rfile.read(length)
            fields, files = parse_multipart(content_type, body)
            try:
                brief = build_collateral_plunge_research(fields, files)
                body_html = COLLATERAL_RESEARCH_PAGE + result_shell(brief)
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html_page(body_html))
            except Exception as exc:
                self.send_error(500, f"Collateral plunge research failed: {exc}")
            return

        if self.path != "/generate":
            self.send_error(404)
            return
        length = int(self.headers.get("Content-Length", "0"))
        content_type = self.headers.get("Content-Type", "")
        body = self.rfile.read(length)
        fields, files = parse_multipart(content_type, body)
        try:
            brief = build_brief(fields, files)
            body_html = PREINVESTMENT_PAGE + result_shell(brief)
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(html_page(body_html))
        except Exception as exc:
            self.send_error(500, f"Brief generation failed: {exc}")


def main() -> None:
    port = int(os.getenv("PORT", "8501"))
    host = os.getenv("HOST", "127.0.0.1")
    server = ThreadingHTTPServer((host, port), Handler)
    print("Advanced DD Brief Generator running locally")
    print(f"Open this on your computer only: http://{host}:{port}")
    print("Press Control+C to stop.")
    server.serve_forever()


if __name__ == "__main__":
    main()
